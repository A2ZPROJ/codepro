# -*- coding: utf-8 -*-
"""Converte um orçamento nosso de EEE (layout A2) para o LAYOUT/CORES do modelo RCE.
v2: mantém a HIERARQUIA (área -> grupo -> subgrupo -> itens -> subtotal por área),
copia a Memória de Cálculo COM FORMATAÇÃO (tabelas), e insere as LOGOS (do André)."""
import openpyxl, re, sys, os
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, get_column_interval
from openpyxl.drawing.image import Image as XLImage

SRC = sys.argv[1] if len(sys.argv)>1 else "NOSSO_SBA5.xlsx"
OUT = sys.argv[2] if len(sys.argv)>2 else "ORC_EEE_FORMATO-RCE_v2.xlsx"
_HERE = os.path.dirname(os.path.abspath(__file__))
LOGO1 = os.path.join(_HERE, "rce_logos", "image1.png")
LOGO2 = os.path.join(_HERE, "rce_logos", "image2.png")

C_TITULO="303030"; C_VERM="A11312"; C_AREA="3A3F47"; C_GRP="6B7280"; C_SUBG="D9DCE1"
C_SUB="E5E7EB"; C_ZEBRA="F5F6F8"; C_AMAR2="FFE08A"; BRANCO="FFFFFF"; FN='Calibri'
def F(sz=9.5,b=False,color=None,it=False): return Font(name=FN,size=sz,bold=b,color=color,italic=it)
def FILL(c): return PatternFill('solid',fgColor=c)
thin=Side(style='thin',color='D0D0D0'); BD=Border(thin,thin,thin,thin)
AL_L=Alignment('left','center',wrap_text=True); AL_C=Alignment('center','center',wrap_text=True)
AL_R=Alignment('right','center'); MONEY='#,##0.00'; PCT='0.00%'

wb=openpyxl.load_workbook(SRC, data_only=True)
src=[s for s in wb.worksheets if 'AMENTO SINT' in s.title.upper()][0]
sb=str(src.cell(6,11).value or '').strip() or 'SB-A5'
cidade=''
if 'Memorial de Cálculo' in wb.sheetnames:
    t1=str(wb['Memorial de Cálculo'].cell(1,1).value or '')
    parts=re.split(r'\s[—–-]\s', t1)   # "MEMORIAL — EEE SB-A5 — CIDADE"
    if len(parts)>=3: cidade=parts[-1].strip()
bdi1=bdi2=None
if 'BDI' in wb.sheetnames:
    b=wb['BDI']
    for r in range(1,60):
        v=b.cell(r,3).value
        if isinstance(v,(int,float)) and 0.1<v<0.4:
            if bdi1 is None: bdi1=v
            elif bdi2 is None and abs(v-bdi1)>1e-6: bdi2=v
bdi1=bdi1 or 0.2449; bdi2=bdi2 or 0.1299
def num(v):
    try: return float(v)
    except: return None

# ---- percorre mantendo a HIERARQUIA: lista de linhas com 'kind' ----
# kind: 'area' | 'grupo' | 'subgrupo' | 'item'
rows=[]
for r in range(14, src.max_row+1):
    cod=src.cell(r,2).value; e=src.cell(r,5).value; g=src.cell(r,7).value
    h=src.cell(r,8).value; i=num(src.cell(r,9).value); un=src.cell(r,10).value; k=num(src.cell(r,11).value)
    desc=str(e).strip() if e else ''
    if e and 'VALOR TOTAL' in desc.upper(): break
    if g in ('Mat','MO'):
        rows.append(dict(kind='item', cod=str(cod or '').strip(),
                         orig=(str(h).strip() if h else 'SANEPAR'), desc=desc,
                         un=str(un or '').strip(), q=i, ku=k,
                         tipo=('MATERIAL' if g=='Mat' else 'OBRA')))
    elif e:
        cods=str(cod).strip() if cod else ''
        if not cods: kind='area'
        else: kind='grupo' if ('.' not in cods) else 'subgrupo'
        rows.append(dict(kind=kind, cod=cods, desc=desc))

# --- realoca "PAVIMENTAÇÃO ASFÁLTICA" (bloco extra no fim) p/ DENTRO da área POÇO ---
# (Daniel: pavimentação não é grupo separado, faz parte dos grupos existentes)
pav=[]; rows2=[]; grab=False
for row in rows:
    if row['kind']=='area' and 'ASF' in row['desc'].upper():
        grab=True; continue                     # descarta o cabeçalho de área isolado
    if grab and row['kind']=='area': grab=False  # acabou o bloco (nova área)
    (pav if grab else rows2).append(row)
if pav:
    ins=len(rows2); seen_poco=False
    for idx,row in enumerate(rows2):
        if row['kind']=='area':
            d=row['desc'].upper()
            if d.startswith('POÇO') and 'VISITA' not in d: seen_poco=True; continue
            if seen_poco: ins=idx; break         # 1ª área após POÇO -> insere antes dela
    rows2[ins:ins]=[dict(kind='subgrupo',cod='',desc='PAVIMENTAÇÃO ASFÁLTICA (recomposição)')]+[
        x for x in pav if x['kind']=='item']
    rows=rows2
else:
    rows=rows2

# --- LIMPEZA DE OBRA: sempre GRUPO SEPARADO e por ÚLTIMO (Daniel) ---
limp=[]; rows3=[]; grabl=False
for row in rows:
    if row['kind'] in ('grupo','subgrupo') and 'LIMPEZA' in row['desc'].upper() and 'OBRA' in row['desc'].upper():
        grabl=True; continue                          # remove o cabeçalho da posição atual
    if grabl and row['kind'] in ('area','grupo','subgrupo'): grabl=False
    (limp if grabl else rows3).append(row)
if limp:
    rows3 += [dict(kind='area',cod='',desc='LIMPEZA DE OBRA')] + [x for x in limp if x['kind']=='item']
rows=rows3

# ---- monta wb formato RCE ----
out=openpyxl.Workbook()
def put_logos(ws):
    # bloco mesclado no topo p/ as logos (como o modelo do André): linhas 1-2 maiores
    ws.row_dimensions[1].height=42; ws.row_dimensions[2].height=42
    try:
        ws.merge_cells('A1:A2')   # logo quadrada (2S)
        ws.merge_cells('C1:E2')   # logo retangular (ACCIONA/SANEPAR)
    except Exception: pass
    for c in ('A1','B1','A2','B2','C1','D1','E1','C2','D2','E2'):
        try: ws[c].fill=FILL(BRANCO)
        except Exception: pass
    try:
        if os.path.exists(LOGO1):
            im=XLImage(LOGO1); im.width=78; im.height=78; ws.add_image(im,'A1')
        if os.path.exists(LOGO2):
            im2=XLImage(LOGO2); im2.width=180; im2.height=81; ws.add_image(im2,'C1')
    except Exception as ex: print('logo:', ex)

# ===== ORÇAMENTO =====
wo=out.active; wo.title='Orçamento'
COLS=[('ITEM',5),('CÓDIGO',12),('ORIGEM',20),('DESCRIÇÃO',54),('UN',7),('QUANT.',12),
      ('CUSTO UNIT (R$)',14),('CUSTO DIRETO (R$)',16),('BDI (%)',9),('PREÇO C/ BDI (R$)',16),('TOTAL (R$)',16),('TIPO',13)]
for idx,(t,w) in enumerate(COLS,1): wo.column_dimensions[get_column_letter(idx)].width=w
def band(ws,r,c1,c2,fill,font):
    for c in range(c1,c2+1):
        cc=ws.cell(r,c); cc.fill=FILL(fill); cc.font=font; cc.border=BD
wo.merge_cells('A1:L2'); wo['A1']='PLANILHA ORÇAMENTÁRIA — ESTAÇÃO ELEVATÓRIA DE ESGOTO (EEE)'
band(wo,1,1,12,C_TITULO,F(15,True,BRANCO)); wo['A1'].alignment=Alignment('center','center'); wo.row_dimensions[1].height=22; wo.row_dimensions[2].height=14
# SB e Município em linhas abaixo do título — label+valor juntos em célula LARGA mesclada
wo.merge_cells('A3:E3'); wo['A3']='Estação Elevatória:  '+sb; wo['A3'].font=F(10.5,True,C_TITULO); wo['A3'].alignment=AL_L
wo.merge_cells('A4:E4'); wo['A4']='Município:  '+(cidade or '—'); wo['A4'].font=F(10.5,True,C_TITULO); wo['A4'].alignment=AL_L
wo.row_dimensions[3].height=18; wo.row_dimensions[4].height=18
wo['I3']='BDI 1 Obras/Serviços:'; wo['I3'].font=F(9.5,True); wo['I3'].alignment=AL_R
wo['K3']=bdi1; wo['K3'].number_format=PCT; wo['K3'].fill=FILL(C_SUB); wo['K3'].font=F(9.5,True); wo['K3'].alignment=AL_C
wo['I4']='BDI 2 Fornec. Materiais:'; wo['I4'].font=F(9.5,True); wo['I4'].alignment=AL_R
wo['K4']=bdi2; wo['K4'].number_format=PCT; wo['K4'].fill=FILL(C_AMAR2); wo['K4'].font=F(9.5,True); wo['K4'].alignment=AL_C
for idx,(t,w) in enumerate(COLS,1):
    c=wo.cell(5,idx,t); c.fill=FILL(C_TITULO); c.font=F(9.5,True,BRANCO); c.border=BD; c.alignment=AL_C
wo.row_dimensions[5].height=28
r=6; item_no=0; area_subs=[]; area_ini=None; area_nome=None; z=0
def close_area():
    global r, area_ini, area_nome
    if area_ini is not None and area_nome is not None:
        wo.cell(r,1,'   Subtotal — '+area_nome); wo.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
        wo.cell(r,8,f'=SUM(H{area_ini}:H{r-1})'); wo.cell(r,11,f'=SUM(K{area_ini}:K{r-1})')
        band(wo,r,1,12,C_SUB,F(10,True,C_TITULO)); wo.cell(r,1).alignment=AL_L
        for c in (8,11): wo.cell(r,c).number_format=MONEY; wo.cell(r,c).alignment=AL_R
        area_subs.append((area_nome, r)); r+=1
for row in rows:
    kind=row['kind']
    if kind=='area':
        close_area()
        wo.merge_cells(start_row=r,start_column=1,end_row=r,end_column=12)
        wo.cell(r,1,row['desc']); band(wo,r,1,12,C_AREA,F(11,True,BRANCO)); wo.cell(r,1).alignment=AL_L
        wo.row_dimensions[r].height=18; r+=1; area_ini=r; area_nome=row['desc']; z=0
    elif kind in ('grupo','subgrupo'):
        fill = C_GRP if kind=='grupo' else C_SUBG
        fcol = BRANCO if kind=='grupo' else C_TITULO
        wo.merge_cells(start_row=r,start_column=1,end_row=r,end_column=12)
        lbl=(row['cod']+'  ' if row['cod'] else '')+row['desc']
        wo.cell(r,1,lbl); band(wo,r,1,12,fill,F(9.5,True,fcol)); wo.cell(r,1).alignment=AL_L; r+=1
    else:  # item
        item_no+=1; z+=1
        wo.cell(r,1,item_no); wo.cell(r,2,row['cod']); wo.cell(r,2).number_format='@'
        wo.cell(r,3,row['orig']); wo.cell(r,4,row['desc']); wo.cell(r,5,row['un'])
        wo.cell(r,6,row['q']); wo.cell(r,7,row['ku'])
        wo.cell(r,8,f'=IF(G{r}="",0,F{r}*G{r})'); wo.cell(r,9,f'=IF(L{r}="MATERIAL",$K$4,$K$3)')
        wo.cell(r,10,f'=IF(G{r}="","",G{r}*(1+I{r}))'); wo.cell(r,11,f'=IF(J{r}="",0,F{r}*J{r})')
        wo.cell(r,12,row['tipo'])
        fillz=C_ZEBRA if z%2==0 else BRANCO
        for c in range(1,13):
            cc=wo.cell(r,c); cc.border=BD; cc.font=F(9.5); cc.fill=FILL(fillz)
            cc.alignment=AL_L if c==4 else AL_C   # tudo centralizado, exceto DESCRIÇÃO (col D)
        for c in (7,8,10,11): wo.cell(r,c).number_format=MONEY
        wo.cell(r,9).number_format=PCT; r+=1
close_area()
cd_row=r
wo.cell(r,1,'CUSTO DIRETO (sem BDI)'); wo.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
wo.cell(r,8,'='+'+'.join(f'H{gr}' for _,gr in area_subs) if area_subs else 0); wo.cell(r,8).number_format=MONEY
band(wo,r,1,12,C_SUB,F(10,True,C_TITULO)); wo.cell(r,1).alignment=AL_L; wo.cell(r,8).alignment=AL_R; r+=1
tot_row=r
wo.cell(r,1,'VALOR TOTAL COM BDI'); wo.merge_cells(start_row=r,start_column=1,end_row=r,end_column=10)
wo.cell(r,11,'='+'+'.join(f'K{gr}' for _,gr in area_subs) if area_subs else 0); wo.cell(r,11).number_format=MONEY
band(wo,r,1,12,C_VERM,F(12,True,BRANCO)); wo.cell(r,1).alignment=AL_L; wo.cell(r,11).alignment=AL_R; wo.cell(r,11).font=F(12,True,BRANCO)
wo.row_dimensions[r].height=22; wo.freeze_panes='A6'

# ===== RESUMO =====
wr=out.create_sheet('Resumo',0)
for k,w in [('A',26),('B',44),('C',22),('D',14),('E',11)]: wr.column_dimensions[k].width=w
wr.row_dimensions[1].height=20; wr.row_dimensions[2].height=20
wr.merge_cells('A3:E3'); wr['A3']='ORÇAMENTO — ESTAÇÃO ELEVATÓRIA DE ESGOTO (EEE)'; band(wr,3,1,5,C_TITULO,F(15,True,BRANCO)); wr['A3'].alignment=AL_L; wr.row_dimensions[3].height=26
wr.merge_cells('A4:E4'); wr['A4']=sb+(('  —  '+cidade) if cidade else '')+'  —  Padrão SANEPAR  —  2S Engenharia'; band(wr,4,1,5,C_TITULO,F(10,True,BRANCO)); wr['A4'].alignment=AL_L
info=[('Obra:',sb),('Município:',cidade or '—'),('Contratante:','ACCIONA / SANEPAR'),('Projetista:','2S Engenharia'),
      ('Base de preços:','SANEPAR MOS 5ª Ed. - JUN/2025'),('Método:','EEE em linha - Padrão SANEPAR'),
      ('BDI 1 Obras/Serviços:','=Orçamento!$K$3'),('BDI 2 Fornec. Materiais:','=Orçamento!$K$4')]
rr=6
for lbl,val in info:
    wr.cell(rr,1,lbl); wr.cell(rr,1).fill=FILL(C_SUB); wr.cell(rr,1).font=F(10,True,C_TITULO); wr.cell(rr,1).alignment=AL_L
    wr.cell(rr,2,val); wr.cell(rr,2).font=F(10); wr.cell(rr,2).alignment=AL_L
    if isinstance(val,str) and val.endswith(('$K$3','$K$4')): wr.cell(rr,2).number_format=PCT
    rr+=1
rr+=1; wr.cell(rr,1,'RESUMO POR ÁREA'); wr.cell(rr,1).font=F(10,True,C_TITULO); rr+=1
for i,t in enumerate(['ITEM','ÁREA','VALOR C/ BDI (R$)','% TOTAL'],1):
    c=wr.cell(rr,i,t); c.fill=FILL(C_TITULO); c.font=F(9.5,True,BRANCO); c.alignment=AL_C; c.border=BD
rr+=1; g_ini=rr; tot_ref=g_ini+len(area_subs)+1
for idx,(nome,gr) in enumerate(area_subs,1):
    wr.cell(rr,1,idx); wr.cell(rr,2,nome); wr.cell(rr,3,f'=Orçamento!K{gr}'); wr.cell(rr,3).number_format=MONEY
    wr.cell(rr,4,f'=IF($C${tot_ref}=0,0,C{rr}/$C${tot_ref})'); wr.cell(rr,4).number_format=PCT
    for c in range(1,5): wr.cell(rr,c).border=BD; wr.cell(rr,c).font=F(9.5)
    wr.cell(rr,2).alignment=AL_L; wr.cell(rr,3).alignment=AL_R; wr.cell(rr,4).alignment=AL_C; wr.cell(rr,1).alignment=AL_C; rr+=1
wr.cell(rr,1,'CUSTO DIRETO (sem BDI)'); wr.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=2)
wr.cell(rr,3,f'=Orçamento!H{cd_row}'); wr.cell(rr,3).number_format=MONEY
band(wr,rr,1,4,C_SUB,F(10,True,C_TITULO)); wr.cell(rr,1).alignment=AL_L; wr.cell(rr,3).alignment=AL_R; rr+=1
wr.cell(rr,1,'VALOR TOTAL COM BDI'); wr.merge_cells(start_row=rr,start_column=1,end_row=rr,end_column=2)
wr.cell(rr,3,f'=Orçamento!K{tot_row}'); wr.cell(rr,3).number_format=MONEY
band(wr,rr,1,4,C_VERM,F(12,True,BRANCO)); wr.cell(rr,1).alignment=AL_L; wr.cell(rr,3).alignment=AL_R; wr.cell(rr,3).font=F(12,True,BRANCO)
put_logos(wr)

# ===== REFERÊNCIAS =====
wf=out.create_sheet('Referências')
for k,w in [('A',4),('B',30),('C',48),('D',30),('E',60)]: wf.column_dimensions[k].width=w
wf.merge_cells('A1:E1'); wf['A1']='REFERÊNCIAS NORMATIVAS — EEE (FONTE DE CADA CRITÉRIO)'; band(wf,1,1,5,C_TITULO,F(15,True,BRANCO)); wf['A1'].alignment=AL_L; wf.row_dimensions[1].height=26
wf.merge_cells('A2:E2'); wf['A2']='De onde vem cada preço e critério: SANEPAR MOS/MPS, cotação de fornecedor, projeto executivo da obra ou norma (ABNT/TCU). Rastreável item a item.'; wf['A2'].font=F(9,color='666666'); wf['A2'].alignment=AL_L
for i,t in enumerate(['Nº','CRITÉRIO','REGRA APLICADA NO ORÇAMENTO','FONTE (de onde puxa)','OBSERVAÇÃO / ONDE CONFERIR'],1):
    c=wf.cell(4,i,t); c.fill=FILL(C_TITULO); c.font=F(9.5,True,BRANCO); c.alignment=AL_C; c.border=BD
REFS=[
 ('1','Base de preços — itens SANEPAR','Custo unitário de cada item com código SANEPAR puxado direto da tabela oficial de preços.','SANEPAR — MOS 5ª Ed. v02 (JUN/2025)','O código na coluna CÓDIGO identifica o item na tabela MOS (aba REF_TAB_PREÇOS_SANEPAR).'),
 ('2','Base de preços — itens sem código (cotação)','Itens com ORIGEM = CP (tubos/conexões PEAD-FD-PPR, motobomba, painéis) não constam no MOS → preço por cotação de fornecedor.','Cotação de fornecedor (mín. 3 orçamentos → menor preço)','Ver pasta COTAÇÕES da obra; anexar os PDFs de referência.'),
 ('3','BDI duplo','BDI 1 (Obras/Serviços) 24,49% e BDI 2 (Fornecimento de Materiais/Equip.) 12,99%, aplicado item a item pelo TIPO (Obra/Mat).','TCU — Acórdão 2622/2013-Plenário','Cálculo detalhado na aba BDI da planilha-base (parcelas AC/SG/R/DF/L).'),
 ('4','Levantamento de quantidades','Toda a coluna QUANT sai da aba Memória de Cálculo, levantada dos projetos executivos da obra.','Memória de Cálculo (esta planilha) + projetos','Cada quantidade rastreável à fórmula/medida na Memória de Cálculo.'),
 ('5','Movimento de solo / escavação','Escavação e reaterro do poço/PV pela geometria de projeto; VCA (vala a céu aberto) até a profundidade de projeto.','Projeto (implantação/perfil) + SANEPAR MOS','Volume = área × profundidade da escavação (Memória de Cálculo).'),
 ('6','Estrutura do poço e PV','Concreto fck 40 MPa, aço CA-50 e fôrma de chapa conforme dimensionamento; PV em anéis de concreto armado + tampão FD.','Projeto Estrutural (PEST) + SANEPAR MOS','Quantidades de concreto/aço/fôrma no PEST.'),
 ('7','Impermeabilização','Impermeabilização em poliuretano nas faces internas do poço, PV e abrigo.','SANEPAR MOS + projeto','Área das faces internas (Memória de Cálculo).'),
 ('8','Conjunto motobomba','Bomba selecionada pela vazão (Q) e altura manométrica (Hman) do hidráulico; fornecimento e montagem por cotação.','Projeto Hidráulico (PHID) + cotação','Q e Hman de projeto → modelo cotado (pasta COTAÇÕES).'),
 ('9','Instalações elétricas','QCM, entrada de energia, cabos, eletrodutos e aterramento conforme elétrico; fornecimentos por cotação.','Projeto Elétrico (PELE) + cotação','Cargas e componentes definidos no PELE.'),
 ('10','Abrigo (arquitetura / acabamento)','Alvenaria, revestimento (chapisco/emboço/reboco) e pintura no padrão SANEPAR (azul claro/escuro).','Projeto Arquitetônico (PARQ) + SANEPAR MOS','Áreas de parede/pintura na Memória de Cálculo.'),
 ('11','Reposição de pavimento','Reposição do pavimento removido na implantação, no mesmo tipo do existente, dentro da área correspondente.','Projeto + SANEPAR MOS','Área reposta lançada dentro da área a que pertence.'),
 ('12','Norma de referência — EEE','Concepção da estação elevatória de esgoto conforme norma técnica.','ABNT NBR 12208 — Estações elevatórias de esgoto sanitário','Complementar: NBR 9649 (redes) e MPS SANEPAR.'),
 ('13','Limpeza final de obra','Limpeza geral ao término, lançada como grupo próprio ao final do orçamento.','SANEPAR MOS','Grupo LIMPEZA DE OBRA (sempre o último).')]
rr=5
for row in REFS:
    for i,v in enumerate(row,1):
        c=wf.cell(rr,i,v); c.border=BD; c.alignment=AL_L; c.font=(F(9,True,C_VERM) if i==4 else F(9 if i==5 else 9.5, i==1))
        if rr%2==0: c.fill=FILL(C_ZEBRA)
    rr+=1

# ===== MEMÓRIA DE CÁLCULO (copia COM formatação) =====
if 'Memorial de Cálculo' in wb.sheetnames:
    m=wb['Memorial de Cálculo']; wm=out.create_sheet('Memória de Cálculo')
    def rce_map(orig):   # recolore a memória p/ a paleta da aba Orçamento (RCE)
        if not orig: return None
        s=str(orig)[-6:].upper()
        return {'002060':C_TITULO,'1F4E79':C_TITULO,'FF2060':C_TITULO,'D9E1F2':C_SUB}.get(s)
    for row in m.iter_rows():
        for cell in row:
            nc=wm.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                nc.font=copy(cell.font); nc.fill=copy(cell.fill); nc.border=copy(cell.border)
                nc.alignment=copy(cell.alignment); nc.number_format=cell.number_format
                of = cell.fill.fgColor.rgb if (cell.fill and getattr(cell.fill,'patternType',None)=='solid') else None
                newc=rce_map(of)
                if newc:
                    nc.fill=FILL(newc)
                    nc.font=F(cell.font.size or 10, True, BRANCO if newc==C_TITULO else C_TITULO)
    # título principal (A1) na cor RCE
    wm['A1'].fill=FILL(C_TITULO); wm['A1'].font=F(13,True,BRANCO)
    for mc in m.merged_cells.ranges: wm.merge_cells(str(mc))
    # MESCLA o título e o subtítulo (não vinham mesclados -> texto sumia ao colorir)
    ncol=m.max_column or 7
    for rr_ in (1,2):
        try: wm.merge_cells(start_row=rr_,start_column=1,end_row=rr_,end_column=ncol)
        except Exception: pass
    wm['A1'].alignment=Alignment('left','center'); wm['A2'].alignment=Alignment('left','center',wrap_text=True)
    wm.row_dimensions[1].height=20
    for kcol,v in m.column_dimensions.items():
        if v.width: wm.column_dimensions[kcol].width=v.width
    for krow,v in m.row_dimensions.items():
        if v.height: wm.row_dimensions[krow].height=v.height

out.save(OUT)
n_it=sum(1 for x in rows if x['kind']=='item')
print('OK ->', OUT, '| itens:', n_it, '| áreas:', len(area_subs))
