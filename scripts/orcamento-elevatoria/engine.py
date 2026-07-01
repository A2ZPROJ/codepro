# -*- coding: utf-8 -*-
"""
MOTOR de orçamento de Elevatória (EEE) — 2S / Daniel Gama.
Gera, a partir do gabarito A2 + um CONFIG de obra:
  - <saida>.xlsx           (3 abas: ORÇAMENTO SINTÉTICO - <SB>, BDI, Memorial de Cálculo)
  - <saida> - ORÇAMENTO.pdf
  - <saida> - MEMORIAL.pdf

O que é FIXO (aqui no motor): estrutura do orçamento (layout A2 de 210 linhas),
o template do Memorial por área, as fórmulas de quantitativo, o BDI por fórmula,
o religamento do BDI e a exportação de PDF.
O que MUDA por obra (vem do CONFIG): identificação (SB/cidade/contrato), os DADOS
DE ENTRADA (quantidades de projeto), os preços de cotação (CP) com fonte, e os
overrides de preço/quant no orçamento.

Uso:  python engine.py caminho\\do\\config.py
Config = módulo Python com: SB, CIDADE, CONTRATO, A2_PATH, OUT_XLSX, DATA, CP,
         PRICE_UPD (opcional), QTY_UPD (opcional).
"""
import win32com.client as win32, openpyxl, json, os, sys, time, pythoncom, zipfile, importlib.util, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------------------------------------------------------ utilidades
def retry(fn, n=10, wait=0.5):
    for i in range(n):
        try: return fn()
        except pythoncom.com_error as e:
            # call rejected / busy / object-not-ready (transitórios pós Copy de aba)
            if e.hresult in (-2147418111, -2147417846, -2146827864) and i < n-1:
                time.sleep(wait); continue
            raise

ERR = {-2146826288:'#NULL!',-2146826281:'#DIV/0!',-2146826273:'#VALUE!',
       -2146826265:'#REF!',-2146826259:'#NAME?',-2146826252:'#NUM!',-2146826246:'#N/A'}
def sv(v): return ERR.get(v, v) if isinstance(v, int) else v

# ------------------------------------------------------ orçamento (layout A2 fixo)
A2_SHEET = 'ORÇAMENTO SINTÉTICO - SB-A2'
AUX_DEL  = ['Memorial','Quantitativo','Preços tubos e conexões','Elétrico',
            'ORÇAMENTO SINTÉTICO - PA4','Quantitativo PA4']
# orçamento_linha -> chave do Memorial (Quant I referencia o Memorial)
MAP_REF = {
 17:'tapume',19:'placa',22:'locacao',24:'spt_o',26:'trado_o',28:'asbuilt',31:'acesso',33:'tapume',34:'tapume',36:'plc1',37:'plc2',
 41:'escpoco',43:'comppoco',46:'frmpoco',48:'acopoco',50:'cmago',52:'c40poco',54:'bombpoco',56:'imppoco',59:'peadass',61:'fd80ass',62:'fd150ass',
 64:'pvd1',65:'pvd2',68:'piso',
 109:'escpv',111:'comppv',114:'frmpv',116:'acopv',118:'c40pv',120:'bombpv',122:'imppv',
 125:'anel1',126:'anel2',127:'tampao',
 131:'frmab',133:'acoab',135:'c40ab',137:'bombab',139:'impab',142:'alv',
 145:'e_qcm',147:'e_ent',
 151:'chap',152:'chap',153:'chap',155:'pint',156:'pint',157:'pint',158:'pint',159:'pint',
 162:'e_pead',163:'e_aco',164:'e_cabo25',165:'e_cabo10',
 167:'e_auxped',168:'e_ped',171:'e_eletr',172:'e_ajud',
 174:'e_pead',176:'e_aco',178:'e_lb',179:'e_t',181:'e_cx',
 183:'e_cabonu',185:'e_haste',187:'e_split',189:'e_cabo10',190:'e_cabo25',
 192:'e_instr',193:'e_cxinsp',194:'e_term',195:'e_eqp',196:'e_chv',
 198:'e_valae',199:'e_valae',201:'e_valat',202:'e_valat',203:'e_insth',
 205:'e_instp',207:'e_poste',209:'limpeza',
}
MAP_VAL = {71:1, 73:1, 75:9, 76:1, 128:0, 144:1, 148:1}   # valores diretos fixos do orçamento

# ------------------------------------------------------ template do Memorial (fixo)
# kinds: ('area',t) ('sub',t)
#  ('in',key,label,nota,un)              valor = DATA[key]  (amarelo)
#  ('qf',key,label,nota,formula,un)      fórmula (refs {key})
#  ('qd',key,label,nota,un)              valor direto = DATA[key]
#  ('cp',key,label,nota,un)              qty = DATA[key]; preço/fonte = CP[key]
TEMPLATE = [
 ('area','1. SERVIÇOS INICIAIS'),
 ('sub','Dados de entrada (projeto)'),
 ('in','cant_c','Canteiro/locação — comprimento','PARQ implantação','m'),
 ('in','cant_l','Canteiro/locação — largura','PARQ implantação','m'),
 ('in','spt','Sondagem SPT — extensão','SPT sempre 25 m','m'),
 ('in','trado','Sondagem a trado — extensão','','m'),
 ('sub','Quantitativos'),
 ('qf','tapume','Tapume/tela + sinalização (perímetro ×2)','','=({cant_c}+{cant_l})*2','m'),
 ('qd','placa','Placa de obra (chapa metálica 2,0×2,0)','Modelo Sanepar','m²'),
 ('qf','locacao','Locação da obra c/ equip. topográfico','','={cant_c}*{cant_l}','m²'),
 ('qf','spt_o','Sondagem SPT (perfuração c/ estadia)','','={spt}','m'),
 ('qf','trado_o','Sondagem a trado','','={trado}','m'),
 ('qd','asbuilt','Cadastro as-built obra localizada','','ud'),
 ('qd','acesso','Acesso provisório','','m²'),
 ('qd','plc1','Placa advertência 1,0×1,0 m','','ud'),
 ('qd','plc2','Placa advertência 1,0×2,0 m','','ud'),

 ('area','2. POÇO (CAIXA EEE)'),
 ('sub','Dados de entrada (projeto)'),
 ('in','esc_c','Escavação caixa — comprimento','PEST cortes','m'),
 ('in','esc_l','Escavação caixa — largura','PEST cortes','m'),
 ('in','esc_p','Escavação caixa — profundidade','±0,00 → fundo + folga','m'),
 ('in','c40_poco','Concreto fck40 — poço','PEST-1: total − PV − base','m³'),
 ('in','frm_poco','Fôrma chapa 12mm — poço','PEST-1: total − PV − base','m²'),
 ('in','aco_pc','Aço CA-50 — caixa EEE','Relação de aço PEST-2','kg'),
 ('in','cmag','Concreto magro fck15 (lastro)','Quantitativo estrutural','m³'),
 ('in','rec160','Recalque interno PEAD DE160','Trecho dentro da EEE','m'),
 ('in','piso_e','Piso concreto — espessura','','m'),
 ('in','calcada','Calçada — área (ao redor da estação)','PARQ planta baixa 1:25','m²'),
 ('in','imp_pa_h','Imperm. parede A — altura','','m'),
 ('in','imp_pa_l','Imperm. parede A — largura','','m'),
 ('in','imp_pb_h','Imperm. parede B — altura','','m'),
 ('in','imp_pb_l','Imperm. parede B — largura','','m'),
 ('in','imp_tp_c','Imperm. tampa — comprimento','','m'),
 ('in','imp_tp_l','Imperm. tampa — largura','','m'),
 ('sub','Movimento de solos'),
 ('qf','escpoco','Escavação mecânica não-vala 0<h≤4m','','={esc_c}*{esc_l}*{esc_p}','m³'),
 ('qf','comppoco','Compactação não em valas (s/ GC)','','={esc_c}*{esc_l}*{esc_p}','m³'),
 ('sub','Fundações / estrutura'),
 ('qf','frmpoco','Fôrma chapa resinada 12mm — poço','','={frm_poco}','m²'),
 ('qf','acopoco','Aço CA-50 — poço','','={aco_pc}','kg'),
 ('qf','c40poco','Concreto fck40 slump12 — poço','','={c40_poco}','m³'),
 ('qf','cmago','Concreto magro fck15 (lastro)','','={cmag}','m³'),
 ('qf','bombpoco','Bombeamento de concreto — poço','','={c40_poco}','m³'),
 ('qf','imppoco','Impermeab. poliuretano — poço','','=(({imp_pa_h}*{imp_pa_l})+({imp_pb_h}*{imp_pb_l}))*2+({imp_tp_c}*{imp_tp_l})*2','m²'),
 ('sub','Assentamentos'),
 ('qf','peadass','Assentamento tubulação PEAD DE160','','={rec160}','m'),
 ('qf','fd80ass','Assentamento tubulação FD DN80','0,5+0,7+1,2+0,7+0,9','=0.5+0.7+1.2+0.7+0.9','m'),
 ('qd','fd150ass','Assentamento tubulação FD DN150','','m'),
 ('sub','Pavimentação / acessos'),
 ('qf','piso','Piso de concreto desempenado (calçada)','','={calcada}*{piso_e}','m³'),
 ('qd','escada','Escada tipo marinheiro PRFV <3,5m','PARQ item 3','ud'),
 ('sub','Materiais hidráulicos — COTAÇÃO (PHID-03)'),
 ('cp','m01','Tubo PEAD PE100 PN10 ocre DE160 L=1,6m','','ud'),
 ('cp','m02','Colarinho PEAD termofusão PE100 PN10 DE160','','ud'),
 ('cp','m03','Toco tubo FD FF 0,25m esgoto DN150','','ud'),
 ('cp','m04','Conjunto motobomba em linha','','cj'),
 ('cp','m05','Tê redução FD FFF PN16 DN100/DN80','','ud'),
 ('cp','m06','Ventosa tríplice esgoto PN10/16 DN100','','ud'),
 ('cp','m07','Tê FD FFF PN16 esgoto DN80','','ud'),
 ('cp','m08','Toco tubo FD FF 700mm aba PN16 DN80','','ud'),
 ('cp','m09','Colarinho PEAD termofusão PE100 PN10 DE90','','ud'),
 ('cp','m10','Válvula guilhotina FD bidirec. PN16 DN80','','ud'),
 ('cp','m11','Tubo FD K9 FF solda interf. 1,20m DN80','','ud'),
 ('cp','m12','Extremidade FD PF 700mm aba PN16 DN80','','ud'),
 ('cp','m13','Curva FD JE 2GS BB 22 esgoto DN80','','ud'),
 ('cp','m14','Tubo FD K9 PP esgoto 0,9m DN80','','ud'),
 ('cp','m15','Conj. motobomba submersível (drenagem)','','ud'),
 ('cp','m16','Adaptador PPR80 bolsa/rosca PN10 DE90','','ud'),
 ('cp','m17','Tubo PPR80 JS PP termofusão 6m PN10 DE90','','ud'),
 ('cp','m18','União PPR flangeada DE90 PN10','','ud'),
 ('cp','m19','Curva longa 90° PPR DE90','','ud'),
 ('cp','m20','Flange aço carbono p/ colarinho PEAD DE160 DN150','','cj'),
 ('cp','m21','Flange aço carbono p/ colarinho PEAD DE90 DN80','','cj'),
 ('cp','m22','Kit parafuso flange colarinho PEAD DN80','','cj'),
 ('cp','m23','Kit parafuso flange colarinho PEAD DN150','','cj'),
 ('cp','m24','Kit parafuso válvula wafer DN80','','cj'),
 ('cp','m25','Kit parafuso esgoto PN16 DN80','','cj'),
 ('cp','m26','Kit parafuso esgoto PN16 DN100','','cj'),
 ('cp','m27','Kit parafuso inox esgoto PN10 DN150','','cj'),

 ('area','3. POÇO DE VISITA (PV CHEGADA)'),
 ('sub','Dados de entrada (projeto)'),
 ('in','pv_c','Escavação PV — comprimento','','m'),
 ('in','pv_l','Escavação PV — largura','','m'),
 ('in','pv_p','Escavação PV — profundidade','PEST: fundo PV','m'),
 ('in','frm_pv','Fôrma chapa 12mm — PV','PEST','m²'),
 ('in','aco_pv','Aço CA-50 — PV entrada','Relação de aço PEST-3','kg'),
 ('in','c40_pv','Concreto fck40 — PV','PEST','m³'),
 ('in','imp_pv','Imperm. PV — área','π·D·H','m²'),
 ('sub','Quantitativos'),
 ('qf','escpv','Escavação mecânica PV 0<h≤4m','','={pv_c}*{pv_l}*{pv_p}','m³'),
 ('qf','comppv','Compactação PV','','={pv_c}*{pv_l}*{pv_p}','m³'),
 ('qf','frmpv','Fôrma chapa resinada 12mm — PV','','={frm_pv}','m²'),
 ('qf','acopv','Aço CA-50 — PV entrada','','={aco_pv}','kg'),
 ('qf','c40pv','Concreto fck40 — PV','','={c40_pv}','m³'),
 ('qf','bombpv','Bombeamento de concreto — PV','','={c40_pv}','m³'),
 ('qf','imppv','Impermeab. poliuretano — PV','','={imp_pv}','m²'),
 ('qd','pvd1','PV tipo D DN1200 — prof até 2,00m','','ud'),
 ('qf','pvd2','PV tipo D DN1200 — acréscimo >2,00m','','={pv_p}-2','m'),
 ('qd','anel1','Poço em anéis DN1200 até 1,00m','','ud'),
 ('qf','anel2','Poço em anéis — acréscimo >1,00m','','={pv_p}-1','m'),
 ('qd','tampao','Tampão FD 850/910mm (PARQ item 1)','','ud'),

 ('area','4. ABRIGO / PAINEL ELÉTRICO'),
 ('sub','Dados de entrada (projeto)'),
 ('in','frm_ab','Fôrma chapa 12mm — base painel','PEST','m²'),
 ('in','aco_ab','Aço CA-50 — base painel','Relação de aço PEST-3','kg'),
 ('in','c40_ab','Concreto fck40 — base painel','PEST','m³'),
 ('in','imp_ab','Imperm. abrigo — área','','m²'),
 ('in','mur_c','Mureta — comprimento','','m'),
 ('in','mur_h','Mureta — altura','','m'),
 ('in','rev','Área revestida (chapisco/pintura)','Mureta 2 faces + cinta','m²'),
 ('sub','Quantitativos'),
 ('qf','frmab','Fôrma chapa resinada 12mm — base painel','','={frm_ab}','m²'),
 ('qf','acoab','Aço CA-50 — base painel','','={aco_ab}','kg'),
 ('qf','c40ab','Concreto fck40 — base painel','','={c40_ab}','m³'),
 ('qf','bombab','Bombeamento de concreto — base painel','','={c40_ab}','m³'),
 ('qf','impab','Impermeab. poliuretano — abrigo','','={imp_ab}','m²'),
 ('qf','alv','Alvenaria tijolo maciço — mureta','','={mur_c}*{mur_h}','m²'),
 ('qf','chap','Chapisco/Emboço/Reboco','','={rev}','m²'),
 ('qf','pint','Pinturas (raspagem→2 demãos azul)','','={rev}','m²'),
 ('cp','e_qcm','Painel elétrico QCM 2 motores inv. 3CV','PELE 02.02','ud'),
 ('qd','e_ent','Entrada de energia trifásica em mureta','PELE 01.06','ud'),

 ('area','5. ELÉTRICO / EXTERNO (LIGAÇÕES PREDIAIS) — PELE'),
 ('sub','Eletrodutos e caixas'),
 ('qd','e_pead','Eletroduto PEAD flex. 2" subterrâneo','PELE 01.01','m'),
 ('qd','e_aco','Eletroduto aço galv. 2" (FG)','PELE 01.02','m'),
 ('qd','e_lb','Condulete LB alumínio 2"','PELE 01.03','ud'),
 ('qd','e_t','Condulete T alumínio 2"','PELE 01.04','ud'),
 ('qd','e_cx','Caixa concreto pré-mold. 0,40³','PELE 01.05','ud'),
 ('cp','e_cxinsp','Caixa inspeção propileno DN300','PELE 04.06','ud'),
 ('cp','e_eqp','Caixa equipotencialização 9 furos','PELE 04.02','ud'),
 ('sub','Cabos e conectores'),
 ('qd','e_cabonu','Cabo de cobre nu 50mm²','PELE 04.01','m'),
 ('qf','e_cabo10','Cabo cobre 10mm² (36+12+12)','PELE 05.01+02+03','=36+12+12','m'),
 ('cp','e_cabo25','Cabo cobre 2,5mm² tetrapolar','PELE 05.04','m'),
 ('qd','e_instr','Cabo instrumentação blindado 1x2#1,5','PELE 05.05','m'),
 ('cp','e_term','Terminal estanhado 1 furo','PELE 04.04','ud'),
 ('qd','e_split','Conector split-bolt','PELE 04.05','ud'),
 ('sub','Aterramento / instrumentação'),
 ('qd','e_haste','Haste de aterramento 3/4" 3,0m','PELE 04.03','ud'),
 ('cp','e_chv','Chave de nível tipo pera NA+NF','PELE 02.03','ud'),
 ('sub','Mão de obra'),
 ('qd','e_eletr','Eletricista industrial (h)','PELE 03.01','h'),
 ('qd','e_ajud','Ajudante de eletricista (h)','PELE 03.02','h'),
 ('qd','e_ped','Pedreiro (h)','PELE 03.03','h'),
 ('qd','e_auxped','Auxiliar de pedreiro (h)','PELE 03.04','h'),
 ('qd','e_comis','Comissionamento/startup 2 CMB c/ CLP (h)','PELE 03.05 — PENDENTE','h'),
 ('sub','Instalações, valas e limpeza'),
 ('qf','e_valae','Vala + assentamento eletroduto','','={e_pead}','m'),
 ('qf','e_valat','Vala + assentamento cabo aterramento','','={e_cabonu}','m'),
 ('qf','e_insth','Instalação de haste de aterramento','','={e_haste}','ud'),
 ('qd','e_instp','Instalação de poste','','ud'),
 ('qd','e_poste','Poste de concreto armado duplo T 7m','','ud'),
 ('qf','limpeza','Limpeza de obra localizada','','={locacao}','m²'),
]

# ------------------------------------------------------ helpers do Memorial
def _ptbr(v):
    try: return ('%.2f' % float(v)).replace('.', ',')
    except Exception: return str(v)

def _grandeza(label):
    """Nome curto da GRANDEZA a partir do rótulo do dado de entrada.
    'Escavação caixa — profundidade' -> 'profundidade'; 'Calçada — área (...)' -> 'área'."""
    s = label
    if '—' in s: s = s.split('—')[-1]
    s = s.split('(')[0].strip()
    return s or label.strip()

# nome da grandeza por chave (derivado dos DADOS DE ENTRADA do TEMPLATE) — usado p/ a fórmula simbólica
GRANDEZA = {t[1]: _grandeza(t[2]) for t in TEMPLATE if t[0] == 'in'}
# rótulo do resultado pela unidade (ex.: m² -> "Área = ...")
_ROTULO_UN = {'m²':'Área','m2':'Área','m³':'Volume','m3':'Volume','kg':'Peso','m':'Compr.','ud':'Qtd.','un':'Qtd.'}

def _fmt_expr(expr, repl):
    s = re.sub(r'\{(\w+)\}', repl, expr)
    s = (s.replace('*', ' × ').replace('/', ' ÷ ').replace('+', ' + ').replace('-', ' − '))
    return re.sub(r'\s+', ' ', s).strip()

def memo_calc(formula, DATA, unit, nota=''):
    """Conta DIDÁTICA em 3 etapas, p/ qualquer um entender o que foi feito:
        FÓRMULA (em grandezas) -> SUBSTITUIÇÃO (valores) -> RESULTADO.
    ex.:  Área = comprimento × largura
          = 8,50 × 3,00
          = 25,50 m²
    Fórmula de valor único (={spt}) vira só o número; fórmula de constantes (=36+12+12)
    pula a linha simbólica (não há grandeza a nomear)."""
    expr = formula.lstrip('=')
    symb = _fmt_expr(expr, lambda m: GRANDEZA.get(m.group(1), m.group(1)))
    nums = _fmt_expr(expr, lambda m: _ptbr(DATA.get(m.group(1), 0)))
    numexpr = re.sub(r'\{(\w+)\}', lambda m: repr(float(DATA.get(m.group(1), 0))), expr)
    try: res = eval(numexpr, {'__builtins__': {}})
    except Exception: res = None
    has_op = any(o in nums for o in (' × ',' ÷ ',' + ',' − '))
    nota_ok = bool(nota and re.search(r'[A-Za-zÀ-ÿ]', nota))
    pref = (nota + '\n') if nota_ok else ''
    if not has_op or res is None:                       # valor único -> só o número
        return (nota + ' → ' if nota_ok else '') + (_ptbr(res) if res is not None else nums)
    if symb == nums:                                    # constantes (sem grandeza p/ nomear)
        return pref + '%s = %s %s' % (nums, _ptbr(res), unit)
    rot = _ROTULO_UN.get(unit, 'Qtd.')
    return pref + '%s = %s\n= %s\n= %s %s' % (rot, symb, nums, _ptbr(res), unit)

def _origem_curta(orig):
    """Coluna ORIGEM (H) do ORÇAMENTO: rótulo CURTO, SEM valor e SEM site.
    O detalhe (preço de internet + link) aparece SÓ na aba Memorial (seção FONTES DE PREÇO).
    Segue os rótulos que o A2 já usa: Sanepar / CP / Estimativa. ('estimativa' é testado
    primeiro porque alguns textos de estimativa citam 'SANEPAR/SINAPI')."""
    t = orig or ''; low = t.lower()
    if 'estimativa' in low: return 'Estimativa'
    if 'sanepar'    in low: return 'Sanepar'
    if 'sinapi'     in low: return 'SINAPI'
    if ('http' in low or 'internet' in low or 'cotaç' in low or 'cotac' in low): return 'CP'
    return (re.split(r'[\s:(]', t.strip())[0] or 'CP')

def set_fonte(cell, text):
    """Escreve a fonte; se tiver URL, vira HYPERLINK azul clicável."""
    text = text or ''
    cell.value = text
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    m = re.search(r'(https?://\S+)', text)
    if m:
        try: cell.hyperlink = m.group(1)
        except Exception: pass
        cell.font = Font(size=8, color='0563C1', underline='single')
    else:
        cell.font = Font(size=8)

# ------------------------------------------------------ build do Memorial (openpyxl)
def build_memorial(cfg, mem_path):
    C_HDR=PatternFill('solid',fgColor='1F4E79'); C_AREA=PatternFill('solid',fgColor='002060')
    C_SUB=PatternFill('solid',fgColor='D9E1F2'); C_IN=PatternFill('solid',fgColor='FFF2CC')
    F_HDR=Font(color='FFFFFF',bold=True,size=10); F_AREA=Font(color='FFFFFF',bold=True,size=11)
    F_SUB=Font(color='1F3864',bold=True,size=10); F_NOTE=Font(size=8,italic=True,color='808080')
    thin=Side(style='thin',color='BFBFBF'); BD=Border(thin,thin,thin,thin)
    AL_L=Alignment(horizontal='left',vertical='center',wrap_text=True)
    AL_C=Alignment(horizontal='center',vertical='center',wrap_text=True)
    num='#,##0.00'; money='#,##0.00'
    DATA, CP = cfg.DATA, cfg.CP

    wb=openpyxl.Workbook(); ws=wb.active; ws.title='Memorial de Cálculo'
    # pass1: linhas + anchors
    anchors={}; r=5; plan=[]
    for t in TEMPLATE:
        plan.append((r,t))
        if t[0] in ('in','qf','qd','cp'): anchors[t[1]]='D%d'%r
        r+=1
    LAST=r-1
    def resolve(f):
        out=f
        for k,a in anchors.items(): out=out.replace('{%s}'%k,a)
        return out
    # título / cabeçalho
    ws['A1']='MEMORIAL DE CÁLCULO — EEE %s — %s'%(cfg.SB, cfg.CIDADE)
    ws['A1'].font=Font(bold=True,size=13,color='1F3864')
    ws['A2']=('Contrato %s · EEE em Linha (poço seco NT-24) · Padrão SANEPAR · Organizado POR ÁREA. '
              'Células AMARELAS = dados de entrada do projeto; coluna QUANT. = fórmula sobre esses dados.'%cfg.CONTRATO)
    ws['A2'].font=F_NOTE
    for c,txt in enumerate(['BLOCO / ITEM','','MEMÓRIA DE CÁLCULO','QUANT.','UN','PREÇO UNIT. (R$)','FONTE / COTAÇÃO (item CP)'],1):
        cc=ws.cell(4,c,txt); cc.fill=C_HDR; cc.font=F_HDR; cc.border=BD; cc.alignment=AL_C
    ws.merge_cells('A4:B4'); ws['A4'].alignment=AL_L
    # pass2: escreve
    miss=[]
    for r,t in plan:
        k=t[0]
        if k=='area':
            ws.cell(r,1,t[1]); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
            for c in range(1,8): ws.cell(r,c).fill=C_AREA; ws.cell(r,c).border=BD
            ws.cell(r,1).font=F_AREA; ws.cell(r,1).alignment=AL_L
        elif k=='sub':
            ws.cell(r,1,t[1]); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
            for c in range(1,8): ws.cell(r,c).fill=C_SUB; ws.cell(r,c).border=BD
            ws.cell(r,1).font=F_SUB; ws.cell(r,1).alignment=AL_L
        elif k=='in':
            _,key,label,nota,un=t
            if key not in DATA: miss.append(key)
            ws.cell(r,2,label).alignment=AL_L
            ws.cell(r,3,nota).font=F_NOTE; ws.cell(r,3).alignment=AL_L
            d=ws.cell(r,4,DATA.get(key)); d.fill=C_IN; d.number_format=num; d.alignment=AL_C; d.font=Font(bold=True)
            ws.cell(r,5,un).alignment=AL_C
            for c in range(2,8): ws.cell(r,c).border=BD
        else:  # qf, qd, cp
            label=t[2]; nota=t[3]
            ws.cell(r,2,label).alignment=AL_L
            if k=='qf':
                formula=t[4]; un=t[5]
                ws.cell(r,4, resolve(formula))
                ctxt = memo_calc(formula, DATA, un, nota)     # 3 etapas: fórmula -> valores -> resultado
            else:
                un=t[4]
                if t[1] not in DATA: miss.append(t[1])
                ws.cell(r,4, DATA.get(t[1]))
                ctxt = nota
            ws.cell(r,3,ctxt).font=F_NOTE; ws.cell(r,3).alignment=AL_L
            d=ws.cell(r,4); d.number_format=num; d.alignment=AL_C
            ws.cell(r,5,un).alignment=AL_C
            if k=='cp':
                preco,fonte = CP.get(t[1],(None,'(sem cotação)'))
                if preco is not None:
                    p=ws.cell(r,6,preco); p.number_format=money; p.alignment=AL_C
                if fonte:
                    set_fonte(ws.cell(r,7), fonte)           # fonte com link clicável
            for c in range(2,8): ws.cell(r,c).border=BD
    # ---- seção: FONTES DE PREÇO de cotação/internet (com link) ----
    extras=[]
    for blk in (getattr(cfg,'AREA_INSERTS',[]) or []):
        for it in blk[2]: extras.append((it[1],it[4],it[5],it[6],it[3]))  # desc,q,un,val,fonte(orig)
    eb=getattr(cfg,'EXTRA_BLOCK',None)
    if eb:
        for it in eb['itens']: extras.append((it[1],it[4],it[5],it[6],it[3]))
    for it in (getattr(cfg,'MEMO_FONTES',[]) or []):
        extras.append(it)  # já em (desc,q,un,val,fonte)
    if extras:
        r+=1; ws.cell(r,1,'FONTES DE PREÇO — COTAÇÃO / INTERNET')
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
        for c in range(1,8): ws.cell(r,c).fill=C_AREA; ws.cell(r,c).border=BD
        ws.cell(r,1).font=F_AREA; ws.cell(r,1).alignment=AL_L
        r+=1; ws.cell(r,1,'Itens fora do catálogo SANEPAR — preço de cotação/internet (clique na FONTE p/ abrir o link)')
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)
        for c in range(1,8): ws.cell(r,c).fill=C_SUB; ws.cell(r,c).border=BD
        ws.cell(r,1).font=F_SUB; ws.cell(r,1).alignment=AL_L
        for desc,q,un,val,fonte in extras:
            r+=1
            ws.cell(r,2,desc).alignment=AL_L
            if q is not None:
                d=ws.cell(r,4,q); d.number_format=num; d.alignment=AL_C
            ws.cell(r,5,un).alignment=AL_C
            if val is not None:
                p=ws.cell(r,6,val); p.number_format=money; p.alignment=AL_C
            set_fonte(ws.cell(r,7), fonte)
            for c in range(2,8): ws.cell(r,c).border=BD
    for col,w in [('A',4),('B',46),('C',44),('D',11),('E',6),('F',14),('G',56)]:
        ws.column_dimensions[col].width=w
    ws.freeze_panes='A5'
    wb.save(mem_path)
    if miss: print('AVISO: chaves sem valor no DATA:', sorted(set(miss)))
    return anchors

# ------------------------------------------------------ pipeline COM
def run(cfg):
    out_xlsx = cfg.OUT_XLSX
    sheet = 'ORÇAMENTO SINTÉTICO - %s' % cfg.SB
    mem_tmp = os.path.join(os.path.dirname(out_xlsx), '_memorial_tmp.xlsx')
    anchors = build_memorial(cfg, mem_tmp)
    for key in set(MAP_REF.values()):
        if key not in anchors: raise SystemExit('chave do orçamento ausente no Memorial: %s' % key)
    import shutil; shutil.copy(cfg.A2_PATH, out_xlsx)

    xl=win32.DispatchEx('Excel.Application'); xl.Visible=False; xl.DisplayAlerts=False
    try:
        wb=retry(lambda: xl.Workbooks.Open(out_xlsx, UpdateLinks=0))
        ws=wb.Worksheets(A2_SHEET)
        retry(lambda: xl.CalculateFull())
        # 1) congelar I,K como valores (quebra refs externas)
        for col in ('I','K'):
            rng=ws.Range('%s15:%s209'%(col,col)); rng.Value=rng.Value
        # 2) título/nome
        ws.Range('K6').Value=cfg.SB; ws.Range('K5').Value='EEE (SES)'
        ws.Name=sheet
        # 3) apagar abas auxiliares do A2
        for s in AUX_DEL:
            try: wb.Worksheets(s).Delete()
            except Exception: pass
        # 4) religar BDI ao custo direto LOCAL + quebrar links externos
        bdi=wb.Worksheets('BDI')
        bdi.Range('F1').Formula = "='%s'!O210" % sheet
        try:
            for src in (wb.LinkSources(1) or []): wb.BreakLink(src,1)
        except Exception: pass
        # 5) importar Memorial novo
        src=retry(lambda: xl.Workbooks.Open(mem_tmp))
        retry(lambda: src.Worksheets('Memorial de Cálculo').Copy(None, wb.Worksheets(wb.Worksheets.Count)))
        src.Close(False)
        time.sleep(0.5)
        ws=retry(lambda: wb.Worksheets(sheet))
        retry(lambda: xl.CalculateFull())
        # 6) Quant (I) -> Memorial
        for r,key in MAP_REF.items():
            retry(lambda r=r,key=key: setattr(ws.Cells(r,9),'Formula',"='Memorial de Cálculo'!%s"%anchors[key]))
        # 7) valores diretos do orçamento
        for r,v in MAP_VAL.items():
            retry(lambda r=r,v=v: setattr(ws.Cells(r,9),'Value',v))
        for r,v in getattr(cfg,'QTY_UPD',{}).items():
            retry(lambda r=r,v=v: setattr(ws.Cells(r,9),'Value',v))
        # 8) BDI: TODA coluna M = fórmula ligada à aba BDI
        for r in range(15,210):
            g=retry(lambda r=r: ws.Cells(r,7).Value)
            if g in ('Mat','MO'):
                retry(lambda r=r: setattr(ws.Cells(r,13),'Formula','=IF(G%d="Mat",BDI!$C$31,IF(G%d="MO",BDI!$C$17,""))'%(r,r)))
        # 9) preços reconciliados no orçamento (K)
        for r,p in getattr(cfg,'PRICE_UPD',{}).items():
            retry(lambda r=r,p=p: setattr(ws.Cells(r,11),'Value',p))
        # 9-cod) código (col B) reconciliado c/ o projeto (ex.: código SANEPAR real do item)
        for r,c in getattr(cfg,'CODE_UPD',{}).items():
            retry(lambda r=r: setattr(ws.Cells(r,2),'NumberFormat','@'))   # texto: preserva zero à esq.
            retry(lambda r=r,c=c: setattr(ws.Cells(r,2),'Value',str(c)))
        # 9-fix) BLINDAR N/O/P de TODO item Mat/MO. Sem isso, item SEM preço (K vazio — ex. Painel
        #         QCM/e_qcm pendente) faz N="" e P=N*I => #VALUE!, que contamina o SUBTOTAL da ÁREA
        #         e zera o VALOR TOTAL (bug achado pelo PREDATOR na SB-A5). Com IF(K="",...) o item
        #         pendente vira "" e o SUM/SUBTOTAL ignora — o total fecha mesmo com item em aberto.
        # O e P do item pendente viram 0 (não ""): os cabeçalhos de grupo somam com "+" explícito
        # (ex. =P144+P145), e numero+"" também dá #VALUE! — só 0 fecha a conta. N fica "" (preço
        # unit. em branco, sinaliza a pendência); O/P=0 -> total geral fecha ignorando o item.
        for r in range(15,210):
            g=retry(lambda r=r: ws.Cells(r,7).Value)
            if g in ('Mat','MO'):
                retry(lambda r=r: setattr(ws.Cells(r,14),'Formula','=IF(K%d="","",K%d*(1+M%d))'%(r,r,r)))
                retry(lambda r=r: setattr(ws.Cells(r,15),'Formula','=IF(K%d="",0,K%d*I%d)'%(r,r,r)))
                retry(lambda r=r: setattr(ws.Cells(r,16),'Formula','=IF(K%d="",0,N%d*I%d)'%(r,r,r)))
        # 9a) inserir itens NOVOS DENTRO do grupo da ÁREA certa (Lucas: "por área")
        area_inserts(xl, ws, cfg)
        # 9b) bloco extra (ex.: PAVIMENTAÇÃO ASFÁLTICA) anexado antes do VALOR TOTAL
        add_extra_block(xl, ws, cfg)
        # 9c) remover linhas de item com QUANT=0 (itens que não se aplicam à obra)
        nrem = prune_zero_qty(xl, ws)
        retry(lambda: xl.CalculateFull())
        vt=210
        for r in range(205,330):
            v=ws.Cells(r,2).Value or ws.Cells(r,5).Value
            if v and 'VALOR TOTAL' in str(v).upper(): vt=r; break
        custo=sv(ws.Cells(vt,15).Value); total=sv(ws.Cells(vt,16).Value)
        wb.Save()
        # 10) PDFs
        pdfs = export_pdfs(xl, wb, sheet, out_xlsx)
        wb.Close(True)
        if os.path.exists(mem_tmp): os.remove(mem_tmp)
        print('Custo direto O210 =', custo)
        print('TOTAL c/BDI  P210 =', total)
        print('xlsx :', out_xlsx)
        for p in pdfs: print('pdf  :', p)
        return total
    finally:
        xl.Quit()

def area_inserts(xl, ws, cfg):
    """Insere itens NOVOS DENTRO de um grupo existente (mesma ÁREA do resto do orçamento),
    estendendo a fórmula do cabeçalho do grupo p/ incluí-los. Assim o subtotal do grupo, o
    cabeçalho da ÁREA e o VALOR TOTAL fecham sozinhos (o Excel ajusta as refs no Insert).
    cfg.AREA_INSERTS = [(after_line, header_line, [item,...]), ...]  (linhas do A2 ORIGINAL).
      after_line  = última linha de item do grupo (insere logo depois)
      header_line = linha do cabeçalho do grupo (cuja fórmula P/O será estendida)
      item        = (cod, desc, tipo, orig, quant, un, valor)
    Processa em ordem DECRESCENTE de after_line p/ não deslocar âncoras ainda não usadas."""
    blocks = getattr(cfg, 'AREA_INSERTS', None)
    if not blocks: return
    xlDown = -4121; ITEM = 41
    for after_line, header_line, items in sorted(blocks, key=lambda x: x[0], reverse=True):
        n = len(items); b = after_line
        for _ in range(n):
            retry(lambda: ws.Rows(ITEM).Copy())
            retry(lambda b=b: ws.Rows(b+1).Insert(xlDown))
        xl.CutCopyMode = False
        for k, it in enumerate(items):
            cod, desc, tipo, orig, q, un, val = it
            r = b+1+k
            retry(lambda r=r: setattr(ws.Cells(r,2), 'NumberFormat', '@'))  # código como TEXTO
            for col, value in [(2,cod),(5,desc),(7,tipo),(8,_origem_curta(orig)),(9,q),(10,un),(11,val)]:
                retry(lambda r=r, col=col, value=value: setattr(ws.Cells(r,col), 'Value', value))
            retry(lambda r=r: setattr(ws.Cells(r,13),'Formula','=IF(G%d="Mat",BDI!$C$31,IF(G%d="MO",BDI!$C$17,""))'%(r,r)))
        # estender a fórmula do cabeçalho do grupo (col P=16 e, se houver, O=15)
        for col in (16, 15):
            f = ws.Cells(header_line, col).Formula
            if not (isinstance(f,str) and f.startswith('=')): continue
            L = 'P' if col==16 else 'O'
            if 'SUBTOTAL' in f.upper() or 'SUM(' in f.upper():
                nf = re.sub(r':%s%d\b' % (L,b), ':%s%d' % (L,b+n), f)
            else:
                nf = f + ''.join('+%s%d' % (L, b+1+k) for k in range(n))
            retry(lambda col=col, nf=nf: setattr(ws.Cells(header_line,col), 'Formula', nf))

def prune_zero_qty(xl, ws):
    """Remove (deleta) as linhas de ITEM com QUANT (col I)=0 — itens do gabarito que NÃO se
    aplicam a esta obra. Antes de deletar, tira o termo +Pr/Pr+ (e +Or/Or+) de qualquer
    cabeçalho que some a linha EXPLICITAMENTE (senão vira #REF!); somas por RANGE
    (SUBTOTAL/SUM) o Excel reajusta sozinho ao deletar. De baixo p/ cima p/ não deslocar
    as linhas ainda não tratadas. Retorna a lista de descrições removidas."""
    vt = None
    for r in range(15, 400):
        v = ws.Cells(r,2).Value or ws.Cells(r,5).Value
        if v and 'VALOR TOTAL' in str(v).upper(): vt = r; break
    if not vt: return []
    targets, descr = [], {}
    for r in range(15, vt):
        g = ws.Cells(r,7).Value
        if g in ('Mat','MO'):
            i = ws.Cells(r,9).Value
            try: iv = float(i) if i not in (None,'') else None
            except Exception: iv = None
            if iv == 0:
                targets.append(r); descr[r] = ws.Cells(r,5).Value
    for r in sorted(targets, reverse=True):
        for h in range(14, vt+1):
            for col, L in ((16,'P'),(15,'O')):
                f = ws.Cells(h,col).Formula
                if not (isinstance(f,str) and f.startswith('=')): continue
                if ('%s%d'%(L,r)) not in f: continue
                if 'SUBTOTAL' in f.upper() or 'SUM(' in f.upper(): continue  # range -> Excel ajusta
                f2 = re.sub(r'\+%s%d\b'%(L,r), '', f)        # remove +Pr (meio/fim)
                if f2 == f: f2 = re.sub(r'%s%d\b\+'%(L,r), '', f)  # ou Pr+ (início)
                if f2 != f: retry(lambda h=h,col=col,f2=f2: setattr(ws.Cells(h,col),'Formula',f2))
        retry(lambda r=r: ws.Rows(r).Delete())
    return [descr[r] for r in sorted(targets)]

def add_extra_block(xl, ws, cfg):
    """Anexa um bloco de itens (ex.: pavimentação asfáltica) ANTES da linha VALOR TOTAL,
    copiando o formato de um cabeçalho de bloco (linha 38) e de um item (linha 41) do A2.
    Não afeta MAP_REF (linhas 17-209), pois insere depois da 209."""
    blk = getattr(cfg, 'EXTRA_BLOCK', None)
    if not blk: return
    itens = blk['itens']; n = len(itens)
    vt = None
    for r in range(190, 320):
        v = ws.Cells(r,2).Value or ws.Cells(r,5).Value
        if v and 'VALOR TOTAL' in str(v).upper(): vt = r; break
    if not vt: raise SystemExit('VALOR TOTAL não encontrado para o bloco extra')
    HDR, ITEM, xlDown = 38, 41, -4121
    retry(lambda: ws.Rows(HDR).Copy())
    retry(lambda: ws.Rows(vt).Insert(xlDown))            # cabeçalho do bloco em vt
    for _ in range(n):
        retry(lambda: ws.Rows(ITEM).Copy())              # re-copiar a cada inserção
        retry(lambda: ws.Rows(vt+1).Insert(xlDown))      # n itens em vt+1..vt+n
    xl.CutCopyMode = False
    hr = vt
    retry(lambda: setattr(ws.Cells(hr,2), 'Value', None))           # limpa código copiado
    retry(lambda: setattr(ws.Cells(hr,5), 'Value', blk['titulo']))  # rótulo do bloco fica na coluna E
    for c in (9,11,13,14):
        retry(lambda c=c: setattr(ws.Cells(hr,c), 'Value', None))
    retry(lambda: setattr(ws.Cells(hr,15), 'Formula', '=SUM(O%d:O%d)' % (vt+1, vt+n)))
    retry(lambda: setattr(ws.Cells(hr,16), 'Formula', '=SUM(P%d:P%d)' % (vt+1, vt+n)))
    for k, it in enumerate(itens):
        cod, desc, tipo, orig, q, un, val = it
        r = vt+1+k
        retry(lambda r=r: setattr(ws.Cells(r,2), 'NumberFormat', '@'))  # código como TEXTO (preserva zero à esquerda do SANEPAR)
        for col, value in [(2,cod),(5,desc),(7,tipo),(8,_origem_curta(orig)),(9,q),(10,un),(11,val)]:
            retry(lambda r=r, col=col, value=value: setattr(ws.Cells(r,col), 'Value', value))
    vtr = vt+n+1
    f = ws.Cells(vtr,16).Formula
    retry(lambda: setattr(ws.Cells(vtr,16), 'Formula', f + '+P%d' % hr))
    fo = ws.Cells(vtr,15).Formula
    if isinstance(fo, str) and fo.startswith('='):
        retry(lambda: setattr(ws.Cells(vtr,15), 'Formula', fo + '+O%d' % hr))

def export_pdfs(xl, wb, sheet, out_xlsx):
    stem=os.path.splitext(out_xlsx)[0]
    pdf_o=stem+' - ORÇAMENTO.pdf'; pdf_m=stem+' - MEMORIAL.pdf'
    o=wb.Worksheets(sheet); ps=o.PageSetup
    ps.Orientation=2; ps.Zoom=False; ps.FitToPagesWide=1; ps.FitToPagesTall=False; ps.CenterHorizontally=True
    ps.PrintTitleRows=''   # A2 repetia a linha 1 (título/logos cortado) em toda página -> só na 1ª
    o.Columns('G:H').Hidden=True    # ocultar TIPO (G) e ORIGEM (H) SÓ no PDF do orçamento
    o.ExportAsFixedFormat(0, pdf_o, 0)
    o.Columns('G:H').Hidden=False   # reexibir -> o .xlsx permanece completo
    m=wb.Worksheets('Memorial de Cálculo'); ps=m.PageSetup
    last=m.UsedRange.Rows.Count
    ps.PrintArea='$A$1:$G$%d'%last; ps.Orientation=2; ps.Zoom=False
    ps.FitToPagesWide=1; ps.FitToPagesTall=False; ps.PrintTitleRows='$4:$4'; ps.CenterHorizontally=True
    for a in ('LeftMargin','RightMargin','TopMargin','BottomMargin'): setattr(ps,a,xl.InchesToPoints(0.4))
    m.ExportAsFixedFormat(0, pdf_m, 0)
    return [pdf_o, pdf_m]

def load_cfg(path):
    spec=importlib.util.spec_from_file_location('cfg', path)
    mod=importlib.util.module_from_spec(spec); spec.loader.exec_module(mod)
    return mod

if __name__=='__main__':
    cfg=load_cfg(sys.argv[1])
    run(cfg)
