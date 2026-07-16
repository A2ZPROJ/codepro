# -*- coding: utf-8 -*-
"""
gerar_mapa.py — MÓDULO NEXUS "Mapa Geral" (SewerGEMS FlexTable -> shape + DXF).

Entrada: --config <json> com:
  { "excel": "...xlsx",            # FlexTable exportado do SewerGEMS
    "shpDir": "...\\SHP",          # (opcional) pasta com MEIO_FIO.shp + ALINHAMENTO.shp
    "saidaDir": "...",             # pasta de saída
    "nomeBase": "MAPA_GERAL" }     # (opcional) prefixo dos arquivos

Saída (última linha stdout = JSON):
  { "ok": true, "dxf": "...", "arquivos": [...], "pvs": N, "redes": N, "anticolisao": true }

Gera: <base>_PVS_COTAS.shp (cotas topo/fundo/prof), <base>_REDES.shp (L/i/mat/DN),
      <base>.dxf (blocos SES-POÇO-DE-VISITA/SES-TL + MLEADER anti-colisão + rótulos rede).
Formato dos rótulos = igual AJUSTARTEXTOPV/ROTULARALINHAMENTOS do NETLOAD.
"""
import sys, os, json, math, argparse
import openpyxl, shapefile, ezdxf
from ezdxf.math import Vec2
from ezdxf.render import mleader
from ezdxf import colors
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import mleader_placer as MP

# ── util ────────────────────────────────────────────────────────────
def numf(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if s == '' or s.startswith('(N/A') or s.lower() == 'none': return None
    s = s.replace('.', '').replace(',', '.')
    try: return float(s)
    except: return None
def br(v, dec):
    return "-" if v is None else ("{:." + str(dec) + "f}").format(v).replace('.', ',')
def emit(obj):
    print(json.dumps(obj, ensure_ascii=False)); sys.stdout.flush()

# ── SewerGEMS FlexTable (col idx fixos, verificados na R7) ───────────
MANHOLE = dict(lab=4, x=46, y=47, gnd=9, inv=11, dep=68)
BOUND = {'Outfall': (4, 36, 37, 'ETE'), 'Wet Well': (4, 41, 42, 'EEE'),
         'Pump': (4, 29, 30, 'EEE'), 'Pressure Junction': (4, 16, 17, 'PJ')}
CONDUIT = dict(ose=0, lab=4, mat=20, dn=31, L=124, i=135, start=126, stop=127)

def carregar_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    node = {}
    def load(sheet, il, ix, iy, ig=None, ii=None, idp=None, kind='PV'):
        if sheet not in wb.sheetnames: return
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            lab = row[il]
            if lab is None: continue
            lab = str(lab).strip()
            if lab == '': continue
            x, y = numf(row[ix]), numf(row[iy])
            if x is None or y is None: continue
            node[lab] = dict(x=x, y=y, kind='TL' if lab.upper().startswith('TL') else kind,
                             ground=numf(row[ig]) if ig is not None else None,
                             invert=numf(row[ii]) if ii is not None else None,
                             depth=numf(row[idp]) if idp is not None else None)
    load('Manhole', MANHOLE['lab'], MANHOLE['x'], MANHOLE['y'], MANHOLE['gnd'], MANHOLE['inv'], MANHOLE['dep'], 'PV')
    for sh, (il, ix, iy, k) in BOUND.items():
        load(sh, il, ix, iy, kind=k)
    tubes = []
    if 'Conduit' in wb.sheetnames:
        for row in wb['Conduit'].iter_rows(min_row=2, values_only=True):
            lab = row[CONDUIT['lab']]
            if lab is None: continue
            s = str(row[CONDUIT['start']]).strip() if row[CONDUIT['start']] else None
            t = str(row[CONDUIT['stop']]).strip() if row[CONDUIT['stop']] else None
            if s not in node or t not in node: continue
            ose = str(row[CONDUIT['ose']]).strip() if row[CONDUIT['ose']] else str(lab).strip()
            tubes.append(dict(label=ose, s=s, t=t, L=numf(row[CONDUIT['L']]), i=numf(row[CONDUIT['i']]),
                              mat=(str(row[CONDUIT['mat']]).strip() if row[CONDUIT['mat']] else ''), dn=numf(row[CONDUIT['dn']])))
    return node, tubes

def carregar_shapes(shpDir, manter_sentido=False):
    """Fonte alternativa: shapes do traçado ajustado (PVS.shp + REDES.shp) em vez do Excel FlexTable.
    PVS: LABEL, ELEV_GND(=CT), ELEV_INV(=CF), STRCTDPTH(=h/prof). REDES: LABEL(=OSE), MATERIAL, D(=DN mm),
    S(=i), START_/STOP_. L = comprimento planimétrico da própria geometria. EEEs_ETEs.shp = fronteira."""
    def rd(name):
        p = os.path.join(shpDir, name)
        r = shapefile.Reader(p)
        f = [x[0] for x in r.fields[1:]]
        gi = lambda n: (f.index(n) if n in f else None)
        return r, gi
    node = {}
    rp, gi = rd('PVS')
    iL, iG, iI, iD = gi('LABEL'), gi('ELEV_GND'), gi('ELEV_INV'), gi('STRCTDPTH')
    for sh, rec in zip(rp.shapes(), rp.records()):
        if iL is None or not sh.points: continue
        lab = str(rec[iL]).strip()
        if not lab: continue
        x, y = sh.points[0][0], sh.points[0][1]
        node[lab] = dict(x=x, y=y, kind='TL' if lab.upper().startswith('TL') else 'PV',
                         ground=numf(rec[iG]) if iG is not None else None,
                         invert=numf(rec[iI]) if iI is not None else None,
                         depth=numf(rec[iD]) if iD is not None else None)
    # fronteira EEE/ETE (opcional)
    if os.path.exists(os.path.join(shpDir, 'EEEs_ETEs.shp')):
        re_, gie = rd('EEEs_ETEs')
        idd = gie('DESCRIC')
        for k, (sh, rec) in enumerate(zip(re_.shapes(), re_.records())):
            if not sh.points: continue
            desc = (str(rec[idd]).strip() if idd is not None and rec[idd] else 'EEE')
            up = desc.upper()
            if 'PV' in up: continue  # PV-EX etc: não é fronteira
            kind = 'ETE' if 'ETE' in up else 'EEE'
            nm = '%s#%d' % (desc, k)
            node[nm] = dict(x=sh.points[0][0], y=sh.points[0][1], kind=kind, ground=None, invert=None, depth=None, disp=desc)
    # tubes
    rr, git = rd('REDES')
    iLab, iMat, iDN, iS, iSt, iSp = git('LABEL'), git('MATERIAL'), git('D'), git('S'), git('START_'), git('STOP_')
    iIvS, iIvP = git('INV_STRT'), git('INV_STP')  # cota de fundo (invert) em cada ponta → orienta o fluxo
    tubes = []
    def synth(name, pt):  # nó externo (EEE/PV-EX) que não está em PVS.shp → cria terminal a partir da geometria
        if name and name not in node:
            node[name] = dict(x=pt[0], y=pt[1], kind='EXT', ground=None, invert=None, depth=None)
    for sh, rec in zip(rr.shapes(), rr.records()):
        s = str(rec[iSt]).strip() if iSt is not None and rec[iSt] else None
        t = str(rec[iSp]).strip() if iSp is not None and rec[iSp] else None
        # FLUXO POR GRAVIDADE: montante = ponta com a cota de fundo (invert) MAIOR. Se o STOP tiver invert
        # maior que o START, o START/STOP está guardado ao contrário do caimento → inverte (mantém o fluxo real).
        ivs = numf(rec[iIvS]) if iIvS is not None else None
        ivp = numf(rec[iIvP]) if iIvP is not None else None
        if not manter_sentido and ivs is not None and ivp is not None and ivp > ivs + 1e-6:
            s, t = t, s  # reorienta pela invert (fonte com START/STOP não confiável). manter_sentido=True preserva o START/STOP da origem (ex. modelo SewerGEMS)
        g = sh.points
        if g and len(g) >= 2:  # sintetiza a ponta faltante pelo extremo da geometria mais distante do nó conhecido
            A, B = g[0], g[-1]
            def far(ref):
                da = (A[0] - ref['x']) ** 2 + (A[1] - ref['y']) ** 2
                db = (B[0] - ref['x']) ** 2 + (B[1] - ref['y']) ** 2
                return A if da > db else B
            if s in node and t not in node: synth(t, far(node[s]))
            elif t in node and s not in node: synth(s, far(node[t]))
            elif s not in node and t not in node: synth(s, A); synth(t, B)
        if s not in node or t not in node: continue
        L = None
        if g and len(g) >= 2:
            L = sum(math.hypot(g[k + 1][0] - g[k][0], g[k + 1][1] - g[k][1]) for k in range(len(g) - 1))
        ose = str(rec[iLab]).strip() if iLab is not None and rec[iLab] else ''
        tubes.append(dict(label=ose, s=s, t=t, L=L, i=numf(rec[iS]) if iS is not None else None,
                          mat=(str(rec[iMat]).strip() if iMat is not None and rec[iMat] else ''),
                          dn=numf(rec[iDN]) if iDN is not None else None,
                          geom=[(p[0], p[1]) for p in g] if g and len(g) >= 2 else None))
    return node, tubes

def segs_shp(path):
    r = shapefile.Reader(path); out = []
    for i in range(len(r)):  # shape-a-shape (r.shapes() em massa quebra em alguns .shp type 13 de CAD)
        try:
            sh = r.shape(i)
        except Exception:
            continue
        pts = sh.points; parts = list(sh.parts) + [len(pts)]
        for k in range(len(parts) - 1):
            seg = pts[parts[k]:parts[k + 1]]
            for a in range(len(seg) - 1):
                out.append((seg[a][0], seg[a][1], seg[a + 1][0], seg[a + 1][1]))
    return out

def achar_prj(shpDir):
    for nm in ("PVS.prj", "MEIO_FIO.prj", "ALINHAMENTO.prj"):
        p = os.path.join(shpDir or '', nm)
        if os.path.exists(p): return open(p).read()
    # SIRGAS 2000 / UTM 22S padrão
    return ('PROJCS["SIRGAS_2000_UTM_Zone_22S",GEOGCS["GCS_SIRGAS_2000",DATUM["D_SIRGAS_2000",'
            'SPHEROID["GRS_1980",6378137.0,298.257222101]],PRIMEM["Greenwich",0.0],UNIT["Degree",0.0174532925199433]],'
            'PROJECTION["Transverse_Mercator"],PARAMETER["False_Easting",500000.0],PARAMETER["False_Northing",10000000.0],'
            'PARAMETER["Central_Meridian",-51.0],PARAMETER["Scale_Factor",0.9996],PARAMETER["Latitude_Of_Origin",0.0],UNIT["Meter",1.0]]')

# ── geometria dos rótulos PV/TL ─────────────────────────────────────
CHAR_W, LINE_H, HPV, HRE = 0.70, 1.60, 1.2, 1.0
def pv_lines(lab, d):
    prof = d['depth']
    if prof is None and d['ground'] is not None and d['invert'] is not None:
        prof = d['ground'] - d['invert']
    return ["%s | h: %sm" % (lab, br(prof, 2)), "CT: %s" % br(d['ground'], 3), "CF: %s" % br(d['invert'], 3)]

def gerar(cfg):
    excel = cfg.get('excel'); saida = cfg['saidaDir']; base = cfg.get('nomeBase') or 'MAPA_GERAL'
    shpDir = cfg.get('shpDir') or ''
    os.makedirs(saida, exist_ok=True)
    if (cfg.get('fonte') or '').lower() == 'shapes':
        node, tubes = carregar_shapes(cfg.get('fonteShapes') or shpDir, bool(cfg.get('manter_sentido')))
    else:
        node, tubes = carregar_excel(excel)
    pvs = [(l, d) for l, d in node.items() if d['kind'] in ('PV', 'TL')]
    downstream = {}
    for tb in tubes:
        downstream.setdefault(tb['s'], tb['t']); downstream.setdefault(tb['t'], tb['s'])

    # ── SHAPES ──
    RAIO, COR, BARRA = 0.7531, 172, 1.045
    prj = achar_prj(shpDir)
    fpv = os.path.join(saida, base + "_PVS_COTAS")
    wp = shapefile.Writer(fpv, shapeType=shapefile.POINT, encoding='utf-8')
    wp.field('LABEL', 'C', 80); wp.field('COTA_TOPO', 'F', 19, 3); wp.field('COTA_FUNDO', 'F', 19, 3)
    wp.field('PROF', 'F', 19, 3); wp.field('TIPO', 'C', 10); wp.field('X', 'F', 19, 3); wp.field('Y', 'F', 19, 3)
    for lab, d in node.items():
        if d['kind'] not in ('PV', 'TL'): continue
        prof = d['depth']
        if prof is None and d['ground'] is not None and d['invert'] is not None: prof = d['ground'] - d['invert']
        wp.point(d['x'], d['y']); wp.record(lab, d['ground'], d['invert'], prof, d['kind'], d['x'], d['y'])
    wp.close()
    fre = os.path.join(saida, base + "_REDES")
    wt = shapefile.Writer(fre, shapeType=shapefile.POLYLINE, encoding='utf-8')
    wt.field('LABEL', 'C', 80); wt.field('MONTANTE', 'C', 80); wt.field('JUSANTE', 'C', 80)
    wt.field('L_M', 'F', 19, 2); wt.field('DECLIV', 'F', 19, 4); wt.field('MATERIAL', 'C', 40); wt.field('DN_MM', 'F', 19, 1)
    for tb in tubes:
        a, b = node[tb['s']], node[tb['t']]
        poly = tb.get('geom') or [(a['x'], a['y']), (b['x'], b['y'])]  # geometria fiel do original (sem deslocar)
        wt.line([[(p[0], p[1]) for p in poly]])
        wt.record(tb['label'], tb['s'], tb['t'], tb['L'], tb['i'], tb['mat'], tb['dn'])
    wt.close()
    for f in (fpv, fre):
        open(f + ".prj", "w").write(prj); open(f + ".cpg", "w").write("UTF-8")

    # ── anti-colisão dos rótulos PV ──
    seg_mf = seg_mu = []
    anti = False
    if shpDir and os.path.exists(os.path.join(shpDir, "MEIO_FIO.shp")):
        seg_mf = segs_shp(os.path.join(shpDir, "MEIO_FIO"))
    if shpDir and os.path.exists(os.path.join(shpDir, "ALINHAMENTO.shp")):
        seg_mu = segs_shp(os.path.join(shpDir, "ALINHAMENTO"))
    seg_re = [(node[tb['s']]['x'], node[tb['s']]['y'], node[tb['t']]['x'], node[tb['t']]['y']) for tb in tubes]
    items = []
    for lab, d in pvs:
        lines = pv_lines(lab, d); mx = max(len(s) for s in lines)
        items.append(dict(x=d['x'], y=d['y'], W=mx * HPV * CHAR_W, Ht=len(lines) * HPV * LINE_H))
    res = MP.place_labels(items, seg_mf, seg_mu, seg_re)
    anti = bool(seg_mf or seg_mu) or True

    # ── DXF ──
    doc = ezdxf.new('R2010', setup=True); msp = doc.modelspace(); doc.header['$INSUNITS'] = 6
    if '2SE-ARIAL' not in doc.styles: doc.styles.add('2SE-ARIAL', font='arial.ttf')
    for name, col in [('ALL-MOBI', COR), ('REDE', 3), ('REDE-SETA', 3), ('SES-TXT', 5), ('REDE-TXT', 7), ('EEE-ETE', 1)]:
        if name not in doc.layers: doc.layers.add(name, color=col)
    def add_fc(blk, r):
        blk.add_circle((0, 0), r, dxfattribs={'color': COR})
        h = blk.add_hatch(color=COR); p = h.paths.add_edge_path(); p.add_arc((0, 0), r, 0, 360)
    bpv = doc.blocks.new('SES-POÇO-DE-VISITA'); add_fc(bpv, RAIO)
    btl = doc.blocks.new('SES-TL'); add_fc(btl, RAIO); btl.add_line((BARRA, -BARRA), (BARRA, BARRA), dxfattribs={'color': COR})
    for lab, d in node.items():
        if d['kind'] in ('ETE', 'EEE'):
            msp.add_circle((d['x'], d['y']), RAIO * 1.6, dxfattribs={'layer': 'EEE-ETE'})
            msp.add_text(d.get('disp', lab), dxfattribs={'layer': 'EEE-ETE', 'height': HPV, 'style': '2SE-ARIAL'}).set_placement((d['x'] + 2, d['y'] + 2))
    for (lab, d), r in zip(pvs, res):
        nx, ny = d['x'], d['y']
        if d['kind'] == 'TL':
            oth = downstream.get(lab); rot = 0.0
            if oth and oth in node: rot = math.degrees(math.atan2(ny - node[oth]['y'], nx - node[oth]['x']))
            msp.add_blockref('SES-TL', (nx, ny), dxfattribs={'layer': 'ALL-MOBI', 'rotation': rot})
        else:
            msp.add_blockref('SES-POÇO-DE-VISITA', (nx, ny), dxfattribs={'layer': 'ALL-MOBI'})
        box = r['box']; cx = (box[0] + box[2]) / 2; cy = (box[1] + box[3]) / 2; dx, dy = r['dir']
        side = mleader.ConnectionSide.left if dx >= 0 else mleader.ConnectionSide.right  # attachment HORIZONTAL (só esquerda/direita)
        ml = msp.add_multileader_mtext('Standard')
        ml.set_content("\n".join(pv_lines(lab, d)), color=colors.BLUE, char_height=HPV,
                       alignment=mleader.TextAlignment.left, style='2SE-ARIAL')
        # attachment type horizontal + left/right = "Underline Top Line" (bottom_of_top_line_underline = 6)
        ml.set_connection_types(left=mleader.HorizontalConnection.bottom_of_top_line_underline,
                                right=mleader.HorizontalConnection.bottom_of_top_line_underline)
        ml.set_arrow_properties(size=0.02)
        ml.add_leader_line(side, [Vec2(nx, ny)])
        ml.build(insert=Vec2(cx, cy))
        ml.multileader.dxf.text_attachment_direction = 0  # 0 = horizontal
    for tb in tubes:
        a, c = node[tb['s']], node[tb['t']]  # montante, jusante (orientado pela invert em carregar_shapes)
        poly = tb.get('geom') or [(a['x'], a['y']), (c['x'], c['y'])]  # geometria REAL da rede (fiel, sem deslocar)
        msp.add_lwpolyline(poly, dxfattribs={'layer': 'REDE'})
        segd = [math.hypot(poly[k + 1][0] - poly[k][0], poly[k + 1][1] - poly[k][1]) for k in range(len(poly) - 1)]
        tot = sum(segd)
        if tot < 1e-6: continue
        half = tot / 2.0; acc = 0.0; mx, my = poly[0]
        for k in range(len(poly) - 1):
            if acc + segd[k] >= half:
                rr2 = (half - acc) / segd[k] if segd[k] > 1e-9 else 0.0
                mx = poly[k][0] + (poly[k + 1][0] - poly[k][0]) * rr2
                my = poly[k][1] + (poly[k + 1][1] - poly[k][1]) * rr2
                break
            acc += segd[k]
        dx, dy = c['x'] - a['x'], c['y'] - a['y']; ln = math.hypot(dx, dy)  # sentido da SETA = fluxo montante->jusante
        if ln < 1e-6: continue
        ux, uy = dx / ln, dy / ln; px, py = -uy, ux
        tip = (mx + ux * 0.8, my + uy * 0.8); bl = (mx - ux * 0.8 + px * 0.45, my - uy * 0.8 + py * 0.45); bb = (mx - ux * 0.8 - px * 0.45, my - uy * 0.8 - py * 0.45)
        h = msp.add_hatch(color=3, dxfattribs={'layer': 'REDE-SETA'}); h.paths.add_polyline_path([tip, bl, bb], is_closed=True)
        ang = math.degrees(math.atan2(dy, dx))
        if ang > 90 or ang <= -90: ang += 180
        ar = math.radians(ang); npx, npy = -math.sin(ar), math.cos(ar); off = 1.1
        mat_dn = ((tb['mat'] + " ") if tb['mat'] else "") + "DN " + br(tb['dn'], 0)
        up = "%s - L=%sm" % (tb['label'], br(tb['L'], 2)); dn_t = "i=%s - %s" % (br(tb['i'], 4), mat_dn)
        t1 = msp.add_mtext(up, dxfattribs={'layer': 'REDE-TXT', 'char_height': HRE, 'style': '2SE-ARIAL'})
        t1.set_location((mx + npx * off, my + npy * off), rotation=ang, attachment_point=5)
        t2 = msp.add_mtext(dn_t, dxfattribs={'layer': 'REDE-TXT', 'char_height': HRE, 'style': '2SE-ARIAL'})
        t2.set_location((mx - npx * off, my - npy * off), rotation=ang, attachment_point=5)
    fdxf = os.path.join(saida, base + ".dxf")
    doc.saveas(fdxf)

    arqs = [fpv + ".shp", fre + ".shp", fdxf]
    return dict(ok=True, dxf=fdxf, arquivos=arqs, pvs=len(pvs), redes=len(tubes),
                fronteira=sum(1 for d in node.values() if d['kind'] in ('ETE', 'EEE')),
                anticolisao=anti)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--config', required=True)
    a = ap.parse_args()
    try:
        cfg = json.load(open(a.config, encoding='utf-8'))
    except Exception as e:
        emit(dict(ok=False, erro="Config inválido: %s" % e)); return
    is_shapes = (cfg.get('fonte') or '').lower() == 'shapes'
    req = ('saidaDir',) if is_shapes else ('excel', 'saidaDir')
    for k in req:
        if not cfg.get(k):
            emit(dict(ok=False, erro="Falta '%s' no config." % k)); return
    if is_shapes:
        sd = cfg.get('fonteShapes') or cfg.get('shpDir') or ''
        if not os.path.exists(os.path.join(sd, 'PVS.shp')) or not os.path.exists(os.path.join(sd, 'REDES.shp')):
            emit(dict(ok=False, erro="PVS.shp/REDES.shp não encontrados em: %s" % sd)); return
    elif not os.path.exists(cfg['excel']):
        emit(dict(ok=False, erro="Excel não encontrado: %s" % cfg['excel'])); return
    try:
        emit(gerar(cfg))
    except Exception as e:
        import traceback
        emit(dict(ok=False, erro=str(e), trace=traceback.format_exc()[-1200:]))

if __name__ == '__main__':
    main()
