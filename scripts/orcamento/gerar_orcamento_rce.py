# -*- coding: utf-8 -*-
"""
Gerador GENERICO do orcamento de REDE COLETORA DE ESGOTO (RCE) - padrao 2S/SANEPAR.

Modulo de LINHA DE COMANDO que o Nexus (Electron) chama com QUALQUER projeto:
recebe o arquivo de OSEs (.xlsx) + os parametros (CLI ou --config JSON) e gera a
planilha orcamentaria completa (5 abas: Resumo / Memoria de Calculo / Orcamento /
Base de Precos / Referencias) com a MESMA logica/metodologia/visual ja validados em
Diamante do Norte (gerar_orcamento_diamante.py).

Metodologia (100% preservada do gerador de Diamante):
  - Divisao MND x VCA por PROFUNDIDADE do trecho (regra trecho-inteiro:
    max(prof. dos dois extremos) > 3,00 m -> VCA ; senao MND/cravacao Navigator).
  - Largura de vala VCA pela Tabela B.1 (NBR 17015 / MOS Mod.4) por profundidade.
  - Escavacao escalonada por faixa, escoramento graduado (descontinuo ate 3 m,
    continuo > 3 m), assentamento 009.002, berco de areia, reaterro, compactacao,
    recomposicao de passeio (lajota sextavada), carga e transporte do excedente.
  - Cavas de PV (lado parametrico) e TL (lado parametrico) sempre escavadas.
  - Cravacao MND - Navigator nos trechos MND.
  - Tubo PVC DN150 como ITEM SEPARADO de material (fornecimento ACCIONA).
  - BDI duplo (Acordao 2622/2013-TCU): Obras ~24,49% / Materiais ~12,99%.
  - Coluna ORIGEM (SANEPAR-MOS / SINAPI-Cotacao), aba Referencias, identidade 2S,
    secao de PARAMETROS editaveis (celulas amarelas alimentam o orcamento por formula).

LEITURA do arquivo de OSEs (robusta, generica):
  - Aba "RESUMO": secao DETALHAMENTO (OSE|Tipo|Nome|C.Topo|C.Fundo|Profundidade) ->
    PVs e TLs com profundidade real. Extensao total = soma da coluna Extensao.
  - Abas "OSE-NNN": tabela de trechos a partir da linha 11 (col A=nome do no,
    col F=distancia ao no anterior, col S="Prof. Vala"). Trecho = par de nos
    consecutivos distintos; VCA se max(S_a, S_b) > 3,00 m.
  - PIT = estaqueamento (nao orca; conta na profundidade do trecho).
  - DN lido de A1/A2 de cada aba OSE (default 150).

Saida (stdout, ultima linha): JSON com {"ok":true,"saida":..,"total":..,"custo_direto":..,
  "ext_mnd":..,"ext_vca":..,"n_pv":..,"n_tl":..,"n_ligacoes":..} para o Nexus capturar.

Autor: Claude Code (Opus) p/ Lucas Abdala / 2S Engenharia.
"""
import os, sys, json, math, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import cm_to_EMU
from openpyxl.comments import Comment


# =====================================================================
# 1) CONFIG / CLI
# =====================================================================
def build_config():
    """Le os parametros via --config <json> e/ou flags CLI. CLI sobrepoe JSON.
    Retorna um dict 'cfg' com todos os campos, aplicando defaults."""
    ap = argparse.ArgumentParser(
        description="Gerador generico de orcamento de Rede Coletora de Esgoto (RCE) - 2S/SANEPAR.")
    ap.add_argument("--config", help="Arquivo JSON com todos os campos (o Nexus passa este).")
    ap.add_argument("--oses", help="Caminho do .xlsx de OSEs (obrigatorio).")
    ap.add_argument("--saida", help="Caminho do .xlsx de saida (obrigatorio).")
    ap.add_argument("--ligacoes", type=int, help="N. de ligacoes prediais (informado manualmente).")
    ap.add_argument("--frentes", type=int, default=2, help="N. de frentes simultaneas (default 2).")
    ap.add_argument("--canteiro-larg", type=float, default=20.0)
    ap.add_argument("--canteiro-compr", type=float, default=30.0)
    ap.add_argument("--placa-area", type=float, default=3.0)
    ap.add_argument("--placas-frente", type=int, default=3)
    ap.add_argument("--fita-frente", type=float, default=50.0)
    ap.add_argument("--tapume-movel-frente", type=float, default=12.0)
    ap.add_argument("--obra", default="Rede Coletora de Esgoto")
    ap.add_argument("--municipio", default="")
    ap.add_argument("--contratante", default="ACCIONA")
    ap.add_argument("--projetista", default="2S Engenharia")
    ap.add_argument("--tubo-preco", type=float, default=85.78)
    ap.add_argument("--tubo-origem", default="SINAPI 12/2024 cod 41936")
    ap.add_argument("--dist-botafora", type=float, default=5.0, help="Dist. media bota-fora (km).")
    ap.add_argument("--asbuilt-qtd", type=int, default=None,
                    help="Qtd de as-built (default = numero de OSEs).")
    ap.add_argument("--base-precos", default="SANEPAR MOS 5a Ed. - JUN/2025")
    args = ap.parse_args()

    # defaults
    cfg = {
        "oses": None, "saida": None, "ligacoes": None, "frentes": 2,
        "canteiro_larg": 20.0, "canteiro_compr": 30.0, "placa_area": 3.0,
        "placas_frente": 3, "fita_frente": 50.0, "tapume_movel_frente": 12.0,
        "obra": "Rede Coletora de Esgoto", "municipio": "", "contratante": "ACCIONA",
        "projetista": "2S Engenharia", "tubo_preco": 85.78,
        "tubo_origem": "SINAPI 12/2024 cod 41936", "dist_botafora": 5.0,
        "asbuilt_qtd": None, "base_precos": "SANEPAR MOS 5a Ed. - JUN/2025",
    }

    # 1) JSON config (se houver)
    if args.config:
        with open(args.config, "r", encoding="utf-8") as f:
            jc = json.load(f)
        # aceita tanto chaves com '-' quanto '_'
        for k, v in jc.items():
            cfg[k.replace("-", "_")] = v

    # 2) CLI sobrepoe (apenas valores explicitos; argparse usa defaults -> so
    #    sobrepoe quando o usuario passou a flag E ela difere do default JSON).
    explicit = {a.split("=")[0].lstrip("-").replace("-", "_")
                for a in sys.argv[1:] if a.startswith("--")}
    cli_map = {
        "oses": args.oses, "saida": args.saida, "ligacoes": args.ligacoes,
        "frentes": args.frentes, "canteiro_larg": args.canteiro_larg,
        "canteiro_compr": args.canteiro_compr, "placa_area": args.placa_area,
        "placas_frente": args.placas_frente, "fita_frente": args.fita_frente,
        "tapume_movel_frente": args.tapume_movel_frente, "obra": args.obra,
        "municipio": args.municipio, "contratante": args.contratante,
        "projetista": args.projetista, "tubo_preco": args.tubo_preco,
        "tubo_origem": args.tubo_origem, "dist_botafora": args.dist_botafora,
        "asbuilt_qtd": args.asbuilt_qtd, "base_precos": args.base_precos,
    }
    for k, v in cli_map.items():
        if k in explicit and v is not None:
            cfg[k] = v

    # validacao
    if not cfg.get("oses"):
        ap.error("--oses (ou config.oses) e obrigatorio.")
    if not cfg.get("saida"):
        ap.error("--saida (ou config.saida) e obrigatorio.")
    if cfg.get("ligacoes") is None:
        ap.error("--ligacoes (ou config.ligacoes) e obrigatorio.")
    return cfg


# =====================================================================
# 2) LEITURA ROBUSTA DO ARQUIVO DE OSEs
# =====================================================================
def _num(v):
    try:
        if v is None:
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def larg_vala(p):
    """Tabela B.1 (NBR 17015 / MOS Mod.4) - largura de vala DN<=150 x profundidade."""
    if p < 2.0:
        return 0.65
    if p < 3.0:
        return 0.85
    return 1.05  # 3-4 m


def ler_oses(path):
    """Le o .xlsx de OSEs e devolve:
        pvs   : [(ose, nome, prof), ...]   profundidade real (col Profundidade do DETALHAMENTO)
        tls   : [(ose, nome, prof), ...]
        segs  : [{ose,a,b,ext,s0,s1,pm,larg}, ...]  trechos VCA (max prof dos extremos > 3 m)
        ext   : extensao total da rede (m) por soma da coluna Extensao do RESUMO
        n_ose : numero de OSEs
        dns   : set de DN encontrados (str)
        ext_por_dn : {dn: metros}
    Robusta: nao assume numero fixo de OSEs nem DN unico.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- aba RESUMO ----
    resumo = None
    for sn in wb.sheetnames:
        if sn.strip().upper().startswith("RESUMO"):
            resumo = wb[sn]
            break
    if resumo is None:
        raise RuntimeError("Aba RESUMO nao encontrada no arquivo de OSEs.")

    # extensao total + numero de OSEs (secao topo: OSE | Extensao | PV | TL | PIT)
    ext_total = 0.0
    n_ose = 0
    # localiza o cabecalho 'OSE' / 'Extensao'
    top_hdr = None
    for r in range(1, resumo.max_row + 1):
        a = resumo.cell(r, 1).value
        b = resumo.cell(r, 2).value
        if isinstance(a, str) and a.strip().upper() == "OSE" and isinstance(b, str) \
           and "EXTENS" in b.strip().upper():
            top_hdr = r
            break
    if top_hdr:
        r = top_hdr + 1
        while r <= resumo.max_row:
            a = resumo.cell(r, 1).value
            if isinstance(a, str) and a.strip().upper().startswith("OSE-"):
                e = _num(resumo.cell(r, 2).value)
                if e is not None:
                    ext_total += e
                    n_ose += 1
                r += 1
            elif a is None:
                # uma linha em branco pode separar; verifica se acabou
                nxt = resumo.cell(r + 1, 1).value
                if not (isinstance(nxt, str) and nxt.strip().upper().startswith("OSE-")):
                    break
                r += 1
            else:
                break

    # secao DETALHAMENTO (OSE | Tipo | Nome | C.Topo | C.Fundo | Profundidade)
    det_hdr = None
    for r in range(1, resumo.max_row + 1):
        a = resumo.cell(r, 1).value
        b = resumo.cell(r, 2).value
        if isinstance(a, str) and a.strip().upper() == "OSE" and isinstance(b, str) \
           and b.strip().upper() == "TIPO":
            det_hdr = r
            break
    pvs, tls = [], []
    if det_hdr:
        r = det_hdr + 1
        while r <= resumo.max_row:
            ose = resumo.cell(r, 1).value
            if not ose:
                r += 1
                # para se as proximas linhas tambem vazias
                if all(resumo.cell(r + k, 1).value is None for k in range(0, 3)):
                    break
                continue
            tipo = (resumo.cell(r, 2).value or "")
            nome = resumo.cell(r, 3).value
            prof = _num(resumo.cell(r, 6).value)
            tipo = str(tipo).strip().upper()
            if tipo == "PV" and prof is not None:
                pvs.append((str(ose).strip(), str(nome).strip(), round(prof, 3)))
            elif tipo == "TL" and prof is not None:
                tls.append((str(ose).strip(), str(nome).strip(), round(prof, 3)))
            r += 1

    # ---- abas OSE-NNN : DN + trechos VCA ----
    ext_por_dn = {}
    dns = set()
    segs = []
    for sn in wb.sheetnames:
        if not sn.strip().upper().startswith("OSE-"):
            continue
        ws = wb[sn]
        # DN em A1/A2 (ex.: A1='150', A2='DN 150')
        dn = None
        for cell in ("A1", "A2"):
            v = ws[cell].value
            if v is None:
                continue
            s = "".join(ch for ch in str(v) if ch.isdigit())
            if s:
                dn = s
                break
        if dn is None:
            dn = "150"
        dns.add(dn)
        # comprimento da OSE (C6) -> soma por DN (fallback p/ ext total se RESUMO falhar)
        comp = _num(ws["C6"].value) or 0.0
        ext_por_dn[dn] = ext_por_dn.get(dn, 0.0) + comp

        # tabela de trechos a partir da linha 11: col A=nome, col F(6)=dist, col S(19)=Prof.Vala
        raw = []
        r = 11
        while r <= ws.max_row:
            nm = ws.cell(r, 1).value
            if nm is None:
                break
            raw.append((str(nm).strip(), _num(ws.cell(r, 6).value), _num(ws.cell(r, 19).value)))
            r += 1
        # colapsa nos duplicados consecutivos (linha principal + linha DEGRAU) -> mantem a 1a
        nodes = []
        for nm, F, S in raw:
            if nodes and nodes[-1][0] == nm:
                continue
            nodes.append((nm, F, S))
        # trechos = pares consecutivos; VCA se max(S_a,S_b) > 3,00
        for i in range(1, len(nodes)):
            a, Fa, Sa = nodes[i - 1]
            b, Fb, Sb = nodes[i]
            if Sa is None or Sb is None:
                continue
            ext = Fb if Fb is not None else 0.0
            if ext <= 0:
                continue
            if max(Sa, Sb) > 3.00:
                pm = (Sa + Sb) / 2.0
                # ext/pm guardados a 4 casas (igual a extracao validada de Diamante);
                # a Memoria/calculo arredonda a 3 casas como o gerador original.
                segs.append(dict(ose=sn, a=a, b=b, ext=round(ext, 4),
                                 s0=round(Sa, 4), s1=round(Sb, 4),
                                 pm=round(pm, 4), larg=larg_vala(round(pm, 4))))

    # fallback: se RESUMO nao deu extensao, usa soma dos comprimentos das OSEs
    if ext_total <= 0:
        ext_total = sum(ext_por_dn.values())
        n_ose = sum(1 for sn in wb.sheetnames if sn.strip().upper().startswith("OSE-"))

    return dict(pvs=pvs, tls=tls, segs=segs, ext=round(ext_total, 2),
                n_ose=n_ose, dns=dns, ext_por_dn=ext_por_dn)


# =====================================================================
# 3) GERACAO DA PLANILHA (logica/visual 100% do gerador de Diamante)
# =====================================================================
def gerar(cfg):
    data = ler_oses(cfg["oses"])
    PVS = [(nm, p) for ose, nm, p in data["pvs"]]
    TLS = [(nm, p) for ose, nm, p in data["tls"]]
    SEGS = data["segs"]
    EXT = data["ext"]
    N_OSE = data["n_ose"]
    N_LIG = int(cfg["ligacoes"])

    _ext_vca_raw = sum(s["ext"] for s in SEGS)
    EXT_VCA = round(_ext_vca_raw, 2)
    EXT_MND = round(EXT - _ext_vca_raw, 2)

    OUT = cfg["saida"]
    # logos (mesmos caminhos do gerador de Diamante)
    LOGO_2S = r"C:\Users\lcabd\OneDrive - A2Z Projetos\INSTALADORES - PROGRAMAÇÃO\PROJETOS\codepro\assets\logo-2s.png"
    LOGO_ACC = r"C:\Users\lcabd\OneDrive - A2Z Projetos\INSTALADORES - PROGRAMAÇÃO\PROJETOS\codepro\assets\logo-acciona.png"

    # ===================== PALETA 2S (refinada - sobria) =====================
    C_GRAF = "303030"; C_GRPMD = "3A3F47"; C_RED = "A11312"; C_SUB = "E5E7EB"; C_COTAR = "FFE08A"
    C_ZEBRA = "F5F6F8"; C_WHITE = "FFFFFF"; C_INFO = "ECEDEF"

    FN_HDR = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
    FN_GRP = Font(name="Calibri", bold=True, color=C_WHITE, size=10.5)
    FN_SUB = Font(name="Calibri", bold=True, color=C_GRAF, size=10)
    FN_TOT = Font(name="Calibri", bold=True, color=C_WHITE, size=12)
    FN_TITLE = Font(name="Calibri", bold=True, color=C_WHITE, size=15)
    FN_SUBTL = Font(name="Calibri", bold=True, color=C_WHITE, size=10)
    FN_NORM = Font(name="Calibri", size=9.5)
    FN_BOLD = Font(name="Calibri", bold=True, size=9.5)
    FN_LBL = Font(name="Calibri", bold=True, size=10, color=C_GRAF)
    FN_NOTE = Font(name="Calibri", italic=True, size=8.5, color="7A5C00")

    F_GRAF = PatternFill("solid", fgColor=C_GRAF); F_RED = PatternFill("solid", fgColor=C_RED)
    F_GRP = PatternFill("solid", fgColor=C_GRPMD)
    F_SUB = PatternFill("solid", fgColor=C_SUB); F_COTAR = PatternFill("solid", fgColor=C_COTAR)
    F_ZEBRA = PatternFill("solid", fgColor=C_ZEBRA); F_INFO = PatternFill("solid", fgColor=C_INFO)
    F_REDLINE = PatternFill("solid", fgColor=C_RED)

    thin = Side(style="thin", color="D9DADC"); med = Side(style="medium", color=C_GRAF)

    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left", vertical="center", wrap_text=True)
    R = Alignment(horizontal="right", vertical="center")
    CC = Alignment(horizontal="center", vertical="center")
    LC = Alignment(horizontal="left", vertical="center")

    FMT_RS = '_-R$ * #,##0.00_-;-R$ * #,##0.00_-;_-R$ * "-"??_-'
    FMT_N2 = '#,##0.00'; FMT_N3 = '#,##0.000'; FMT_PCT = '0.00%'; FMT_PCT1 = '0.0%'

    def box(ws, r1, c1, r2, c2, side=thin):
        bd = Border(left=side, right=side, top=side, bottom=side)
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(r, c).border = bd

    # ===================== DADOS DA OBRA =====================
    muni = cfg["municipio"] or ""
    OBRA = {
        "titulo": cfg["obra"],
        "contratante": cfg["contratante"],
        "projetista": cfg["projetista"],
        "base": cfg["base_precos"],
        "ext": EXT,
        "n_ose": N_OSE,
        "municipio": muni,
    }

    # premissas / parametros
    DIST_BOTAFORA = float(cfg["dist_botafora"])
    QTD_ASBUILT = int(cfg["asbuilt_qtd"]) if cfg["asbuilt_qtd"] is not None else N_OSE
    PV_LADO = 1.40; PV_DIA = 0.80; TL_LADO = 0.75; TL_DIA = 0.15
    ESCOR_MIN = 1.25; BERCO_ESP = 0.10; ESCOR_CONT_MIN = 3.00; DN_TUBO = 0.150

    # ===================== PRECOS MOS =====================
    TUBO_PRECO = float(cfg["tubo_preco"])
    TUBO_ORIGEM_TXT = cfg["tubo_origem"]
    P = {
        "placa": ("001.005.001", "Placa de obra em chapa metálica", "m2", 552.66),
        "mobiliz": ("006.002.001", "Mobilização e desmobilização (obras)", "ud", 351.69),
        "locacao": ("002.001.039", "Locação e nivelamento p/ assentamento (OSE)", "m", 2.45),
        "asbuilt": ("002.007.005", "Cadastro de obra localizada (As Built)", "ud", 154.12),
        "mnd_150": ("016.012.004", "Cravação MND - Navigator DN150", "m", 139.73),
        "pv_base": ("009.015.001", "PV Tipo A DN800 - parcela até 1,00 m de prof.", "ud", 1346.10),
        "pv_acr": ("009.015.002", "PV Tipo A DN800 - acréscimo por m acima de 1,00 m", "m", 638.92),
        "tl": ("009.013.002", "Terminal de Limpeza TL PVC JE DN150", "ud", 59.75),
        "lig": ("017.019.002", "Ligação predial de esgoto - passeio pavimentado", "ud", 482.78),
        "til": ("009.012.001", "Tubo de Inspeção e Limpeza TIL PVC JE DN100", "ud", 55.50),
        "berco": ("009.029.002", "Berço / embasamento de areia", "m3", 175.49),
        "esc_2": ("004.002.001", "Escavação mec. - prof. 0 < h <= 2,00 m", "m3", 18.43),
        "esc_4": ("004.002.002", "Escavação mec. - prof. 2,01 a 4,00 m", "m3", 19.90),
        "reat": ("004.013.002", "Reaterro mecânico", "m3", 2.45),
        "comp": ("004.014.002", "Compactação mecânica de reaterro", "m3", 7.18),
        "escor_d": ("005.001.002", "Escoramento descontínuo", "m2", 47.47),
        "escor_c": ("005.001.003", "Escoramento contínuo (prof > 3,00 m)", "m2", 71.27),
        "assent": ("009.002.002", "Assentamento de tubo PVC esgoto DN150", "m", 9.19),
        "passeio": ("010.003.001", "Recomposição de passeio - lajota sextavada de concreto", "m2", 57.50),
        "carga": ("004.018.001", "Carga de material (solo exceto rocha)", "m3", 3.93),
        "transp": ("004.019.002", "Transporte/bota-fora (caminho de serviço)", "m3xkm", 4.69),
        "tapume": ("001.004.003", "Tapume de chapa de madeira compensada (canteiro)", "m", 135.68),
        "placa_adv": ("003.006.001", "Placa de advertência de trânsito 1,00 x 1,00 m", "ud", 151.33),
        "fita": ("003.005.001", "Fita plástica de sinalização", "m", 0.30),
        "tap_mov": ("003.005.003", "Tapume móvel descontínuo (proteção de segurança)", "m", 4.00),
        "tubo": ("SINAPI 41936", "Fornecimento de tubo PVC esgoto DN150 (coletor JEI - NBR 7362)", "m", TUBO_PRECO),
    }

    ORIGEM_MOS = "SANEPAR - MOS JUN/25"
    ORIGEM_COT = TUBO_ORIGEM_TXT

    def origem_de(cod):
        return ORIGEM_COT if str(cod).upper().startswith("SINAPI") or str(cod).strip() == "" else ORIGEM_MOS

    # ===================== WORKBOOK =====================
    wb = openpyxl.Workbook()
    WS_RES, WS_MEM, WS_ORC, WS_BASE = "Resumo", "Memória de Cálculo", "Orçamento", "Base de Preços (MOS)"

    titulo_obra = OBRA["titulo"]
    muni_sufixo = (" - " + muni) if muni else ""

    # =====================================================================
    # ABA: MEMORIA DE CALCULO
    # =====================================================================
    mem = wb.active; mem.title = WS_MEM; mem.sheet_view.showGridLines = False

    mem.merge_cells("A1:I1")
    mem["A1"] = f"MEMÓRIA DE CÁLCULO  -  {titulo_obra}  (MND + VCA por profundidade)"
    mem["A1"].font = FN_TITLE; mem["A1"].fill = F_GRAF; mem["A1"].alignment = CC
    mem.row_dimensions[1].height = 26

    mem.merge_cells("A2:I2")
    mem["A2"] = ("Rede repartida por PROFUNDIDADE do trecho (regra trecho-inteiro: max prof dos extremos > "
                 "3,00 m -> VCA, senão MND). MND = cravação Navigator (sem vala). VCA = vala aberta: escavação "
                 "escalonada + escoramento (contínuo se prof > 3 m) + assentamento 009.002 + berço + reaterro + "
                 "recomposição de passeio. Cavas de PV (lado 1,40 m) e TL (lado 0,75 m) sempre escavadas; "
                 "escoramento da cava/vala se prof > 1,25 m; berço de areia 0,10 m. Profundidades = valores reais "
                 "do projeto (col. 'Prof. Vala' do arquivo fonte).")
    mem["A2"].font = FN_NOTE; mem["A2"].alignment = LC
    mem.row_dimensions[2].height = 34

    # ---- premissas editaveis (amarelas) topo ----
    mem["A4"] = "Dist. média bota-fora (km):"; mem["A4"].font = FN_LBL; mem["A4"].alignment = R
    mem.merge_cells("A4:C4")
    mem["D4"] = DIST_BOTAFORA; mem["D4"].fill = F_COTAR; mem["D4"].number_format = FMT_N2
    mem["D4"].alignment = CC; mem["D4"].font = FN_BOLD
    P_DIST = "'%s'!$D$4" % WS_MEM
    mem["E4"] = "Folga cava PV (lado, m):"; mem["E4"].font = FN_LBL; mem["E4"].alignment = R
    mem.merge_cells("E4:G4")
    mem["H4"] = PV_LADO; mem["H4"].fill = F_COTAR; mem["H4"].number_format = FMT_N2
    mem["H4"].alignment = CC; mem["H4"].font = FN_BOLD
    P_PVLADO = "'%s'!$H$4" % WS_MEM
    box(mem, 4, 4, 4, 4); box(mem, 4, 8, 4, 8)

    # ---- SECAO 0: PARAMETROS DA OBRA (editaveis) ----
    PR0 = 6
    mem.merge_cells(start_row=PR0, start_column=1, end_row=PR0, end_column=10)
    hc = mem.cell(PR0, 1, "0. PARÂMETROS DA OBRA  (editáveis - células amarelas alimentam o Orçamento por fórmula)")
    hc.font = FN_HDR; hc.fill = F_GRAF; hc.alignment = L
    for col in range(1, 11):
        mem.cell(PR0, col).fill = F_GRAF
    mem.row_dimensions[PR0].height = 20

    def param(row, label, value, fmt=FMT_N2, formula=None, note=None):
        mem.cell(row, 1, label).font = FN_LBL; mem.cell(row, 1).alignment = R
        mem.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = mem.cell(row, 4)
        if formula is not None:
            c.value = formula
        else:
            c.value = value
        c.fill = F_COTAR; c.number_format = fmt; c.alignment = CC; c.font = FN_BOLD
        box(mem, row, 4, row, 4)
        if note:
            # nota e TEXTO; se comecar com '=' o openpyxl/Excel a trataria como formula
            # (gerando #NAME?). Forca string (data_type 's').
            nc = mem.cell(row, 5, note)
            if isinstance(note, str) and note.startswith("="):
                nc.data_type = "s"
            nc.font = FN_NOTE; nc.alignment = LC
            mem.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
        return f"'{WS_MEM}'!$D${row}"

    pr = PR0 + 1
    P_NFRENTES = param(pr, "N. de frentes de serviço (N_FRENTES):", int(cfg["frentes"]), fmt='0',
                       note="Frentes de serviço simultâneas. Alimenta placas de advertência, fita e tapume móvel.")
    pr += 1
    P_CANT_LARG = param(pr, "Canteiro - largura (m):", float(cfg["canteiro_larg"]))
    CLARG_ROW = pr; pr += 1
    P_CANT_COMPR = param(pr, "Canteiro - comprimento (m):", float(cfg["canteiro_compr"]))
    CCOMPR_ROW = pr; pr += 1
    P_CANT_PERIM = param(pr, "Canteiro - PERÍMETRO (m):", None,
                         formula=f"=2*(D{CLARG_ROW}+D{CCOMPR_ROW})",
                         note="=2*(LARG+COMPR). Alimenta o tapume fixo (cod 001.004.003).")
    pr += 1
    P_PLACA_AREA = param(pr, "Placa de obra - ÁREA (m2):", float(cfg["placa_area"]),
                         note="Modelo 2,0 x 1,5 m = 3,0 m2. Alimenta a placa de obra (cod 001.005.001, m2).")
    pr += 1
    P_PLACAS_PF = param(pr, "Placas de advertência por frente (ud):", int(cfg["placas_frente"]), fmt='0',
                        note="Premissa por frente. QUANT = N_FRENTES x este valor.")
    pr += 1
    P_FITA_PF = param(pr, "Fita de sinalização por frente (m):", float(cfg["fita_frente"]),
                      note="Premissa por frente. QUANT = N_FRENTES x este valor.")
    pr += 1
    P_TAPMOV_PF = param(pr, "Tapume móvel por frente (m - perim. poço MND):", float(cfg["tapume_movel_frente"]),
                        note="Premissa por frente (perímetro do poço MND). QUANT = N_FRENTES x este valor.")
    pr += 1
    box(mem, PR0, 1, pr - 1, 10)

    # ---- SECAO 1: POCOS DE VISITA ----
    r = pr + 1
    mem.cell(r, 1, f"1. POÇOS DE VISITA  ({len(PVS)} un. - cava a céu aberto, sempre escavada)").font = FN_LBL
    r += 1
    hdr = ["PV", "Prof. (m)", "Lado cava (m)", "Vol. cava (m3)", "Vol. PV (m3)",
           "Reaterro (m3)", "Berço (m3)", "Escor. (m2)", "Acrésc. (m>1,00)", "Faixa escav."]
    PVH = r
    for j, h in enumerate(hdr, 1):
        cc = mem.cell(r, j, h); cc.font = FN_HDR; cc.fill = F_GRAF; cc.alignment = C
    mem.row_dimensions[r].height = 34
    r += 1
    PV_FIRST = r
    for i, (nm, pf) in enumerate(PVS):
        rr = r + i
        mem.cell(rr, 1, nm)
        mem.cell(rr, 2, round(pf, 3))
        mem.cell(rr, 3).value = f"={P_PVLADO}"
        mem.cell(rr, 4).value = f"=C{rr}^2*B{rr}"
        mem.cell(rr, 5).value = f"=PI()/4*{PV_DIA}^2*B{rr}"
        mem.cell(rr, 6).value = f"=MAX(0,D{rr}-E{rr})"
        mem.cell(rr, 7).value = f"=C{rr}^2*{BERCO_ESP}"
        mem.cell(rr, 8).value = f'=IF(B{rr}>{ESCOR_MIN},4*C{rr}*B{rr},0)'
        mem.cell(rr, 9).value = f"=MAX(0,B{rr}-1)"
        mem.cell(rr, 10).value = f'=IF(B{rr}<=2,"<=2m","2-4m")'
    PV_LAST = r + len(PVS) - 1
    for rr in range(PV_FIRST, PV_LAST + 1):
        for col in (2, 3, 4, 5, 6, 7, 8, 9):
            mem.cell(rr, col).number_format = FMT_N3
        if (rr - PV_FIRST) % 2 == 1:
            for col in range(1, 11):
                mem.cell(rr, col).fill = F_ZEBRA
        for col in range(1, 11):
            mem.cell(rr, col).alignment = CC; mem.cell(rr, col).font = FN_NORM
    box(mem, PVH, 1, PV_LAST, 10)
    rpv = PV_LAST + 1
    mem.cell(rpv, 1, "TOTAL PV")
    for col, let in [(4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H"), (9, "I")]:
        c = mem.cell(rpv, col); c.value = f"=SUM({let}{PV_FIRST}:{let}{PV_LAST})"; c.number_format = FMT_N3
    for col in range(1, 11):
        mem.cell(rpv, col).fill = F_SUB; mem.cell(rpv, col).alignment = CC; mem.cell(rpv, col).font = FN_SUB
    box(mem, rpv, 1, rpv, 10)
    PV_TOT = rpv

    # ---- SECAO 2: TERMINAIS DE LIMPEZA ----
    r = rpv + 2
    mem.cell(r, 1, f"2. TERMINAIS DE LIMPEZA  ({len(TLS)} un. - cava menor, lado 0,75 m)").font = FN_LBL
    r += 1
    hdr = ["TL", "Prof. (m)", "Lado cava (m)", "Vol. cava (m3)", "Vol. TL (m3)",
           "Reaterro (m3)", "Escor. (m2)", "Faixa escav."]
    TLH = r
    for j, h in enumerate(hdr, 1):
        cc = mem.cell(r, j, h); cc.font = FN_HDR; cc.fill = F_GRAF; cc.alignment = C
    mem.row_dimensions[r].height = 30
    r += 1
    TL_FIRST = r
    for i, (nm, pf) in enumerate(TLS):
        rr = r + i
        mem.cell(rr, 1, nm)
        mem.cell(rr, 2, round(pf, 3))
        mem.cell(rr, 3, TL_LADO)
        mem.cell(rr, 4).value = f"=C{rr}^2*B{rr}"
        mem.cell(rr, 5).value = f"=PI()/4*{TL_DIA}^2*B{rr}"
        mem.cell(rr, 6).value = f"=MAX(0,D{rr}-E{rr})"
        mem.cell(rr, 7).value = f'=IF(B{rr}>{ESCOR_MIN},4*C{rr}*B{rr},0)'
        mem.cell(rr, 8).value = f'=IF(B{rr}<=2,"<=2m","2-4m")'
    TL_LAST = r + len(TLS) - 1
    for rr in range(TL_FIRST, TL_LAST + 1):
        for col in (2, 3, 4, 5, 6, 7):
            mem.cell(rr, col).number_format = FMT_N3
        if (rr - TL_FIRST) % 2 == 1:
            for col in range(1, 9):
                mem.cell(rr, col).fill = F_ZEBRA
        for col in range(1, 9):
            mem.cell(rr, col).alignment = CC; mem.cell(rr, col).font = FN_NORM
    box(mem, TLH, 1, TL_LAST, 8)
    rtl = TL_LAST + 1
    mem.cell(rtl, 1, "TOTAL TL")
    for col, let in [(4, "D"), (5, "E"), (6, "F"), (7, "G")]:
        c = mem.cell(rtl, col); c.value = f"=SUM({let}{TL_FIRST}:{let}{TL_LAST})"; c.number_format = FMT_N3
    for col in range(1, 9):
        mem.cell(rtl, col).fill = F_SUB; mem.cell(rtl, col).alignment = CC; mem.cell(rtl, col).font = FN_SUB
    box(mem, rtl, 1, rtl, 8)
    TL_TOT = rtl

    # ---- SECAO 2B: TRECHOS VCA ----
    r = rtl + 2
    n_vca_oses = len({s["ose"] for s in SEGS})
    mem.cell(r, 1, (f"2B. TRECHOS EM VALA COMUM ABERTA - VCA  ({len(SEGS)} trechos / "
                    f"{EXT_VCA:.2f} m em {n_vca_oses} OSEs - prof. do trecho > 3,00 m)")).font = FN_LBL
    r += 1
    hdr = ["Trecho (OSE: mont.-jus.)", "Ext. (m)", "Prof. méd (m)", "Larg. vala (m)",
           "Escav. (m3)", "Escor. (m2)", "Tipo escor.", "Berço (m3)", "Reaterro (m3)", "Passeio (m2)"]
    VH = r
    for j, h in enumerate(hdr, 1):
        cc = mem.cell(r, j, h); cc.font = FN_HDR; cc.fill = F_GRAF; cc.alignment = C
    mem.row_dimensions[r].height = 34
    r += 1
    V_FIRST = r
    if SEGS:
        for i, s in enumerate(SEGS):
            rr = r + i
            mem.cell(rr, 1, f"{s['ose']}: {s['a']}-{s['b']}").alignment = L
            mem.cell(rr, 2, round(s["ext"], 3))
            mem.cell(rr, 3, round(s["pm"], 3))
            mem.cell(rr, 4).value = f'=IF(C{rr}<2,0.65,IF(C{rr}<3,0.85,1.05))'
            mem.cell(rr, 5).value = f"=D{rr}*C{rr}*B{rr}"
            mem.cell(rr, 6).value = f'=IF(C{rr}>{ESCOR_MIN},2*C{rr}*B{rr},0)'
            mem.cell(rr, 7).value = (f'=IF(C{rr}<={ESCOR_MIN},"-",'
                                     f'IF(C{rr}>{ESCOR_CONT_MIN},"CONTINUO","DESCONT"))')
            mem.cell(rr, 8).value = f"=D{rr}*{BERCO_ESP}*B{rr}"
            mem.cell(rr, 9).value = f"=MAX(0,E{rr}-PI()/4*{DN_TUBO}^2*B{rr}-H{rr})"
            mem.cell(rr, 10).value = f"=D{rr}*B{rr}"
        V_LAST = r + len(SEGS) - 1
    else:
        # sem trechos VCA: cria uma linha-zero para as formulas de SUM funcionarem
        mem.cell(r, 1, "(sem trechos VCA)").alignment = L
        for col in (2, 3, 5, 6, 8, 9, 10):
            mem.cell(r, col, 0)
        mem.cell(r, 4, 0); mem.cell(r, 7, "-")
        V_LAST = r
    for rr in range(V_FIRST, V_LAST + 1):
        for col in (2, 3, 4, 5, 6, 8, 9, 10):
            mem.cell(rr, col).number_format = FMT_N3
        if (rr - V_FIRST) % 2 == 1:
            for col in range(1, 11):
                mem.cell(rr, col).fill = F_ZEBRA
        for col in range(1, 11):
            if col != 1:
                mem.cell(rr, col).alignment = CC
            mem.cell(rr, col).font = FN_NORM
    box(mem, VH, 1, V_LAST, 10)
    rv = V_LAST + 1
    mem.cell(rv, 1, "TOTAL VCA")
    for col, let in [(2, "B"), (5, "E"), (6, "F"), (8, "H"), (9, "I"), (10, "J")]:
        c = mem.cell(rv, col); c.value = f"=SUM({let}{V_FIRST}:{let}{V_LAST})"; c.number_format = FMT_N3
    for col in range(1, 11):
        mem.cell(rv, col).fill = F_SUB; mem.cell(rv, col).alignment = CC; mem.cell(rv, col).font = FN_SUB
    mem.cell(rv, 1).alignment = R
    box(mem, rv, 1, rv, 10)
    VCA_TOT = rv

    # ---- SECAO 3: REDE E LIGACOES ----
    r = VCA_TOT + 2
    mem.cell(r, 1, "3. REDE E LIGAÇÕES").font = FN_LBL
    r += 1
    for j, h in enumerate(["Item", "Valor", "Unid."], 1):
        cc = mem.cell(r, j, h); cc.font = FN_HDR; cc.fill = F_GRAF; cc.alignment = C
    r += 1
    RDF = r
    _rbox = {"r": r}

    def kv(label, val, unit):
        rr = _rbox["r"]
        mem.cell(rr, 1, label).font = FN_NORM; mem.cell(rr, 1).alignment = L
        c = mem.cell(rr, 2, val); c.font = FN_BOLD; c.alignment = R; c.number_format = FMT_N2
        mem.cell(rr, 3, unit).font = FN_NORM; mem.cell(rr, 3).alignment = CC
        _rbox["r"] = rr + 1
        return rr

    EXTTOT_ROW = kv("Extensão total da rede DN150", EXT, "m")
    EXT_ROW = kv("  do qual em MND (cravação Navigator)", EXT_MND, "m")
    EXTVCA_ROW = kv("  do qual em VCA (vala, prof. trecho > 3 m)", EXT_VCA, "m")
    NOSE_ROW = kv("Número de OSEs", N_OSE, "ud")
    NLIG_ROW = kv("Ligações prediais (informado)", N_LIG, "ud")
    mem.cell(NLIG_ROW, 2).fill = F_COTAR
    NPV_ROW = kv("Número de PVs", len(PVS), "ud")
    NTL_ROW = kv("Número de TLs", len(TLS), "ud")
    r = _rbox["r"]
    for rr in range(RDF, r):
        if (rr - RDF) % 2 == 1:
            for col in range(1, 4):
                mem.cell(rr, col).fill = F_ZEBRA
    box(mem, RDF - 1, 1, r - 1, 3)

    # ---- SECAO 4: RESUMO DE QUANTITATIVOS ----
    r += 1
    mem.cell(r, 1, "4. RESUMO DE QUANTITATIVOS  (alimenta a aba Orçamento)").font = FN_LBL
    r += 1
    for j, h in enumerate(["Quantitativo", "Unid.", "Valor"], 1):
        cc = mem.cell(r, j, h); cc.font = FN_HDR; cc.fill = F_GRAF; cc.alignment = C
    QHEAD = r; r += 1; QF = r
    Q = {}
    _qbox = {"r": r}

    def qrow(label, unit, formula):
        rr = _qbox["r"]
        mem.cell(rr, 1, label).font = FN_NORM; mem.cell(rr, 1).alignment = L
        mem.cell(rr, 2, unit).font = FN_NORM; mem.cell(rr, 2).alignment = CC
        c = mem.cell(rr, 3); c.value = formula; c.number_format = FMT_N3; c.alignment = R; c.font = FN_BOLD
        _qbox["r"] = rr + 1
        return rr

    def esc_faixa(cond_pv, cond_tl):
        tp = [f'IF(AND({cond_pv.format(rr=rr)}),D{rr},0)' for rr in range(PV_FIRST, PV_LAST + 1)]
        tt = [f'IF(AND({cond_tl.format(rr=rr)}),D{rr},0)' for rr in range(TL_FIRST, TL_LAST + 1)]
        return "=" + "+".join(tp + tt)

    def vca_sum(col_let, cond):
        t = [f'IF({cond.format(rr=rr)},{col_let}{rr},0)' for rr in range(V_FIRST, V_LAST + 1)]
        return "=" + "+".join(t)

    Q["ext_tot"] = qrow("Extensão TOTAL da rede DN150 (MND+VCA)", "m", f"=B{EXTTOT_ROW}")
    Q["ext"] = qrow("Extensão em MND (cravação Navigator)", "m", f"=B{EXT_ROW}")
    Q["ext_vca"] = qrow("Extensão em VCA (vala)", "m", f"=B{EXTVCA_ROW}")
    Q["n_pv"] = qrow("Número de PVs", "ud", f"=B{NPV_ROW}")
    Q["pv_acr"] = qrow("Acréscimo PV (Sigma prof-1,00 m)", "m", f"=I{PV_TOT}")
    Q["n_tl"] = qrow("Número de TLs", "ud", f"=B{NTL_ROW}")
    Q["n_lig"] = qrow("Ligações prediais", "ud", f"=B{NLIG_ROW}")
    Q["assent"] = qrow("Assentamento tubo PVC DN150 (VCA)", "m", f"=B{EXTVCA_ROW}")
    Q["passeio"] = qrow("Recomposição de passeio (VCA)", "m2", f"=J{VCA_TOT}")
    Q["berco_vca"] = qrow("Berço de areia (VCA)", "m3", f"=H{VCA_TOT}")
    Q["berco_pv"] = qrow("Berço de areia (PV)", "m3", f"=G{PV_TOT}")
    Q["berco"] = qrow("Berço de areia TOTAL (PV + VCA)", "m3", f"=C{Q['berco_pv']}+C{Q['berco_vca']}")
    Q["esc2"] = qrow("Escavação - faixa <= 2,00 m (cavas + VCA)", "m3",
                     "=" + esc_faixa("B{rr}<=2", "B{rr}<=2")[1:] + "+" + vca_sum("E", "C{rr}<=2")[1:])
    Q["esc4"] = qrow("Escavação - faixa 2,01-4,00 m (cavas + VCA)", "m3",
                     "=" + esc_faixa("B{rr}>2", "B{rr}>2")[1:] + "+" + vca_sum("E", "C{rr}>2")[1:])
    Q["escav_tot"] = qrow("Escavação TOTAL (cavas PV+TL + VCA)", "m3", f"=D{PV_TOT}+D{TL_TOT}+E{VCA_TOT}")
    Q["reat"] = qrow("Reaterro TOTAL (cavas + VCA)", "m3", f"=F{PV_TOT}+F{TL_TOT}+I{VCA_TOT}")
    Q["compact"] = qrow("Compactação (= reaterro)", "m3", f"=C{Q['reat']}")
    Q["botafora"] = qrow("Excedente / bota-fora (escav - reaterro)", "m3",
                         f"=C{Q['escav_tot']}-C{Q['reat']}")
    Q["transp"] = qrow("Transporte bota-fora (m3 x km)", "m3xkm", f"=C{Q['botafora']}*{P_DIST}")
    _vca_desc = vca_sum("F", 'AND(C{rr}>' + f'{ESCOR_MIN}' + ',C{rr}<=' + f'{ESCOR_CONT_MIN}' + ')')[1:]
    _vca_cont = vca_sum("F", 'C{rr}>' + f'{ESCOR_CONT_MIN}')[1:]
    Q["escor_d"] = qrow("Escoramento DESCONTÍNUO (cavas + VCA 1,25-3 m)", "m2",
                        f"=H{PV_TOT}+G{TL_TOT}+" + _vca_desc)
    Q["escor_c"] = qrow("Escoramento CONTÍNUO (VCA prof > 3,00 m)", "m2", "=" + _vca_cont)
    Q["escor"] = qrow("Escoramento TOTAL (descont. + contínuo)", "m2",
                      f"=C{Q['escor_d']}+C{Q['escor_c']}")
    r = _qbox["r"]
    QL = r - 1
    for rr in range(QF, QL + 1):
        if (rr - QF) % 2 == 1:
            for col in range(1, 4):
                mem.cell(rr, col).fill = F_ZEBRA
    box(mem, QHEAD, 1, QL, 3)

    # ---- SECAO 5: MEMORIA DE CALCULO DO BDI ----
    r += 1
    mem.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    hc = mem.cell(r, 1, "5. MEMÓRIA DE CÁLCULO DO BDI  -  Acórdão 2622/2013 (TCU Plenário)  -  composições do orçamento aprovado (Santa Fé)")
    hc.font = FN_HDR; hc.fill = F_GRAF; hc.alignment = L
    for col in range(1, 8):
        mem.cell(r, col).fill = F_GRAF
    mem.row_dimensions[r].height = 22; r += 1

    mem.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    t1 = mem.cell(r, 1, "BDI 1 - OBRAS E SERVIÇOS"); t1.font = FN_SUB; t1.fill = F_SUB; t1.alignment = CC
    for col in range(1, 4):
        mem.cell(r, col).fill = F_SUB
    mem.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    t2 = mem.cell(r, 5, "BDI 2 - FORNECIMENTO DE MATERIAIS"); t2.font = FN_SUB; t2.fill = F_SUB; t2.alignment = CC
    for col in range(5, 8):
        mem.cell(r, col).fill = F_SUB
    r += 1

    for j, h in enumerate(["Parcela", "Sigla", "%"], 1):
        cc = mem.cell(r, j, h); cc.font = FN_HDR; cc.alignment = C; cc.fill = F_GRAF
    for j, h in zip((5, 6, 7), ["Parcela", "Sigla", "%"]):
        cc = mem.cell(r, j, h); cc.font = FN_HDR; cc.alignment = C; cc.fill = F_GRAF
    BDI_HEAD = r; r += 1

    BDI_PARC = [("Administração Central", "AC", 0.0500, 0.0250),
                ("Seguros e Garantias", "S+G", 0.0050, 0.0050),
                ("Riscos", "R", 0.0130, 0.0080),
                ("Despesas Financeiras", "DF", 0.0075, 0.0085),
                ("Lucro", "L", 0.0800, 0.0400),
                ("COFINS", "COFINS", 0.0300, 0.0300),
                ("PIS", "PIS", 0.0065, 0.0065),
                ("ISS", "ISS", 0.0300, 0.0000),
                ("CPRB", "CPRB", 0.0000, 0.0000)]
    parc_obra = {}; parc_mat = {}
    for i, (desc, sig, vo, vm) in enumerate(BDI_PARC):
        rr = r + i
        mem.cell(rr, 1, desc).font = FN_NORM; mem.cell(rr, 1).alignment = L
        mem.cell(rr, 2, sig).font = FN_BOLD; mem.cell(rr, 2).alignment = CC
        cv = mem.cell(rr, 3, vo); cv.number_format = FMT_PCT; cv.alignment = R; cv.font = FN_BOLD; cv.fill = F_COTAR
        parc_obra[sig] = f"C{rr}"
        mem.cell(rr, 5, desc).font = FN_NORM; mem.cell(rr, 5).alignment = L
        mem.cell(rr, 6, sig).font = FN_BOLD; mem.cell(rr, 6).alignment = CC
        cm = mem.cell(rr, 7, vm); cm.number_format = FMT_PCT; cm.alignment = R; cm.font = FN_BOLD; cm.fill = F_COTAR
        parc_mat[sig] = f"G{rr}"
        if i % 2 == 1:
            for col in (1, 2, 5, 6):
                mem.cell(rr, col).fill = F_ZEBRA
    BDI_LAST = r + len(BDI_PARC) - 1
    box(mem, BDI_HEAD, 1, BDI_LAST, 3); box(mem, BDI_HEAD, 5, BDI_LAST, 7)
    r = BDI_LAST + 1

    def _bdi_formula(pc):
        AC, SG, R_, DF = pc["AC"], pc["S+G"], pc["R"], pc["DF"]
        L_, COF, PIS_, ISS_, CPRB_ = pc["L"], pc["COFINS"], pc["PIS"], pc["ISS"], pc["CPRB"]
        return f"=((1+{AC}+{SG}+{R_})*(1+{DF})*(1+{L_}))/(1-({COF}+{PIS_}+{ISS_}+{CPRB_}))-1"

    mem.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    bc = mem.cell(r, 1, "BDI 1 Obras"); bc.font = FN_TOT; bc.alignment = R; bc.fill = F_GRAF
    mem.cell(r, 2).fill = F_GRAF
    bres = mem.cell(r, 3); bres.value = _bdi_formula(parc_obra); bres.number_format = FMT_PCT
    bres.font = FN_TOT; bres.alignment = CC; bres.fill = F_GRAF
    BDI_CALC_REF = f"'{WS_MEM}'!$C${r}"

    mem.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    bm = mem.cell(r, 5, "BDI 2 Materiais"); bm.font = FN_TOT; bm.alignment = R; bm.fill = F_GRAF
    mem.cell(r, 6).fill = F_GRAF
    mres = mem.cell(r, 7); mres.value = _bdi_formula(parc_mat); mres.number_format = FMT_PCT
    mres.font = FN_TOT; mres.alignment = CC; mres.fill = F_GRAF
    BDI_MAT_REF = f"'{WS_MEM}'!$G${r}"
    mem.row_dimensions[r].height = 20
    box(mem, BDI_HEAD, 1, r, 3, side=med); box(mem, BDI_HEAD, 5, r, 7, side=med); r += 1

    mem.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    mem.cell(r, 1, "BDI = ( (1+AC+S+G+R) x (1+DF) x (1+L) ) / ( 1 - (COFINS+PIS+ISS+CPRB) ) - 1").font = FN_NOTE
    mem.cell(r, 1).alignment = L; r += 1
    mem.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    mem.cell(r, 1, ("Dois BDIs conforme orçamento aprovado de Santa Fé: BDI 1 (Obras e Serviços, ~24,49%) p/ a maioria dos itens; "
                    "BDI 2 (Fornecimento de Materiais, ~12,99%) somente p/ itens de fornecimento puro de material "
                    "(no Santa Fé: placa de obra em chapa metálica 001.005.001 e tapume de chapa de madeira compensada 001.004.003). "
                    "Parcelas editáveis nas células amarelas. Coluna 'Tipo' da aba Orçamento define qual BDI cada item usa.")).font = FN_NOTE
    mem.cell(r, 1).alignment = L

    wmem = {1: 30, 2: 11, 3: 13, 4: 4, 5: 30, 6: 11, 7: 13, 8: 12, 9: 14, 10: 11}
    for col, w in wmem.items():
        mem.column_dimensions[get_column_letter(col)].width = w
    mem.freeze_panes = "A3"

    def qref(key):
        return f"'{WS_MEM}'!C{Q[key]}"

    # =====================================================================
    # ABA: ORCAMENTO
    # =====================================================================
    orc = wb.create_sheet(WS_ORC); orc.sheet_view.showGridLines = False
    orc.merge_cells("A1:L1")
    orc["A1"] = f"PLANILHA ORÇAMENTÁRIA  -  {titulo_obra}"
    orc["A1"].font = FN_TITLE; orc["A1"].fill = F_GRAF; orc["A1"].alignment = CC
    orc.row_dimensions[1].height = 26
    orc.merge_cells("A2:L2")
    orc["A2"] = (f"Base: {OBRA['base']}   |   {OBRA['titulo']}   |   Contratante: {OBRA['contratante']}   |   "
                 f"Projeto: {OBRA['projetista']}   |   MND {EXT_MND:,.2f} m  +  VCA {EXT_VCA:,.2f} m  =  "
                 f"{EXT:,.2f} m / {N_OSE} OSEs")
    orc["A2"].font = FN_SUBTL; orc["A2"].fill = F_GRAF; orc["A2"].alignment = CC
    orc.row_dimensions[2].height = 18

    orc["I3"] = "BDI 1 Obras/Serviços:"; orc["I3"].font = FN_BOLD; orc["I3"].alignment = R
    orc.merge_cells("I3:J3")
    orc["K3"] = f"={BDI_CALC_REF}"; orc["K3"].number_format = FMT_PCT; orc["K3"].font = FN_BOLD
    orc["K3"].alignment = CC; orc["K3"].fill = F_SUB
    box(orc, 3, 11, 3, 11)
    orc["I4"] = "BDI 2 Fornec. Materiais:"; orc["I4"].font = FN_BOLD; orc["I4"].alignment = R
    orc.merge_cells("I4:J4")
    orc["K4"] = f"={BDI_MAT_REF}"; orc["K4"].number_format = FMT_PCT; orc["K4"].font = FN_BOLD
    orc["K4"].alignment = CC; orc["K4"].fill = F_COTAR
    box(orc, 4, 11, 4, 11)
    BDI_REF = f"'{WS_ORC}'!$K$3"
    BDI_MAT_DISP = f"'{WS_ORC}'!$K$4"

    MATERIAL_CODES = {"001005001", "001004003", "008021001", "008025001", "008025009", "009031001"}

    def _is_material(cod):
        return str(cod).replace(".", "").strip() in MATERIAL_CODES

    HR = 5
    NCOL = 12
    heads = ["ITEM", "CÓDIGO", "ORIGEM", "DESCRIÇÃO", "UN", "QUANT.", "CUSTO UNIT (R$)",
             "CUSTO DIRETO (R$)", "BDI (%)", "PREÇO C/ BDI (R$)", "TOTAL (R$)", "TIPO (Obra/Mat)"]
    for j, h in enumerate(heads, 1):
        c = orc.cell(HR, j, h); c.font = FN_HDR; c.fill = F_GRAF; c.alignment = C
    orc.row_dimensions[HR].height = 32

    def it(p_key, desc=None, un=None, qf=None, manual=None, cod=None, pu=None, destac=False, note=None):
        cod_, desc_, un_, pu_ = (None, None, None, None)
        if p_key:
            cod_, desc_, un_, pu_ = P[p_key]
        return (cod or cod_, desc or desc_, un or un_, qf, manual,
                pu if pu is not None else pu_, destac, note)

    GRUPOS = [
        ("1. SERVIÇOS PRELIMINARES, CANTEIRO E SINALIZAÇÃO", [
            it("mobiliz", qf=None, manual=1),
            it("placa", qf=f"{P_PLACA_AREA}",
               note="QUANT = ÁREA da placa (parâmetro 'Placa de obra - ÁREA' na Memória, modelo 2,0x1,5 = 3,0 m2). Material -> BDI 2."),
            it("tapume", qf=f"{P_CANT_PERIM}",
               note="QUANT = PERÍMETRO do canteiro = 2*(LARG+COMPR) (parâmetros na Memória). Material -> BDI 2."),
            it("placa_adv", qf=f"{P_NFRENTES}*{P_PLACAS_PF}",
               note="QUANT = N_FRENTES x PLACAS_POR_FRENTE (parâmetros na Memória). Sinalização -> Obra (BDI 1)."),
            it("fita", qf=f"{P_NFRENTES}*{P_FITA_PF}",
               note="QUANT = N_FRENTES x FITA_POR_FRENTE (parâmetros na Memória). Sinalização -> Obra (BDI 1)."),
            it("tap_mov", qf=f"{P_NFRENTES}*{P_TAPMOV_PF}",
               note="QUANT = N_FRENTES x TAPUME_MOVEL_POR_FRENTE (perim. poço MND, parâmetros na Memória). Segurança -> Obra (BDI 1)."),
        ]),
        ("2. SERVIÇOS TÉCNICOS", [
            it("locacao", qf=qref("ext_tot")),
            it("asbuilt", qf=qref("n_pv"), desc="Cadastro de obra localizada (As Built) - por OSE", un="ud", pu=154.12),
        ]),
        ("3. REDE COLETORA EM MND (Cravação Navigator)", [
            it("mnd_150", desc="Cravação MND - Navigator DN150 (trechos prof <= 3,00 m)", qf=qref("ext")),
        ]),
        ("4. REDE COLETORA EM VCA (vala - trechos prof > 3,00 m)", [
            it("assent", qf=qref("assent")),
            it("berco", desc="Berço / embasamento de areia (rede VCA)", qf=qref("berco_vca")),
            it("passeio", qf=qref("passeio")),
        ]),
        (f"5. POÇOS DE VISITA ({len(PVS)} un. Tipo A DN800)", [
            it("pv_base", qf=qref("n_pv")),
            it("pv_acr", qf=qref("pv_acr")),
            it("berco", desc="Berço / embasamento de areia (PV)", qf=qref("berco_pv")),
        ]),
        (f"6. TERMINAIS DE LIMPEZA ({len(TLS)} un. DN150)", [
            it("tl", qf=qref("n_tl")),
        ]),
        (f"7. LIGAÇÕES PREDIAIS ({N_LIG} un.)", [
            it("lig", qf=qref("n_lig"), destac=True),
            it("til", qf=qref("n_lig"), destac=True),
        ]),
        ("8. MOVIMENTO DE SOLOS E ESCORAMENTO (cavas PV/TL + trechos VCA)", [
            it("esc_2", desc="Escavação mec. - prof. até 2,00 m (cavas + VCA)", qf=qref("esc2")),
            it("esc_4", desc="Escavação mec. - prof. 2,01 a 4,00 m (cavas + VCA)", qf=qref("esc4")),
            it("escor_d", desc="Escoramento descontínuo (cavas + VCA 1,25-3,00 m)", qf=qref("escor_d")),
            it("escor_c", desc="Escoramento contínuo (VCA prof > 3,00 m)", qf=qref("escor_c")),
            it("reat", qf=qref("reat")),
            it("comp", qf=qref("compact")),
            it("carga", qf=qref("botafora")),
            it("transp", qf=qref("transp")),
        ]),
        ("9. FORNECIMENTO DE MATERIAIS (ACCIONA)", [
            it("tubo", qf=qref("ext_tot"), destac=True,
               note=("[Fornecimento ACCIONA - a cotar (SINAPI/cotação); excluir se não aplicável]. "
                     "ATENÇÃO: o assentamento VCA 009.002.002 (Grupo 4) já embute material de tubo nos "
                     f"{EXT_VCA:.2f} m de VCA -> possível dupla contagem nesses metros; o Lucas decide. "
                     "Item de MATERIAL -> BDI 2 (12,99%).")),
        ]),
    ]

    ASBUILT_QTY_REF = f"'{WS_MEM}'!B{NOSE_ROW}"
    TUBO_COD = P["tubo"][0]

    r = HR + 1; item_no = 0; group_subtotal_rows = []; group_names = []; zebra = 0
    n_material_items = 0
    F_MAT = PatternFill("solid", fgColor="FFF2CC")
    for gname, items in GRUPOS:
        orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        gc = orc.cell(r, 1, gname); gc.font = FN_GRP; gc.fill = F_GRP; gc.alignment = L
        for col in range(1, NCOL + 1):
            orc.cell(r, col).fill = F_GRP
        grp_first = r + 1; r += 1
        for (cod, desc, un, qf, manual, pu, destac, note) in items:
            item_no += 1
            is_cot = str(cod).upper().startswith("SINAPI") or str(cod).strip() == ""
            is_mat = _is_material(cod) or is_cot
            if is_mat:
                n_material_items += 1
            orc.cell(r, 1, item_no).alignment = CC; orc.cell(r, 1).font = FN_NORM
            orc.cell(r, 2, cod).alignment = CC; orc.cell(r, 2).font = FN_NORM
            og = orc.cell(r, 3, origem_de(cod)); og.alignment = CC; og.font = FN_NOTE if is_cot else FN_NORM
            dc = orc.cell(r, 4, desc); dc.alignment = L; dc.font = FN_NORM
            orc.cell(r, 5, un).alignment = CC; orc.cell(r, 5).font = FN_NORM
            cq = orc.cell(r, 6)
            if cod == "002.007.005":
                cq.value = f"={ASBUILT_QTY_REF}"
            elif qf is None:
                cq.value = manual
            else:
                cq.value = f"={qf}"
            cq.number_format = FMT_N2; cq.alignment = R; cq.font = FN_NORM
            cpu = orc.cell(r, 7); cpu.value = pu
            cpu.number_format = FMT_RS; cpu.alignment = R; cpu.font = FN_NORM
            ccd = orc.cell(r, 8); ccd.value = f'=IF(G{r}="",0,F{r}*G{r})'
            ccd.number_format = FMT_RS; ccd.alignment = R; ccd.font = FN_NORM
            cb = orc.cell(r, 9); cb.value = f"={BDI_MAT_DISP if is_mat else BDI_REF}"; cb.number_format = FMT_PCT
            cb.alignment = CC; cb.font = FN_NORM
            cpb = orc.cell(r, 10); cpb.value = f'=IF(G{r}="","",G{r}*(1+I{r}))'
            cpb.number_format = FMT_RS; cpb.alignment = R; cpb.font = FN_NORM
            ct = orc.cell(r, 11); ct.value = f'=IF(J{r}="",0,F{r}*J{r})'
            ct.number_format = FMT_RS; ct.alignment = R; ct.font = FN_BOLD
            tp = orc.cell(r, 12, "MATERIAL" if is_mat else "OBRA"); tp.alignment = CC
            tp.font = FN_BOLD if is_mat else FN_NORM
            if note:
                dc.comment = Comment(note, "2S")
            if is_cot:
                cpu.fill = F_COTAR
                for col in (6, 10, 11):
                    orc.cell(r, col).fill = F_COTAR
            elif destac:
                dc.comment = Comment("[Item destacável - ligações prediais]", "2S")
                for col in (6, 10, 11):
                    orc.cell(r, col).fill = F_COTAR
            else:
                if is_mat:
                    for col in range(1, NCOL + 1):
                        orc.cell(r, col).fill = F_MAT
                elif zebra % 2 == 1:
                    for col in range(1, NCOL + 1):
                        cur = orc.cell(r, col).fill
                        if cur is None or cur.fgColor.rgb in (None, "00000000"):
                            orc.cell(r, col).fill = F_ZEBRA
                zebra += 1
            if is_mat:
                tp.comment = Comment("BDI 2 - Fornecimento de Materiais. Demais itens usam BDI 1 (Obras).", "2S")
            r += 1
        grp_last = r - 1
        orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        sc = orc.cell(r, 1, f"   Subtotal grupo {gname.split('.')[0]}"); sc.font = FN_SUB; sc.alignment = R
        cd_st = orc.cell(r, 8, f"=SUM(H{grp_first}:H{grp_last})"); cd_st.number_format = FMT_RS
        cd_st.font = FN_SUB; cd_st.alignment = R
        st = orc.cell(r, 11, f"=SUM(K{grp_first}:K{grp_last})"); st.number_format = FMT_RS
        st.font = FN_SUB; st.alignment = R
        for col in range(1, NCOL + 1):
            orc.cell(r, col).fill = F_SUB
        group_subtotal_rows.append(r); group_names.append(gname); r += 1

    ITEM_LAST = r - 1
    box(orc, HR, 1, ITEM_LAST, NCOL)

    r += 1
    cd_sum = "+".join(f"H{sr}" for sr in group_subtotal_rows)
    tot_sum = "+".join(f"K{sr}" for sr in group_subtotal_rows)
    orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    orc.cell(r, 1, "CUSTO DIRETO TOTAL (sem BDI)").font = FN_BOLD; orc.cell(r, 1).alignment = R
    cdc = orc.cell(r, 11); cdc.value = f"={cd_sum}"; cdc.number_format = FMT_RS; cdc.font = FN_BOLD
    for col in range(1, NCOL + 1):
        orc.cell(r, col).fill = F_SUB
    CD_ROW = r; r += 1
    orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    orc.cell(r, 1, "BDI sobre custo direto (mix Obras 24,49% + Materiais 12,99%)").font = FN_BOLD; orc.cell(r, 1).alignment = R
    bc = orc.cell(r, 11); bc.value = f"={tot_sum}-K{CD_ROW}"; bc.number_format = FMT_RS; bc.font = FN_BOLD
    for col in range(1, NCOL + 1):
        orc.cell(r, col).fill = F_SUB
    BDI_ROW = r; r += 1
    orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    orc.cell(r, 1, "BDI EFETIVO (mix) = BDI total / Custo Direto").font = FN_BOLD; orc.cell(r, 1).alignment = R
    be = orc.cell(r, 11); be.value = f"=IF(K{CD_ROW}=0,0,K{BDI_ROW}/K{CD_ROW})"; be.number_format = FMT_PCT
    be.font = FN_BOLD; be.alignment = R
    for col in range(1, NCOL + 1):
        orc.cell(r, col).fill = F_SUB
    BDIEFF_ROW = r; r += 1
    orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    orc.cell(r, 1, "VALOR TOTAL COM BDI").font = FN_TOT; orc.cell(r, 1).alignment = R
    tc = orc.cell(r, 11); tc.value = f"={tot_sum}"; tc.number_format = FMT_RS; tc.font = FN_TOT
    for col in range(1, NCOL + 1):
        orc.cell(r, col).fill = F_RED
    orc.row_dimensions[r].height = 22; TOT_ROW = r
    box(orc, CD_ROW, 1, TOT_ROW, NCOL, side=med)

    r += 2
    orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    orc.cell(r, 1, (f"Rede dividida por PROFUNDIDADE do trecho (regra trecho-inteiro): {EXT_MND:,.2f} m em MND "
                    f"(cravação Navigator, prof <= 3 m) e {EXT_VCA:,.2f} m em VCA (vala aberta, prof do trecho > 3 m). "
                    "Nos trechos VCA: escavação escalonada por faixa, escoramento (contínuo onde prof > 3 m), "
                    "assentamento 009.002, berço de areia, reaterro, compactação, recomposição de passeio "
                    "(lajota sextavada 010.003.001), carga e transporte do excedente. Cavas de PV/TL sempre "
                    "escavadas. Grupo 7 (Ligações) e Grupo 9 (Tubo ACCIONA) destacáveis em amarelo.")).font = FN_NOTE
    r += 1
    orc.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    orc.cell(r, 1, ("Coluna ORIGEM: SANEPAR - MOS JUN/25 (itens com código MOS) ou SINAPI/Cotação (tubo a cotar). "
                    "BDI por item: coluna TIPO indica Obra/Serviço (BDI 1 ~24,49%) ou Material (BDI 2 ~12,99%). "
                    "Tubo PVC DN150 = fornecimento ACCIONA, preço em célula (a cotar); ver nota da célula "
                    "sobre possível dupla contagem do material de tubo nos metros de VCA.")).font = FN_NOTE

    worc = {1: 5, 2: 11, 3: 20, 4: 50, 5: 7, 6: 12, 7: 14, 8: 15, 9: 8, 10: 15, 11: 17, 12: 13}
    for col, w in worc.items():
        orc.column_dimensions[get_column_letter(col)].width = w
    orc.freeze_panes = "A6"

    # =====================================================================
    # ABA: RESUMO (com logos)
    # =====================================================================
    res = wb.create_sheet(WS_RES, 0); res.sheet_view.showGridLines = False
    COL_W = [10, 44, 22, 12, 8]
    for col in range(1, 6):
        res.column_dimensions[get_column_letter(col)].width = COL_W[col - 1]
    LOGO_CM = 2.30
    res.row_dimensions[1].height = 72; res.row_dimensions[2].height = 3
    res.merge_cells("A1:E1"); res["A1"].fill = PatternFill("solid", fgColor=C_WHITE)
    res.merge_cells("A2:E2")
    for col in range(1, 6):
        res.cell(2, col).fill = F_REDLINE

    PT_PER_CHARW = 5.40; EMU_PER_PT = 12700

    def colw_to_emu(w):
        return int(round(w * PT_PER_CHARW * EMU_PER_PT))

    def emu_x_to_colmarker(left_emu, col_emu_widths):
        acc = 0
        for ci, cw in enumerate(col_emu_widths):
            if left_emu < acc + cw:
                return ci, int(left_emu - acc)
            acc += cw
        last = len(col_emu_widths) - 1
        return last, int(col_emu_widths[last])

    def add_logo_emu_x(ws, path, src_w, src_h, target_h_cm, row0, left_emu, col_emu_widths, row_off_emu=0):
        img = XLImage(path)
        h_emu = cm_to_EMU(target_h_cm); w_emu = int(h_emu * (src_w / src_h))
        col0, coloff = emu_x_to_colmarker(left_emu, col_emu_widths)
        marker = AnchorMarker(col=col0, colOff=int(coloff), row=row0, rowOff=int(row_off_emu))
        img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(w_emu, h_emu))
        ws.add_image(img); return w_emu

    try:
        col_emu_widths = [colw_to_emu(w) for w in COL_W]
        table_w_emu = sum(col_emu_widths)
        v_off = cm_to_EMU(0.12); margem_lat = cm_to_EMU(0.12)
        if os.path.exists(LOGO_2S):
            add_logo_emu_x(res, LOGO_2S, 232, 232, LOGO_CM, row0=0, left_emu=margem_lat,
                           col_emu_widths=col_emu_widths, row_off_emu=v_off)
        if os.path.exists(LOGO_ACC):
            w_acc = int(cm_to_EMU(LOGO_CM) * (640 / 288))
            left_acc = max(0, table_w_emu - w_acc - margem_lat)
            add_logo_emu_x(res, LOGO_ACC, 640, 288, LOGO_CM, row0=0, left_emu=left_acc,
                           col_emu_widths=col_emu_widths, row_off_emu=v_off)
    except Exception as e:
        print("AVISO logos:", e, file=sys.stderr)

    res.merge_cells("A3:E3")
    res["A3"] = "ORÇAMENTO  -  REDE COLETORA DE ESGOTO (RCE)"
    res["A3"].font = FN_TITLE; res["A3"].fill = F_GRAF; res["A3"].alignment = CC
    res.row_dimensions[3].height = 28
    res.merge_cells("A4:E4")
    sub4 = (muni if muni else titulo_obra) + "  -  MND + VCA (por profundidade)  -  " + OBRA["projetista"]
    res["A4"] = sub4
    res["A4"].font = FN_SUBTL; res["A4"].fill = F_GRAF; res["A4"].alignment = CC
    res.row_dimensions[4].height = 18

    info = [("Obra:", OBRA["titulo"]), ("Contratante:", OBRA["contratante"]),
            ("Projetista:", OBRA["projetista"]), ("Base de preços:", OBRA["base"]),
            ("Extensão total / OSEs:", f"{EXT:,.2f} m  /  {N_OSE} OSEs"),
            ("  em MND (Navigator):", f"{EXT_MND:,.2f} m  (trechos prof <= 3,00 m)"),
            ("  em VCA (vala):", f"{EXT_VCA:,.2f} m  (trechos prof > 3,00 m)"),
            ("Método:", "DN150 PVC - MND (cravação) + VCA (vala) por profundidade"),
            ("Estruturas:", f"{len(PVS)} PV Tipo A DN800  /  {len(TLS)} TL DN150  /  {N_LIG} ligações"),
            ("BDI 1 Obras/Serviços:", None),
            ("BDI 2 Fornec. Materiais:", None),
            ("BDI efetivo (mix):", None)]
    rr = 6
    for lbl, val in info:
        res.cell(rr, 1, lbl).font = FN_LBL; res.cell(rr, 1).alignment = R; res.cell(rr, 1).fill = F_INFO
        if lbl == "BDI 1 Obras/Serviços:":
            c = res.cell(rr, 2); c.value = f"={BDI_REF}"; c.number_format = FMT_PCT; c.font = FN_BOLD; c.alignment = L
        elif lbl == "BDI 2 Fornec. Materiais:":
            c = res.cell(rr, 2); c.value = f"={BDI_MAT_DISP}"; c.number_format = FMT_PCT; c.font = FN_BOLD; c.alignment = L
        elif lbl == "BDI efetivo (mix):":
            c = res.cell(rr, 2); c.value = f"='{WS_ORC}'!K{BDIEFF_ROW}"; c.number_format = FMT_PCT; c.font = FN_BOLD; c.alignment = L
        else:
            res.cell(rr, 2, val).font = FN_NORM; res.cell(rr, 2).alignment = L
        res.merge_cells(start_row=rr, start_column=2, end_row=rr, end_column=5)
        rr += 1
    box(res, 6, 1, rr - 1, 5)

    rr += 1
    res.cell(rr, 1, "RESUMO POR GRUPO").font = FN_LBL; rr += 1
    RH = rr
    for j, h in enumerate(["ITEM", "GRUPO", "VALOR C/ BDI (R$)", "% TOTAL"], 1):
        c = res.cell(rr, j, h); c.font = FN_HDR; c.fill = F_GRAF; c.alignment = C
    res.merge_cells(start_row=RH, start_column=2, end_row=RH, end_column=2)
    res.row_dimensions[rr].height = 26; rr += 1
    G_FIRST = rr
    for i, (gname, subrow) in enumerate(zip(group_names, group_subtotal_rows)):
        res.cell(rr, 1, gname.split(".")[0]).alignment = CC; res.cell(rr, 1).font = FN_NORM
        res.cell(rr, 2, gname[gname.find(".") + 1:].strip()).alignment = L; res.cell(rr, 2).font = FN_NORM
        vc = res.cell(rr, 3); vc.value = f"='{WS_ORC}'!K{subrow}"; vc.number_format = FMT_RS; vc.alignment = R; vc.font = FN_NORM
        pc = res.cell(rr, 4); pc.alignment = CC; pc.number_format = FMT_PCT1; pc.font = FN_NORM
        if "LIGAÇÕES" in gname or "FORNECIMENTO" in gname:
            for col in range(1, 5):
                res.cell(rr, col).fill = F_COTAR
        elif i % 2 == 1:
            for col in range(1, 5):
                res.cell(rr, col).fill = F_ZEBRA
        rr += 1
    G_LAST = rr - 1
    box(res, RH, 1, G_LAST, 4)

    res.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
    res.cell(rr, 1, "CUSTO DIRETO (sem BDI)").font = FN_BOLD; res.cell(rr, 1).alignment = R
    cdr = res.cell(rr, 3); cdr.value = f"='{WS_ORC}'!K{CD_ROW}"; cdr.number_format = FMT_RS; cdr.font = FN_BOLD; cdr.alignment = R
    for col in range(1, 5):
        res.cell(rr, col).fill = F_SUB
    RES_CD = rr; rr += 1
    res.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
    res.cell(rr, 1, "BDI").font = FN_BOLD; res.cell(rr, 1).alignment = R
    bdr = res.cell(rr, 3); bdr.value = f"='{WS_ORC}'!K{BDI_ROW}"; bdr.number_format = FMT_RS; bdr.font = FN_BOLD; bdr.alignment = R
    for col in range(1, 5):
        res.cell(rr, col).fill = F_SUB
    RES_BDI = rr; rr += 1
    res.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
    res.cell(rr, 1, "VALOR TOTAL COM BDI").font = FN_TOT; res.cell(rr, 1).alignment = R
    tgr = res.cell(rr, 3); tgr.value = f"='{WS_ORC}'!K{TOT_ROW}"; tgr.number_format = FMT_RS; tgr.font = FN_TOT; tgr.alignment = R
    for col in range(1, 5):
        res.cell(rr, col).fill = F_RED
    res.row_dimensions[rr].height = 22; RES_TOT = rr
    box(res, RES_CD, 1, RES_TOT, 4, side=med)
    for r2 in range(G_FIRST, G_LAST + 1):
        res.cell(r2, 4).value = f"=IF($C${RES_TOT}=0,0,C{r2}/$C${RES_TOT})"

    rr += 2
    res.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)
    res.cell(rr, 1, ("Grupo 7 (Ligações) e Grupo 9 (Tubo - fornecimento ACCIONA) destacados em amarelo. "
                     "Rede repartida MND/VCA por profundidade do trecho. Valores com BDI por item "
                     "(BDI 1 Obras 24,49% / BDI 2 Materiais 12,99% - Acórdão TCU 2622/2013).")).font = FN_NOTE
    res.freeze_panes = "A5"

    # =====================================================================
    # ABA: BASE DE PRECOS
    # =====================================================================
    base = wb.create_sheet(WS_BASE); base.sheet_view.showGridLines = False
    base.merge_cells("A1:E1")
    base["A1"] = "BASE DE PREÇOS  (custo direto, sem BDI)  -  SANEPAR MOS 5a Ed. JUN/2025 + Fornecimento (a cotar)"
    base["A1"].font = FN_TITLE; base["A1"].fill = F_GRAF; base["A1"].alignment = CC
    base.row_dimensions[1].height = 26
    BH = 3
    for j, h in enumerate(["CÓDIGO", "ORIGEM", "DESCRIÇÃO", "UN", "CUSTO UNIT (R$)"], 1):
        c = base.cell(BH, j, h); c.font = FN_HDR; c.fill = F_GRAF; c.alignment = C
    base.row_dimensions[BH].height = 24
    ordem = ["mobiliz", "placa", "tapume", "placa_adv", "fita", "tap_mov", "locacao", "asbuilt", "mnd_150", "assent", "passeio",
             "pv_base", "pv_acr", "berco", "tl", "lig", "til", "esc_2", "esc_4", "escor_d", "escor_c",
             "reat", "comp", "carga", "transp", "tubo"]
    r = BH + 1
    for i, k in enumerate(ordem):
        cod, desc, un, pu = P[k]
        is_cot = str(cod).upper().startswith("SINAPI")
        base.cell(r, 1, cod).alignment = CC; base.cell(r, 1).font = FN_NORM
        og = base.cell(r, 2, origem_de(cod)); og.alignment = CC; og.font = FN_NOTE if is_cot else FN_NORM
        base.cell(r, 3, desc).alignment = L; base.cell(r, 3).font = FN_NORM
        base.cell(r, 4, un).alignment = CC; base.cell(r, 4).font = FN_NORM
        pc = base.cell(r, 5, pu); pc.number_format = FMT_RS; pc.alignment = R; pc.font = FN_NORM
        if is_cot:
            pc.fill = F_COTAR
            pc.comment = Comment("[Fornecimento ACCIONA - a cotar (SINAPI/cotacao)]", "2S")
        if i % 2 == 1 and not is_cot:
            for col in range(1, 6):
                base.cell(r, col).fill = F_ZEBRA
        r += 1
    box(base, BH, 1, r - 1, 5)
    base.column_dimensions["A"].width = 12; base.column_dimensions["B"].width = 20
    base.column_dimensions["C"].width = 56; base.column_dimensions["D"].width = 8
    base.column_dimensions["E"].width = 16
    base.freeze_panes = "A4"

    # =====================================================================
    # ABA: REFERENCIAS
    # =====================================================================
    WS_REF = "Referências"
    ref = wb.create_sheet(WS_REF); ref.sheet_view.showGridLines = False

    ref.merge_cells("A1:E1")
    ref["A1"] = "REFERÊNCIAS NORMATIVAS  -  FONTE EXATA DE CADA CRITÉRIO TÉCNICO DO ORÇAMENTO"
    ref["A1"].font = FN_TITLE; ref["A1"].fill = F_GRAF; ref["A1"].alignment = CC
    ref.row_dimensions[1].height = 26
    ref.merge_cells("A2:E2")
    ref["A2"] = ("Blindagem do orçamento: cada premissa de medição/preço está amarrada ao trecho literal do manual que a embasa. "
                 "MOS = Manual de Obras de Saneamento SANEPAR, 5a Ed., versão 02, jun/2025. "
                 "MPS 11.3 = Manual de Projetos de Saneamento, Módulo 11.3 - Unidades Lineares, versão 2025/R0. "
                 "Páginas conferidas no PDF real (número do PDF + label interno do módulo).")
    ref["A2"].font = FN_NOTE; ref["A2"].alignment = LC
    ref.row_dimensions[2].height = 30

    REF_HDR_TXT = ["Nº", "CRITÉRIO", "REGRA APLICADA NO ORÇAMENTO", "FONTE (Manual - Módulo - Página)", "TRECHO CITADO (transcrição literal do manual)"]
    RH = 4
    for j, h in enumerate(REF_HDR_TXT, 1):
        c = ref.cell(RH, j, h); c.font = FN_HDR; c.fill = F_GRAF; c.alignment = C
    ref.row_dimensions[RH].height = 30

    REFS = [
        ("Faixa de escavação por profundidade (cod 004.002)",
         "Escavação mecânica de vala precificada/medida por faixa de profundidade escalonada: 0<h<=2 m / <=4 m / <=6 m / <=8 m (itens 004.002.001 a 004.002.008). No orçamento, a escavação das cavas de PV/TL e dos trechos VCA é separada nas faixas <=2 m e 2,01-4,00 m.",
         "MOS - Módulo 4 (Movimento de Solos) - Regulamentação de Preços - PDF pág. 269 (label interno 28/37).",
         "\"004002 ESCAVAÇÃO MECÂNICA DE VALAS EM QUALQUER TIPO DE SOLO, EXCETO ROCHA - 004002001 Solos em geral, profundidade 0 m < h <= 2 m; 004002002 ... 0 m < h <= 4 m; ... <= 6 m; ... <= 8 m.\"  Nota (c): \"A medição deve ser escalonada. Ex.: O 1o metro de solo arenoso, paga-se 004001001, o 2o metro 004001002 e assim sucessivamente.\""),
        ("Acréscimo de PV por profundidade (cod 009.015)",
         "PV Tipo A DN800 cobrado como parcela fixa \"até 1,00 m\" (009.015.001) + acréscimo medido em metros para a profundidade acima de 1,00 m (009.015.002). No orçamento, qtd do acréscimo = soma de (prof. de cada PV - 1,00 m).",
         "MOS - Módulo 9 (Assentamentos) - Regulamentação de Preços - PDF pág. 500 (label interno 63/100).",
         "\"009015 POÇO DE VISITA TIPO A - DN 800 - 009015001 Com profundidade até 1,00 m; 009015002 Acréscimo para prof. superior a 1,00 m.\"  Critério de medição: \"009015 a 009018 - Por unidade, ud, sendo os acréscimos medidos em metros.\""),
        ("Largura de vala por DN x profundidade (Tabela B.1)",
         "Largura da vala dos trechos VCA (DN150) pela Tabela B.1: 0<=h<2 -> 0,65 m; 2<=h<3 -> 0,85 m; 3<=h<4 -> 1,05 m (coluna contínuo/descontínuo - madeira). Base ABNT NBR 17015.",
         "MOS - Módulo 4 (Movimento de Solos) - PDF pág. 244-245 (label interno 3/37 e 4/37).",
         "\"Conforme orientação da NBR 17015, A largura da vala para obras de esgoto e de água, em função do diâmetro da tubulação e da cota de corte (profundidade da vala), são especificadas nas Tabelas B1 e B2.\"  Tabela B.1 (DN 100 e 150): \"0<=h<2: 0,65 m;  2<=h<3: 0,85 m;  3<=h<4: 1,05 m.\""),
        ("Escoramento obrigatório acima de 1,25 m + tipos por profundidade",
         "Trechos/cavas com prof. > 1,25 m recebem escoramento. Tipo por profundidade: descontínuo até 3,00 m; contínuo para prof. > 3,00 m.",
         "MOS - Módulo 5 (Escoramento) - PDF pág. 280-281 (label interno 2/22 e 3/22).",
         "\"É obrigatório o escoramento para valas de profundidade superior a 1,25 m, conforme estabelece a portaria n. 3214 do Ministério do Trabalho ... regulamentada pela NR 18 ... NBR 9061.\"  Escoramento descontínuo: \"limitado a valas de até 3,00 m de profundidade\"; pontalete: \"até 2,00 m\"; acima -> contínuo."),
        ("Reaterro VR = VE - VEX + VMS (desconta volume de tubos/poços)",
         "Reaterro dos trechos VCA e das cavas calculado descontando o volume das estruturas (tubo DN150, PV, TL) do volume escavado: VR = VE - VEX + VMS.",
         "MOS - Módulo 4 (Movimento de Solos) - Regulamentação de Preços - PDF pág. 273 (label interno 32/37).",
         "\"004013 ATERRO/REATERRO EM VALAS E CAVAS - Volume, em m3, calculado através da fórmula: VR = VE - VEX + VMS, sendo: VR = Volume do reaterro; VE = Volume do material escavado; VEX = Volume do material a ser exportado (volume de tubulações, caixas, poços); VMS = Volume do material importado para substituição.\""),
        ("Berço / embasamento de areia 10 cm (cod 009.029)",
         "Berço de areia de espessura 0,10 m sob o tubo (trechos VCA) e sob os PV/TL, medido em m3 (item 009.029 Embasamento).",
         "MOS - Módulo 9 (Assentamentos) - PDF pág. 487-488 (label interno 50/100 e 51/100).",
         "\"009029 EMBASAMENTO - ... 2- ... Todas as tubulações devem possuir embasamento com espessura mínima de 10 cm, executado de forma nivelada e devidamente compactada para garantir o adequado apoio e estabilidade das tubulações.\""),
        ("Recobrimento mínimo da rede de esgoto (0,80 / 1,00 / 1,20 m)",
         "Premissa de profundidade da rede: recobrimento mínimo 0,80 m em passeio, 1,00 m em via pavimentada, 1,20 m em via de terra.",
         "MPS - Módulo 11.3 (Unidades Lineares) - PDF pág. 21-22 (label interno 21/32 e 22/32).",
         "\"Nas tubulações de esgoto deve ser observado um recobrimento mínimo final de 0,80 m nos passeios, 1,00 m nas ruas pavimentadas ou com greide definido por meio fio e sarjeta e 1,20 m para via de terra ou com greide indefinido.\""),
        ("Profundidade máxima da rede = 3,00 m (base da divisão MND x VCA)",
         "Regra trecho-inteiro: trecho com max(prof. extremos) <= 3,00 m -> MND (Navigator); trecho com prof. > 3,00 m -> VCA (vala aberta). O limite de 3,00 m é o teto normativo para a rede coletora.",
         "MPS - Módulo 11.3 (Unidades Lineares) - PDF pág. 23 (label interno 23/32).",
         "\"A profundidade máxima das redes coletoras será de 3,00 (três) metros. Em situações especiais, para permitir o atendimento de várias habitações solicitar aprovação prévia da fiscalização da Sanepar.\""),
        ("MND - Cravação Navigator (cod 016.012)",
         "Rede em trechos prof. <= 3,00 m executada por cravação MND - Navigator (perfuração direcional), sem vala. Item 016.012.004 (DN150).",
         "MOS - Módulo 16 (Serviços Diversos) - PDF pág. 795-796 (label interno 8/28 e 9/28).",
         "\"016012 CRAVAÇÃO MND - 'NAVIGATOR' - Processo também denominado como perfuração direcional. É executado com equipamento hidrostático de alta pressão que possui uma cabeça de perfuração (broca) com um dispositivo eletrônico ... Para o devido controle da declividade, deve ser elaborada uma planilha contendo dados de nivelamento a cada 3,00 m. No caso de cravação para redes de esgoto, este método está limitado a uma extensão de 80,00 m, quando deve ser, obrigatoriamente, ser executado um poço de visita.\""),
        ("BDI duplo - Acórdão 2622/2013-TCU (Obras 24,49% / Materiais 12,99%)",
         "Dois BDIs: BDI 1 (Obras e Serviços) ~24,49% para a maioria dos itens; BDI 2 (Fornecimento de Materiais) ~12,99% para fornecimento puro de material. Composições (AC, S+G, R, DF, L, tributos) conforme faixas de referência do Acórdão 2622/2013-TCU Plenário.",
         "TCU - Acórdão 2622/2013 - Plenário (fonte EXTERNA aos manuais SANEPAR; jurisprudência de referência para BDI de obras públicas).",
         "Acórdão 2622/2013-TCU (Plenário): estabelece as faixas referenciais de BDI por tipo de obra de saneamento e a fórmula BDI = [(1+AC+S+G+R) x (1+DF) x (1+L)] / [1-(COFINS+PIS+ISS+CPRB)] - 1, com parcela reduzida para fornecimento de materiais e equipamentos.  (Transcrição do acórdão NÃO consta nos PDFs SANEPAR - texto integral no portal do TCU.)"),
        ("Preço do tubo PVC DN150 (fornecimento de material)",
         f"Tubo coletor de esgoto PVC JEI DN150 (NBR 7362) cotado como item separado de fornecimento (a ACCIONA fornece o tubo). Custo unitário R$ {TUBO_PRECO:,.2f}/m, BDI de Materiais (12,99%).",
         TUBO_ORIGEM_TXT,
         "Adotada a SINAPI de dezembro/2024 por ser a referência mais recente publicada oficialmente pela CAIXA para o Paraná: as referências de 2025 e 2026 não constam no portal oficial (caixa.gov.br/site/Paginas/downloads.aspx) - retornam HTTP 404. Atualizar quando a CAIXA publicar referência mais recente do PR."),
    ]

    C_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
    CC_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
    FN_TRECHO = Font(name="Calibri", italic=True, size=8.5, color="333333")
    r = RH + 1
    for i, (crit, regra, fonte, trecho) in enumerate(REFS):
        ref.cell(r, 1, i + 1).font = FN_BOLD; ref.cell(r, 1).alignment = CC_TOP
        ref.cell(r, 2, crit).font = FN_BOLD; ref.cell(r, 2).alignment = C_TOP
        ref.cell(r, 3, regra).font = FN_NORM; ref.cell(r, 3).alignment = C_TOP
        cf = ref.cell(r, 4, fonte); cf.alignment = C_TOP; cf.font = Font(name="Calibri", bold=True, size=9, color=C_RED)
        ref.cell(r, 5, trecho).font = FN_TRECHO; ref.cell(r, 5).alignment = C_TOP
        if i % 2 == 1:
            for col in range(1, 6):
                ref.cell(r, col).fill = F_ZEBRA
        est_lines = max(4, len(trecho) // 70 + 2)
        ref.row_dimensions[r].height = est_lines * 12.5
        r += 1
    REF_LAST = r - 1
    box(ref, RH, 1, REF_LAST, 5)

    r += 1
    ref.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ref.cell(r, 1, ("Todas as páginas acima foram conferidas no PDF real dos manuais (pdfplumber). A paginação do MOS é por PDF; "
                    "o label interno (ex.: 28/37) é a numeração impressa no rodapé do módulo. Itens 1 a 9 = critérios SANEPAR (MOS/MPS). "
                    "Item 10 (BDI) é norma externa (TCU); o texto integral do Acórdão 2622/2013 deve ser anexado a partir do portal do TCU.")).font = FN_NOTE
    ref.cell(r, 1).alignment = LC; ref.row_dimensions[r].height = 30

    ref.column_dimensions["A"].width = 4
    ref.column_dimensions["B"].width = 30
    ref.column_dimensions["C"].width = 44
    ref.column_dimensions["D"].width = 30
    ref.column_dimensions["E"].width = 70
    ref.freeze_panes = "A5"

    # ordem das abas
    order = [WS_RES, WS_MEM, WS_ORC, WS_BASE, WS_REF]
    wb._sheets.sort(key=lambda s: order.index(s.title))
    wb.active = 0
    wb.save(OUT)

    return dict(saida=OUT, ext=EXT, ext_mnd=EXT_MND, ext_vca=EXT_VCA,
                n_pv=len(PVS), n_tl=len(TLS), n_ose=N_OSE, n_lig=N_LIG,
                cd_ref=("Orçamento", CD_ROW, 11), tot_ref=("Orçamento", TOT_ROW, 11))


# =====================================================================
# 4) CALCULO PYTHON INDEPENDENTE (espelha as formulas do Excel) p/ stdout JSON
# =====================================================================
def calcular_totais(cfg):
    """Recalcula CD e TOTAL em Python puro (mesma metodologia das formulas Excel),
    para reportar no JSON de saida sem depender do Excel COM."""
    data = ler_oses(cfg["oses"])
    PVS = [round(p, 3) for ose, nm, p in data["pvs"]]
    TLS = [round(p, 3) for ose, nm, p in data["tls"]]
    SEGS = data["segs"]
    EXT = data["ext"]; N_OSE = data["n_ose"]; N_LIG = int(cfg["ligacoes"])
    _ev = sum(s["ext"] for s in SEGS)
    EXT_VCA = round(_ev, 2); EXT_MND = round(EXT - _ev, 2)

    PV_LADO = 1.40; PV_DIA = 0.80; TL_LADO = 0.75; TL_DIA = 0.15
    ESCOR_MIN = 1.25; ESCOR_CONT_MIN = 3.00; BERCO = 0.10; DN = 0.150
    DIST = float(cfg["dist_botafora"])

    pv_cava = pv_reat = pv_berco = pv_escor = pv_acr = 0.0; pv_e2 = pv_e4 = 0.0
    for p in PVS:
        cava = PV_LADO**2 * p
        pvvol = math.pi / 4 * PV_DIA**2 * p
        pv_cava += cava; pv_reat += max(0, cava - pvvol)
        pv_berco += PV_LADO**2 * BERCO
        pv_escor += 4 * PV_LADO * p if p > ESCOR_MIN else 0
        pv_acr += max(0, p - 1)
        if p <= 2:
            pv_e2 += cava
        else:
            pv_e4 += cava
    tl_cava = tl_reat = tl_escor = 0.0; tl_e2 = tl_e4 = 0.0
    for p in TLS:
        cava = TL_LADO**2 * p
        tlvol = math.pi / 4 * TL_DIA**2 * p
        tl_cava += cava; tl_reat += max(0, cava - tlvol)
        tl_escor += 4 * TL_LADO * p if p > ESCOR_MIN else 0
        if p <= 2:
            tl_e2 += cava
        else:
            tl_e4 += cava

    def larg(p):
        if p < 2:
            return 0.65
        if p < 3:
            return 0.85
        return 1.05
    vca_escav = vca_escor_d = vca_escor_c = vca_berco = vca_reat = vca_passeio = 0.0; vca_e2 = vca_e4 = 0.0
    for s in SEGS:
        pm = round(s["pm"], 3); ext = round(s["ext"], 3); lg = larg(pm)
        esc = lg * pm * ext; vca_escav += esc
        if pm <= 2:
            vca_e2 += esc
        else:
            vca_e4 += esc
        if pm > ESCOR_MIN:
            area = 2 * pm * ext
            if pm > ESCOR_CONT_MIN:
                vca_escor_c += area
            else:
                vca_escor_d += area
        b = lg * BERCO * ext; vca_berco += b
        tubo = math.pi / 4 * DN**2 * ext
        vca_reat += max(0, esc - tubo - b)
        vca_passeio += lg * ext

    berco_pv = pv_berco
    esc2 = pv_e2 + tl_e2 + vca_e2
    esc4 = pv_e4 + tl_e4 + vca_e4
    reat = pv_reat + tl_reat + vca_reat
    escav_tot = pv_cava + tl_cava + vca_escav
    botafora = escav_tot - reat
    transp = botafora * DIST
    escor_d = pv_escor + tl_escor + vca_escor_d
    escor_c = vca_escor_c

    P = dict(mobiliz=351.69, placa=552.66, tapume=135.68, locacao=2.45, asbuilt=154.12,
             mnd=139.73, assent=9.19, passeio=57.50, pv_base=1346.10, pv_acr=638.92, berco=175.49,
             tl=59.75, lig=482.78, til=55.50, esc2=18.43, esc4=19.90, escor_d=47.47, escor_c=71.27,
             reat=2.45, comp=7.18, carga=3.93, transp=4.69, placa_adv=151.33, fita=0.30, tap_mov=4.00)
    TUBO_PU = float(cfg["tubo_preco"])

    # parametros canteiro/sinalizacao
    frentes = int(cfg["frentes"])
    placa_area = float(cfg["placa_area"])
    perim = 2 * (float(cfg["canteiro_larg"]) + float(cfg["canteiro_compr"]))
    placas = frentes * int(cfg["placas_frente"])
    fita = frentes * float(cfg["fita_frente"])
    tapmov = frentes * float(cfg["tapume_movel_frente"])
    asbuilt_qtd = int(cfg["asbuilt_qtd"]) if cfg["asbuilt_qtd"] is not None else N_OSE

    CD = 0.0; mat_cd = 0.0
    items = []  # (q,pu,is_material)

    def add(q, pu, is_mat=False):
        nonlocal CD, mat_cd
        CD += q * pu
        if is_mat:
            mat_cd += q * pu

    # G1 prelim/canteiro/sinalizacao
    add(1, P["mobiliz"])
    add(placa_area, P["placa"], True)       # material
    add(perim, P["tapume"], True)           # material
    add(placas, P["placa_adv"])
    add(fita, P["fita"])
    add(tapmov, P["tap_mov"])
    # G2 serv tec
    add(EXT, P["locacao"]); add(asbuilt_qtd, P["asbuilt"])
    # G3 MND
    add(EXT_MND, P["mnd"])
    # G4 VCA rede
    add(EXT_VCA, P["assent"]); add(vca_berco, P["berco"]); add(vca_passeio, P["passeio"])
    # G5 PV
    add(len(PVS), P["pv_base"]); add(pv_acr, P["pv_acr"]); add(berco_pv, P["berco"])
    # G6 TL
    add(len(TLS), P["tl"])
    # G7 lig
    add(N_LIG, P["lig"]); add(N_LIG, P["til"])
    # G8 mov solos
    add(esc2, P["esc2"]); add(esc4, P["esc4"])
    add(escor_d, P["escor_d"]); add(escor_c, P["escor_c"])
    add(reat, P["reat"]); add(reat, P["comp"])
    add(botafora, P["carga"]); add(transp, P["transp"])
    # G9 tubo (material)
    add(EXT, TUBO_PU, True)

    BDI1 = ((1 + 0.05 + 0.005 + 0.013) * (1 + 0.0075) * (1 + 0.08)) / (1 - (0.03 + 0.0065 + 0.03 + 0.0)) - 1
    BDI2 = ((1 + 0.025 + 0.005 + 0.008) * (1 + 0.0085) * (1 + 0.04)) / (1 - (0.03 + 0.0065 + 0.0 + 0.0)) - 1
    obra_cd = CD - mat_cd
    BDI_total = obra_cd * BDI1 + mat_cd * BDI2
    TOT = CD + BDI_total
    return dict(custo_direto=round(CD, 2), total=round(TOT, 2),
                bdi_total=round(BDI_total, 2), bdi1=BDI1, bdi2=BDI2,
                ext_mnd=EXT_MND, ext_vca=EXT_VCA)


# =====================================================================
# MAIN
# =====================================================================
def main():
    cfg = build_config()
    info = gerar(cfg)
    tot = calcular_totais(cfg)
    out = {
        "ok": True,
        "saida": info["saida"],
        "total": tot["total"],
        "custo_direto": tot["custo_direto"],
        "ext_mnd": info["ext_mnd"],
        "ext_vca": info["ext_vca"],
        "n_pv": info["n_pv"],
        "n_tl": info["n_tl"],
        "n_ligacoes": info["n_lig"],
        "n_ose": info["n_ose"],
        "ext_total": info["ext"],
    }
    # ultima linha do stdout = JSON (o Nexus captura)
    print(json.dumps(out, ensure_ascii=False))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc(file=sys.stderr)
        print(json.dumps({"ok": False, "erro": str(e)}, ensure_ascii=False))
        sys.exit(1)
