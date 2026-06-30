# -*- coding: utf-8 -*-
import openpyxl, glob, os, math, re
OSE_DIR = r"\\2s-eng-servidor\maringa\PLANILHAS FINAIS\UBIRATÃ\PO-01"
files = [f for f in glob.glob(os.path.join(OSE_DIR, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
print("arquivos:", [os.path.basename(f) for f in files])

bands = {"ate_1_25":0.0, "1_25_2_00":0.0, "2_00_3_00":0.0, "3_00_4_00":0.0, "acima_4_00":0.0}
def classify(p):
    if p <= 1.25: return "ate_1_25"
    if p <= 2.00: return "1_25_2_00"
    if p <= 3.00: return "2_00_3_00"
    if p <= 4.00: return "3_00_4_00"
    return "acima_4_00"

OSE_RE = re.compile(r"^OSE[\s_-]*\d", re.I)
tot_seg_len = 0.0
seg_sem_prof = 0
n_ose = 0
n_files_ose = 0
for f in files:
    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    osesheets = [s for s in wb.sheetnames if OSE_RE.match(s)]
    if osesheets:
        n_files_ose += 1
    for sn in osesheets:
        s = wb[sn]
        nodes = []
        for row in s.iter_rows(min_row=11, values_only=True):
            if len(row) < 19:
                continue
            nome, x, y = row[0], row[1], row[2]
            prof = row[18]  # coluna 19 (0-based 18)
            if nome is None or x is None or y is None:
                continue
            try:
                x = float(x); y = float(y)
            except Exception:
                continue
            p = prof if isinstance(prof, (int, float)) else None
            # dedup consecutivos
            if nodes and nodes[-1][0] == str(nome).strip():
                if p is not None: nodes[-1] = (nodes[-1][0], nodes[-1][1], nodes[-1][2], p)
                continue
            nodes.append((str(nome).strip(), x, y, p))
        if len(nodes) >= 2:
            n_ose += 1
        for i in range(len(nodes)-1):
            a, b = nodes[i], nodes[i+1]
            L = math.hypot(b[1]-a[1], b[2]-a[2])
            if L <= 0 or L > 500:  # descarta saltos absurdos (troca de OSE/erro)
                continue
            pa, pb = a[3], b[3]
            ps = [v for v in (pa, pb) if v is not None]
            if not ps:
                seg_sem_prof += 1
                pm = None
            else:
                pm = sum(ps)/len(ps)
            tot_seg_len += L
            if pm is not None:
                bands[classify(pm)] += L
            else:
                bands["ate_1_25"] += 0  # nao soma em banda; fica fora
    wb.close()

print("arquivos com abas OSE:", n_files_ose, "| OSEs:", n_ose, "| segmentos sem prof:", seg_sem_prof)
print("\n--- EXTENSAO POR FAIXA (m) ---")
soma = 0.0
labels = [("Até 1,25 m","ate_1_25"),("Até 2,00 m","1_25_2_00"),("Até 3,00 m","2_00_3_00"),
          ("Até 4,00 m","3_00_4_00"),("Até 5,00 m","acima_4_00")]
for lab,k in labels:
    print("  %-12s %12.2f" % (lab, bands[k]))
    soma += bands[k]
print("  %-12s %12.2f" % ("TOTAL(bandas)", soma))
print("  total seg (incl. sem prof): %.2f" % tot_seg_len)
print("  alvo quantitativo imagem:   35215.99")
