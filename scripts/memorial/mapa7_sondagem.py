# -*- coding: utf-8 -*-
"""Mapa 7 — LOCAÇÃO DAS SONDAGENS (SPT) — identidade 2S.

Plota os furos de sondagem de um município sobre imagem de satélite (Esri),
cada furo como marcador identificado pelo NÚMERO DO FURO, colorido pela
incidência de nível d'água, no padrão cartográfico 2S (title block, rosa-dos-
ventos, escala gráfica, malha UTM, legenda de rodapé, datum).

Fonte dos furos: Survey123 "Sondagem - Acciona" (XLSX exportado do ArcGIS),
campos: n_mero_do_furo, munic_pio, profundidade_real_m, incidencia_de_agua,
profundidade_agua, Longitude, Latitude (WGS84).

Parametrização por ambiente (o pipeline define antes de rodar; standalone cai
nos defaults de Amaporã):
    MEMORIAL_SONDAGEM_XLSX  -> caminho do RELATORIO_SONDAGEM_*.xlsx
    MEMORIAL_MUNICIPIO      -> nome do municipio (ex.: "Amaporã")
    MEMORIAL_MAPAS          -> pasta de saida do PNG (C.OUT)
"""
import os, sys, math, unicodedata
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import cartolib as C

# ---- parametros ----
_DEF_XLSX = (r"C:\Users\lcabd\OneDrive - 2S ENGENHARIA DE AGRIMENSURA E GEOTECNOLOGIA"
             r"\001. SERVIDOR PARANÁ\002. ACCIONA\004. CT-027.2025 - PROJETOS"
             r"\000. CRONOGRAMAS\DASHBOARD ONLINE\RELATORIO_SONDAGEM_2026-05-29.xlsx")
XLSX = os.environ.get("MEMORIAL_SONDAGEM_XLSX", _DEF_XLSX)
MUNIC = os.environ.get("MEMORIAL_MUNICIPIO", "Amaporã")
OUT_PNG = os.path.join(C.OUT, "Mapa7_Sondagem.png")

# colunas (0-based) do XLSX Survey123
COL_NUM, COL_MUN, COL_PROF = 6, 3, 12
COL_AGUA, COL_PROFAGUA = 50, 51
COL_LON, COL_LAT = 721, 722


def _norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode("ascii")
    return s.strip().upper()


def carregar_furos(xlsx, municipio):
    """Le o XLSX e devolve os furos do municipio: lista de dicts
    {num, lon, lat, prof, agua(bool), prof_agua}."""
    import openpyxl
    alvo = _norm(municipio)
    # aceita match por prefixo (ex.: "Eng Beltrao" no ArcGIS vs "Engenheiro...")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sondagem"] if "Sondagem" in wb.sheetnames else wb[wb.sheetnames[0]]
    furos = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        mun = _norm(r[COL_MUN])
        if not (mun == alvo or mun.startswith(alvo) or alvo.startswith(mun) or alvo in mun):
            continue
        lon, lat = r[COL_LON], r[COL_LAT]
        if lon in (None, "") or lat in (None, ""):
            continue
        try:
            lon = float(str(lon).replace(",", ".")); lat = float(str(lat).replace(",", "."))
        except ValueError:
            continue
        def fnum(v):
            try: return float(str(v).replace(",", "."))
            except Exception: return None
        ag = _norm(r[COL_AGUA])
        agua = ag.startswith("SIM") or ag.startswith("PRESEN") or ag in ("S", "1", "TRUE")
        furos.append(dict(
            num=str(r[COL_NUM] or "").strip(),
            lon=lon, lat=lat,
            prof=fnum(r[COL_PROF]),
            agua=agua,
            prof_agua=fnum(r[COL_PROFAGUA]),
        ))
    return furos


def reprojetar(furos):
    """WGS84 (lon/lat) -> SIRGAS 2000 / UTM 22S (EPSG:31982)."""
    from pyproj import Transformer
    tr = Transformer.from_crs("EPSG:4326", "EPSG:31982", always_xy=True)
    for f in furos:
        f["x"], f["y"] = tr.transform(f["lon"], f["lat"])
    return furos


def main():
    furos = carregar_furos(XLSX, MUNIC)
    if not furos:
        print(f"[mapa7] Nenhum furo encontrado para '{MUNIC}'.")
        return None
    furos = reprojetar(furos)
    # descarta furos com coordenada grosseiramente errada (ex.: 1 ponto no
    # Oceano Atlantico) que estouram o enquadramento: mantem so o cluster do
    # projeto (raio robusto a partir da mediana).
    if len(furos) >= 5:
        _xs = np.array([f["x"] for f in furos]); _ys = np.array([f["y"] for f in furos])
        mcx, mcy = np.median(_xs), np.median(_ys)
        dist = np.hypot(_xs - mcx, _ys - mcy)
        lim = max(8000.0, 6.0 * np.median(dist))
        mant = [f for f, dd in zip(furos, dist) if dd <= lim]
        if len(mant) >= 3 and len(mant) < len(furos):
            print("[mapa7] descartados %d furo(s) fora do cluster (>%.0f m)" %
                  (len(furos) - len(mant), lim))
            furos = mant
    xs = np.array([f["x"] for f in furos]); ys = np.array([f["y"] for f in furos])
    n = len(furos)
    n_agua = sum(1 for f in furos if f["agua"])
    profs = [f["prof"] for f in furos if f["prof"] is not None]
    print(f"[mapa7] {MUNIC}: {n} furos | {n_agua} com NA | prof "
          f"{min(profs):.1f}-{max(profs):.1f} m" if profs else f"[mapa7] {MUNIC}: {n} furos")

    # ---- extensao com margem (zoom mais fechado p/ espalhar os furos) ----
    xmin, xmax = xs.min(), xs.max(); ymin, ymax = ys.min(), ys.max()
    spanx, spany = xmax - xmin, ymax - ymin
    span = max(spanx, spany, 200.0)
    mx = max(span * 0.08, 80.0)
    # centraliza num quadro quadrado p/ aspecto equilibrado
    cx, cy = (xmin + xmax) / 2, (ymin + ymax) / 2
    half = span / 2 + mx
    xmin, xmax = cx - half, cx + half
    ymin, ymax = cy - half, cy + half

    # ---- figura A4 paisagem ----
    fig = plt.figure(figsize=(11.69, 8.27), dpi=300)
    ax = fig.add_axes(C.map_axes_rect())
    ax.set_xlim(xmin, xmax); ax.set_ylim(ymin, ymax)
    ax.set_aspect("equal", adjustable="box")

    # ---- basemap satelite Esri ----
    src = C.add_basemap(ax, provider="imagery")

    # ---- malha UTM + moldura (passo p/ ~5 marcas, mesmas E e N) ----
    full = xmax - xmin
    step = C._nice_step(full)
    while full / step > 6:
        step *= 2
    while full / step < 3:
        step /= 2
    C.utm_grid(ax, xmin, xmax, ymin, ymax, step=step, grid_alpha=0.5)
    C.frame(ax)

    # ---- furos ----
    COR_SECO = "#FFD23F"   # amarelo — sem nivel d'agua
    COR_AGUA = "#27A0E6"   # azul — com nivel d'agua
    from matplotlib import patheffects as _pe
    for f in furos:
        cor = COR_AGUA if f["agua"] else COR_SECO
        # marcador circular menor (nao cobre o rotulo)
        ax.plot(f["x"], f["y"], marker="o", ms=5.6, mfc=cor, mec=C.INK, mew=0.9,
                zorder=18, clip_on=True)
        ax.plot(f["x"], f["y"], marker="o", ms=1.7, mfc=C.INK, mec="none",
                zorder=19, clip_on=True)  # miolo
        # rotulo: numero do furo (halo branco), logo acima do ponto
        lbl = f["num"] if f["num"] else "?"
        ax.annotate(lbl, (f["x"], f["y"]), xytext=(0, 5.5), textcoords="offset points",
                    ha="center", va="bottom", fontsize=C.FS_SMALL - 1.0,
                    fontweight="bold", color=C.INK, zorder=20,
                    path_effects=[_pe.withStroke(linewidth=2.0, foreground="white")])

    # ---- mobilia cartografica ----
    C.north_arrow(ax)
    C.scale_bar(ax, xmin, xmax, ymin, ymax, frac=0.38, loc="lower right", y_frac=0.035)
    C.title_block(fig, "MAPA DE LOCAÇÃO DAS SONDAGENS (SPT)",
                  f"Investigação Geotécnica · {MUNIC} — PR",
                  project=f"REDE COLETORA DE ESGOTO · {_norm(MUNIC).title()} — PR")

    # ---- legenda de rodape ----
    handles = [
        Line2D([0], [0], marker="o", ls="none", mfc=COR_SECO, mec=C.INK, mew=0.9,
               ms=8, label=f"Furo de sondagem — sem nível d'água ({n - n_agua})"),
        Line2D([0], [0], marker="o", ls="none", mfc=COR_AGUA, mec=C.INK, mew=0.9,
               ms=8, label=f"Furo de sondagem — com nível d'água ({n_agua})"),
    ]
    extra = (f"INVESTIGAÇÃO GEOTÉCNICA\n{n} furos (SPT) · NBR 6484\n"
             "SIRGAS 2000 · UTM 22 S · MC 51° W")
    C.footer_legend(fig, handles, title="LEGENDA", ncol=1, extra_right=extra, map_ax=ax)
    C.credits(fig, fonte=(src or "Esri World Imagery"))

    os.makedirs(C.OUT, exist_ok=True)
    fig.savefig(OUT_PNG, dpi=300)
    plt.close(fig)
    print("[mapa7] salvo:", OUT_PNG)
    return OUT_PNG


if __name__ == "__main__":
    main()
