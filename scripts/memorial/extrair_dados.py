# -*- coding: utf-8 -*-
"""
extrair_dados.py — Pipeline PARAMETRICO de extracao dos dados brutos de um
projeto de Rede Coletora de Esgoto (RCE) para o gerador de Memorial Descritivo.

Recebe um CONFIG (dict ou caminho de JSON) com os caminhos dos arquivos brutos
e PRODUZ, num diretorio de saida:
    - dados.json          (quantitativos consolidados — DADOS_JSON)
    - dados_extra.json    (TQ/degrau U12, DN, precisoes GNSS — DADOS_EXTRA_JSON)
    - geo/rede_trechos.geojson  (rede projetada, EPSG do projeto)
    - geo/rede_trechos.csv
    - geo/estruturas_pv_tl.csv
    - geo/soleiras.csv
    - geo/soleiras_shp/    (copia do shapefile de soleiras, usado pelos mapas)

Toda a logica antes hardcoded para Amapora foi refatorada para ler do config,
SEM perder nenhuma regra ja implementada:
    - TQ / DEGRAU pela celula U12 de cada aba OSE-NNN ('T.Q. X,XXX m' / 'DEGRAU X,XXX m')
    - classificacao MND / VCA por PROFUNDIDADE > 3,00 m (nao pelo nome do arquivo)
    - soleiras filtradas espacialmente pela bacia (envoltoria da rede + buffer)
    - precisoes GNSS (HRMS/VRMS/PDOP/...) com filtro de outliers
    - drenagem com geometria corrigida (read_lines_fixed do cartolib)
    - tabela DN automatica

Uso:
    python extrair_dados.py --config config.json
    python extrair_dados.py --config config.json --out C:/saida   # sobrescreve out_dir

Tambem importavel:  from extrair_dados import extrair ; extrair(cfg)
"""
import os
import sys
import json
import csv
import glob
import math
import zipfile
import shutil
import tempfile
import argparse
from collections import Counter, defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import openpyxl
import shapefile  # pyshp
import re as _re
import unicodedata


# ============================================================================
#  Helpers de descoberta recursiva + auto-classificacao de shapes
# ============================================================================
def _norm(s):
    """minusculas, sem acento — p/ casar nomes ('agua'=='água')."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower()


def _walk_shps(root):
    """Varre root (recursivo) e devolve listas de shapefiles (.shp) e zips
    encontrados. Caminhos absolutos."""
    shps, zips = [], []
    for dp, _dn, fns in os.walk(root):
        for fn in fns:
            low = fn.lower()
            full = os.path.join(dp, fn)
            if low.endswith(".shp"):
                shps.append(full)
            elif low.endswith(".zip"):
                zips.append(full)
    return shps, zips


def _expand_zip_shps(zips, tmp_root, tag):
    """Extrai cada zip num subdir do tmp e devolve os .shp de dentro."""
    out = []
    for i, zp in enumerate(zips):
        try:
            dst = _ensure_dir(os.path.join(tmp_root, "_zip_%s_%d" % (tag, i)))
            with zipfile.ZipFile(zp) as z:
                z.extractall(dst)
            out += glob.glob(os.path.join(dst, "**", "*.shp"), recursive=True)
        except Exception as e:
            print("  [aviso] falha ao abrir zip %s: %s" % (zp, e))
    return out


def _gather_shps(path, tmp_root, tag, expand_zips=True):
    """Resolve um caminho informado (PASTA, .shp ou .zip) numa lista de .shp,
    fazendo busca RECURSIVA quando for pasta. Retrocompativel:
      - .shp  -> [esse shp]
      - .zip  -> shapes de dentro do zip
      - pasta -> todos os .shp das subpastas + shapes dos .zip encontrados
    """
    if not path:
        return []
    low = str(path).lower()
    if low.endswith(".shp"):
        return [path] if os.path.exists(path) else []
    if low.endswith(".zip"):
        return _expand_zip_shps([path], tmp_root, tag) if os.path.exists(path) else []
    if os.path.isdir(path):
        shps, zips = _walk_shps(path)
        if expand_zips and zips:
            shps = shps + _expand_zip_shps(zips, tmp_root, tag)
        return shps
    return []


def _classify_shp(shp_path, keywords_all=None, keywords_any=None, keywords_not=None):
    """Pontua um .shp pelo NOME (sem acento, minusculo). keywords_all: todas
    precisam estar presentes; keywords_any: pelo menos uma (cada match soma);
    keywords_not: penaliza/exclui. Retorna score (None se nao casa o minimo)."""
    name = _norm(os.path.splitext(os.path.basename(shp_path))[0])
    if keywords_not:
        for kw in keywords_not:
            if _norm(kw) in name:
                return None
    score = 0
    if keywords_all:
        for kw in keywords_all:
            if _norm(kw) not in name:
                return None
            score += 2
    if keywords_any:
        hit = False
        for kw in keywords_any:
            if _norm(kw) in name:
                score += 1
                hit = True
        if keywords_any and not hit and not keywords_all:
            return None
    return score


def _pick_best(shps, label, keywords_all=None, keywords_any=None, keywords_not=None):
    """Escolhe o .shp mais provavel da categoria. Loga candidatos."""
    cands = []
    for sp in shps:
        sc = _classify_shp(sp, keywords_all, keywords_any, keywords_not)
        if sc is not None:
            cands.append((sc, sp))
    if not cands:
        return None
    cands.sort(key=lambda t: (-t[0], len(t[1])))
    best = cands[0][1]
    if len(cands) > 1:
        print("  [classif] %s: %d candidato(s) -> escolhido '%s'" %
              (label, len(cands), os.path.basename(best)))
        for sc, sp in cands:
            mark = "*" if sp == best else " "
            print("           %s [score %d] %s" % (mark, sc, sp))
    else:
        print("  [classif] %s -> '%s'" % (label, os.path.basename(best)))
    return best


# ============================================================================
#  Helpers de config / caminhos
# ============================================================================
def _get(cfg, *keys, default=None):
    """Le cfg aceitando varios nomes de chave (aliases)."""
    for k in keys:
        if k in cfg and cfg[k] not in (None, ""):
            return cfg[k]
    return default


def _ensure_dir(p):
    os.makedirs(p, exist_ok=True)
    return p


# ============================================================================
#  Geometria (convex hull + buffer + ponto-em-poligono + dist a segmento)
# ============================================================================
def convex_hull(points):
    pts = sorted(set(map(tuple, points)))
    if len(pts) <= 2:
        return pts

    def cross(o, a, b):
        return (a[0] - o[0]) * (b[1] - o[1]) - (a[1] - o[1]) * (b[0] - o[0])

    lo = []
    for p in pts:
        while len(lo) >= 2 and cross(lo[-2], lo[-1], p) <= 0:
            lo.pop()
        lo.append(p)
    up = []
    for p in reversed(pts):
        while len(up) >= 2 and cross(up[-2], up[-1], p) <= 0:
            up.pop()
        up.append(p)
    return lo[:-1] + up[:-1]


def buffer_hull(hull, d):
    if not hull:
        return hull
    cx = sum(p[0] for p in hull) / len(hull)
    cy = sum(p[1] for p in hull) / len(hull)
    out = []
    for x, y in hull:
        dx, dy = x - cx, y - cy
        L = math.hypot(dx, dy) or 1
        out.append((x + dx / L * d, y + dy / L * d))
    return out


def _pt_in_poly(px, py, poly):
    inside = False
    n = len(poly)
    j = n - 1
    for i in range(n):
        xi, yi = poly[i]
        xj, yj = poly[j]
        if ((yi > py) != (yj > py)) and \
           (px < (xj - xi) * (py - yi) / ((yj - yi) or 1e-12) + xi):
            inside = not inside
        j = i
    return inside


def _dist_pt_segs(px, py, segs):
    best = 1e18
    for seg in segs:
        for k in range(len(seg) - 1):
            ax_, ay_ = seg[k]
            bx_, by_ = seg[k + 1]
            dx, dy = bx_ - ax_, by_ - ay_
            L2 = dx * dx + dy * dy
            if L2 == 0:
                d = math.hypot(px - ax_, py - ay_)
            else:
                t = max(0.0, min(1.0, ((px - ax_) * dx + (py - ay_) * dy) / L2))
                cx, cy = ax_ + t * dx, ay_ + t * dy
                d = math.hypot(px - cx, py - cy)
            if d < best:
                best = d
    return best


# ============================================================================
#  1) OSE.xlsx  ->  trechos, PVs/TLs, DN, metodo, TQ/degrau (U12), MND/VCA
# ============================================================================
def _parse_u12(val):
    """Interpreta a celula U12 de uma aba OSE.
    Retorna ('TQ'|'DEGRAU'|None, altura_float|None).
    Formatos: 'T.Q. 1,168 m', 'DEGRAU 0,005 m', 'DEGRAU  m' (vazio)."""
    if val is None:
        return None, None
    s = str(val).strip()
    if not s:
        return None, None
    up = s.upper()
    kind = None
    if "T.Q" in up or up.startswith("TQ") or "TUBO DE QUEDA" in up:
        kind = "TQ"
    elif "DEGRAU" in up:
        kind = "DEGRAU"
    else:
        return None, None
    # extrai numero (aceita virgula ou ponto)
    import re
    m = re.search(r"(-?\d+(?:[.,]\d+)?)", up.replace("T.Q.", "").replace("DEGRAU", ""))
    if not m:
        return kind, None   # rotulo presente mas sem valor (template vazio)
    try:
        h = float(m.group(1).replace(",", "."))
    except Exception:
        return kind, None
    return kind, h


def extrair_ose(xlsx_path):
    """Extrai tudo da planilha de OSEs. Retorna um dict com os blocos crus."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ---------- RESUMO ----------
    ose_list = []
    total_row = None
    if "RESUMO" in wb.sheetnames:
        ws = wb["RESUMO"]
        r = 4
        while r <= 400:
            a = ws.cell(r, 1).value
            if isinstance(a, str) and a.strip().upper() == "TOTAL":
                total_row = dict(ext=ws.cell(r, 2).value, pv=ws.cell(r, 3).value,
                                 tl=ws.cell(r, 4).value, pit=ws.cell(r, 5).value)
                break
            if isinstance(a, str) and a.startswith("OSE-"):
                ose_list.append(dict(ose=a.strip(), ext=ws.cell(r, 2).value,
                                     pv=ws.cell(r, 3).value, tl=ws.cell(r, 4).value,
                                     pit=ws.cell(r, 5).value))
            r += 1
        # DETALHAMENTO (col 'C. Topo' marca o inicio)
        det_start = None
        for rr in range(1, ws.max_row + 1):
            if ws.cell(rr, 4).value == "C. Topo":
                det_start = rr + 1
                break
        detalhamento = []
        if det_start:
            for rr in range(det_start, ws.max_row + 1):
                ose = ws.cell(rr, 1).value
                tipo = ws.cell(rr, 2).value
                if not (ose and tipo):
                    continue
                detalhamento.append(dict(
                    ose=str(ose).strip(), tipo=str(tipo).strip(),
                    nome=str(ws.cell(rr, 3).value).strip() if ws.cell(rr, 3).value else None,
                    ctopo=ws.cell(rr, 4).value, cfundo=ws.cell(rr, 5).value,
                    prof=ws.cell(rr, 6).value))
    else:
        detalhamento = []

    # ---------- abas OSE-NNN ----------
    DN_set = Counter()
    method_set = Counter()
    pav_set = Counter()
    points = {}
    trechos = []
    ose_meta = []
    tq_list = []
    deg_list = []
    # VCA pela regra do projeto: extensao (comprimento da aba OSE) cujo MAIOR
    # prof do trecho > 3,00 m. Calculado DIRETO das abas OSE-NNN (independe do
    # RESUMO), portanto cobre TODAS as bacias mescladas.
    ext_vca_sheets = 0.0
    ext_total_sheets = 0.0
    ose_sheets = [s for s in wb.sheetnames if s.startswith("OSE-")]
    for sn in ose_sheets:
        s = wb[sn]
        dn_b9 = s.cell(9, 2).value
        bacia = s.cell(5, 2).value
        metodo = s.cell(6, 2).value
        comp = s.cell(6, 3).value
        rua = s.cell(7, 2).value
        pav = s.cell(8, 2).value
        bairro = s.cell(4, 2).value
        if dn_b9:
            DN_set[str(dn_b9).strip()] += 1
        if metodo:
            method_set[str(metodo).strip()] += 1
        if pav:
            pav_set[str(pav).strip()] += 1
        ose_meta.append(dict(ose=sn, dn=str(dn_b9).strip() if dn_b9 else None,
                             bacia=bacia, metodo=str(metodo).strip() if metodo else None,
                             comprimento=comp, rua=rua, pavimento=pav, bairro=bairro))

        # TQ / DEGRAU da celula U12 (coluna 21, linha 12)
        kind, h = _parse_u12(s.cell(12, 21).value)
        if kind == "TQ" and h is not None:
            tq_list.append(dict(ose=sn, altura_m=round(h, 3)))
        elif kind == "DEGRAU" and h is not None:
            deg_list.append(dict(ose=sn, altura_m=round(h, 3)))

        # geometria (sequencia de PVs/TLs a partir da linha 11)
        seq = []
        sheet_max_prof = 0.0
        for rr in range(11, s.max_row + 1):
            nome = s.cell(rr, 1).value
            x = s.cell(rr, 2).value
            y = s.cell(rr, 3).value
            if nome is None or x is None or y is None:
                continue
            nome = str(nome).strip()
            try:
                x = float(x); y = float(y)
            except Exception:
                continue
            z = s.cell(rr, 4).value
            prof = s.cell(rr, 19).value
            if isinstance(prof, (int, float)) and prof > sheet_max_prof:
                sheet_max_prof = prof
            tipo = ("PV" if nome.startswith("PV") else
                    ("TL" if nome.startswith("TL") else
                     ("PIT" if nome.startswith("PIT") else "OUT")))
            if tipo in ("PV", "TL") and nome not in points:
                points[nome] = dict(name=nome, x=x, y=y, z=z, prof=prof, tipo=tipo, ose=sn)
            if not seq or seq[-1]["name"] != nome:
                seq.append(dict(name=nome, x=x, y=y))
        # acumula VCA (prof > 3 m) pela extensao da propria aba
        if isinstance(comp, (int, float)):
            ext_total_sheets += comp
            if sheet_max_prof > 3.00:
                ext_vca_sheets += comp
        if len(seq) >= 2:
            trechos.append(dict(ose=sn, dn=str(dn_b9).strip() if dn_b9 else None,
                                comprimento=comp,
                                vertices=[(round(p["x"], 3), round(p["y"], 3)) for p in seq],
                                montante=seq[0]["name"], jusante=seq[-1]["name"]))

    # ---------- faixas de profundidade ----------
    faixas = {"ate_1_25": 0, "1_25_a_2_00": 0, "2_00_a_3_00": 0,
              "3_00_a_4_00": 0, "acima_4_00": 0}
    for d in detalhamento:
        p = d["prof"]
        if isinstance(p, (int, float)):
            if p <= 1.25:
                faixas["ate_1_25"] += 1
            elif p <= 2.00:
                faixas["1_25_a_2_00"] += 1
            elif p <= 3.00:
                faixas["2_00_a_3_00"] += 1
            elif p <= 4.00:
                faixas["3_00_a_4_00"] += 1
            else:
                faixas["acima_4_00"] += 1

    # ---------- MND / VCA por profundidade > 3,00 m ----------
    prof_by_ose = defaultdict(list)
    for d in detalhamento:
        if isinstance(d["prof"], (int, float)):
            prof_by_ose[d["ose"]].append(d["prof"])
    ext_by_ose = {o["ose"]: o["ext"] for o in ose_list}
    ext_mnd = 0.0
    ext_vca = 0.0
    n_ose_vca = 0
    for ose, ext in ext_by_ose.items():
        if ext is None:
            continue
        pmax = max(prof_by_ose.get(ose, [0]) or [0])
        if pmax > 3.00:
            ext_vca += ext
            n_ose_vca += 1
        else:
            ext_mnd += ext

    ext_by_dn = defaultdict(float)
    for m in ose_meta:
        if m["comprimento"] and m["dn"]:
            ext_by_dn[m["dn"]] += m["comprimento"]

    return dict(
        ose_list=ose_list, total_row=total_row, detalhamento=detalhamento,
        DN_set=DN_set, method_set=method_set, pav_set=pav_set,
        points=points, trechos=trechos, ose_meta=ose_meta,
        tq_list=tq_list, deg_list=deg_list,
        faixas=faixas, ext_mnd=ext_mnd, ext_vca=ext_vca, n_ose_vca=n_ose_vca,
        ext_vca_sheets=round(ext_vca_sheets, 2),
        ext_total_sheets=round(ext_total_sheets, 2),
        ext_by_dn=dict(ext_by_dn),
        arquivo_fonte=os.path.basename(xlsx_path))


def extrair_ose_multi(xlsx_paths):
    """Extrai e MESCLA varias planilhas de OSE num unico bloco cru, no mesmo
    formato de extrair_ose(). Concatena listas, soma contadores/quantitativos e
    deduplica PVs/TLs por nome (mantem o 1o). Aceita um caminho str ou lista."""
    if isinstance(xlsx_paths, str):
        xlsx_paths = [xlsx_paths]
    parts = [extrair_ose(p) for p in xlsx_paths]
    if len(parts) == 1:
        return parts[0]

    merged = dict(
        ose_list=[], total_row=None, detalhamento=[],
        DN_set=Counter(), method_set=Counter(), pav_set=Counter(),
        points={}, trechos=[], ose_meta=[], tq_list=[], deg_list=[],
        faixas={"ate_1_25": 0, "1_25_a_2_00": 0, "2_00_a_3_00": 0,
                "3_00_a_4_00": 0, "acima_4_00": 0},
        ext_mnd=0.0, ext_vca=0.0, n_ose_vca=0, ext_by_dn=defaultdict(float),
        ext_vca_sheets=0.0, ext_total_sheets=0.0,
        arquivo_fonte=" + ".join(os.path.basename(p) for p in xlsx_paths))

    tr_sum = {"ext": 0, "pv": 0, "tl": 0, "pit": 0}
    tem_total = False
    for O in parts:
        merged["ose_list"].extend(O["ose_list"])
        merged["detalhamento"].extend(O["detalhamento"])
        merged["trechos"].extend(O["trechos"])
        merged["ose_meta"].extend(O["ose_meta"])
        merged["tq_list"].extend(O["tq_list"])
        merged["deg_list"].extend(O["deg_list"])
        merged["DN_set"].update(O["DN_set"])
        merged["method_set"].update(O["method_set"])
        merged["pav_set"].update(O["pav_set"])
        for nome, p in O["points"].items():
            merged["points"].setdefault(nome, p)
        for k in merged["faixas"]:
            merged["faixas"][k] += O["faixas"].get(k, 0)
        merged["ext_mnd"] += O["ext_mnd"]
        merged["ext_vca"] += O["ext_vca"]
        merged["n_ose_vca"] += O["n_ose_vca"]
        merged["ext_vca_sheets"] += O.get("ext_vca_sheets", 0.0)
        merged["ext_total_sheets"] += O.get("ext_total_sheets", 0.0)
        for dn, v in O["ext_by_dn"].items():
            merged["ext_by_dn"][dn] += v
        if O.get("total_row"):
            tem_total = True
            for k in tr_sum:
                v = O["total_row"].get(k)
                if isinstance(v, (int, float)):
                    tr_sum[k] += v
    merged["ext_by_dn"] = dict(merged["ext_by_dn"])
    merged["total_row"] = tr_sum if tem_total else None
    # PV/TL do total mesclado: o somatorio por planilha DUPLICA os dispositivos
    # que aparecem em mais de uma OSE (o mesmo PV e limite de duas OSEs vizinhas).
    # Recontamos pelos pontos DEDUPLICADOS por nome (merged["points"]), que e a
    # contagem UNICA correta. (Lucas 07/07)
    if merged["total_row"] is not None:
        n_pv_u = sum(1 for p in merged["points"].values() if p.get("tipo") == "PV")
        n_tl_u = sum(1 for p in merged["points"].values() if p.get("tipo") == "TL")
        if n_pv_u:
            merged["total_row"]["pv"] = n_pv_u
        if n_tl_u:
            merged["total_row"]["tl"] = n_tl_u
    return merged


# ============================================================================
#  2) Soleiras (shp ou zip) -> linhas + filtro espacial pela bacia
# ============================================================================
def _open_soleiras_shp(soleiras_path, tmp_root):
    """Aceita .shp direto, .zip (extrai), ou PASTA (busca recursiva). Quando ha
    varios .shp, auto-classifica pelo nome (contem 'sol'/'soleira').
    Retorna (reader, shp_basepath)."""
    shps = _gather_shps(soleiras_path, tmp_root, "sol")
    if not shps:
        raise FileNotFoundError(
            "Nenhum shapefile de soleiras encontrado em: %s" % soleiras_path)
    base = None
    if len(shps) == 1:
        base = shps[0][:-4]
        print("  [classif] soleiras -> '%s'" % os.path.basename(shps[0]))
    else:
        # auto-classifica: nome contem 'soleira' (mais especifico) ou 'sol'
        pick = (_pick_best(shps, "soleiras", keywords_any=["soleira"]) or
                _pick_best(shps, "soleiras", keywords_any=["sol"]))
        if pick is None:
            pick = shps[0]
            print("  [aviso] soleiras: nenhum nome casou 'sol/soleira'; "
                  "usando 1o .shp encontrado: %s" % os.path.basename(pick))
        base = pick[:-4]
    return shapefile.Reader(base), base


def extrair_soleiras(soleiras_path, rede_segs, pv_xy, tmp_root, serve_buf=90.0):
    """Le soleiras e devolve (todas as linhas, contagem_total_bacia,
    n_pos_bacia, n_neg_bacia, idx_pct, shp_base_para_copia).
    Filtro de bacia = ponto dentro da envoltoria da rede (buffer serve_buf) OU
    a <= serve_buf de qualquer trecho da rede projetada."""
    sr, base = _open_soleiras_shp(soleiras_path, tmp_root)
    flds = [f[0] for f in sr.fields[1:]]
    sol_rows = []
    for i in range(len(sr)):
        d = dict(zip(flds, list(sr.record(i))))
        st = str(d.get("STATUS", "")).strip()
        sol_rows.append(dict(x=d.get("X"), y=d.get("Y"), status=st,
                             edificacao=d.get("EDIFICACAO"), dif_cota=d.get("DIF_COTA"),
                             frente=d.get("FRENTE"), local=str(d.get("LOCAL", "")).strip()))

    # ---- filtro espacial pela bacia (envoltoria da rede + PVs) ----
    hull = convex_hull([(x, y) for (x, y) in pv_xy] + [v for seg in rede_segs for v in seg])
    hull_serve = buffer_hull(hull, serve_buf)

    def in_bacia(x, y):
        try:
            x = float(x); y = float(y)
        except Exception:
            return False
        return _pt_in_poly(x, y, hull_serve) or _dist_pt_segs(x, y, rede_segs) <= serve_buf

    n_pos = n_neg = 0
    for r in sol_rows:
        st = str(r["status"]).strip().upper()
        if not in_bacia(r["x"], r["y"]):
            continue
        if st.startswith("SOLEIRA-POSITIVA"):
            n_pos += 1
        elif st.startswith("SOLEIRA-NEGATIVA"):
            n_neg += 1
    n_tot = n_pos + n_neg
    idx = round(100.0 * n_pos / n_tot, 2) if n_tot else None
    return sol_rows, n_tot, n_pos, n_neg, idx, base


# ============================================================================
#  3) Topografia (TXT GNSS) -> contagem, cotas, bbox, precisoes
# ============================================================================
def extrair_topografia(txt_dir, txt_glob="*.txt"):
    """Le os TXT GNSS. Colunas fixas: 0 PONTO,1 DESC,2 N(Y),3 E(X),4 Z.
    Os campos GNSS aceitam DOIS formatos:
      (a) CHAVE:VALOR a partir da col 5 (ex.: HRMS:0.005,...,PDOP:1.085,STATUS:FIXED)
      (b) posicional legado (5 PDOP,6 HDOP,7 VDOP,8 HRMS,9 VRMS,14 status).
    Busca RECURSIVA (varre subpastas). Filtro de outliers (raio 5000 m do
    centroide + |z-med|<=4*std)."""
    import numpy as np
    # busca recursiva: pega TXT na pasta e em todas as subpastas
    files = []
    for root, _dirs, fnames in os.walk(txt_dir):
        for fn in fnames:
            if fn.lower().endswith(".txt"):
                files.append(os.path.join(root, fn))
    files = sorted(files)
    rows = []
    per_file = []
    for fp in files:
        nf = 0
        zf = []
        with open(fp, encoding="latin-1") as fh:
            for line in fh:
                p = line.rstrip("\n").split(",")
                if len(p) < 5:
                    continue
                try:
                    y = float(p[2]); x = float(p[3]); z = float(p[4])
                except Exception:
                    continue
                rec = dict(x=x, y=y, z=z, file=os.path.basename(fp))

                # CHAVE:VALOR (formato CHC/LandStar atual)
                kv = {}
                for tok in p[5:]:
                    if ":" in tok:
                        k, v = tok.split(":", 1)
                        kv[k.strip().upper()] = v.strip()

                def fcol(i):
                    try:
                        return float(p[i])
                    except Exception:
                        return None

                def fkv(key, posidx):
                    if key in kv:
                        try:
                            return float(kv[key])
                        except Exception:
                            return None
                    return fcol(posidx)  # fallback posicional legado

                rec["pdop"] = fkv("PDOP", 5)
                rec["hdop"] = fkv("HDOP", 6)
                rec["vdop"] = fkv("VDOP", 7)
                rec["hrms"] = fkv("HRMS", 8)
                rec["vrms"] = fkv("VRMS", 9)
                rec["status"] = kv.get("STATUS") or (p[14].strip() if len(p) > 14 else "")
                rows.append(rec)
                nf += 1
                zf.append(z)
        per_file.append(dict(file=os.path.basename(fp), pontos=nf,
                             z_min=round(min(zf), 3) if zf else None,
                             z_max=round(max(zf), 3) if zf else None))

    if not rows:
        return dict(arquivos=len(files), total_pontos=0, por_arquivo=per_file), [], None

    arr = np.array([(r["x"], r["y"], r["z"]) for r in rows], dtype=float)
    cx, cy = arr[:, 0].mean(), arr[:, 1].mean()
    rad = np.hypot(arr[:, 0] - cx, arr[:, 1] - cy)
    zmed = np.median(arr[:, 2])
    zstd = arr[:, 2].std()
    keep = (rad < 5000.0) & (np.abs(arr[:, 2] - zmed) <= 4 * zstd)
    kr = [r for i, r in enumerate(rows) if keep[i]]

    xs = [r["x"] for r in kr]
    ys = [r["y"] for r in kr]
    zs = [r["z"] for r in kr]

    def stats(vals):
        a = np.array([v for v in vals if v is not None], dtype=float)
        if a.size == 0:
            return {}
        return dict(minimo=round(float(a.min()), 3), maximo=round(float(a.max()), 3),
                    medio=round(float(a.mean()), 3),
                    tipico_mediana=round(float(np.median(a)), 3),
                    p90=round(float(np.percentile(a, 90)), 3))

    hrms = stats([r["hrms"] for r in kr])
    vrms = stats([r["vrms"] for r in kr])
    pdop = stats([r["pdop"] for r in kr])
    hdop = stats([r["hdop"] for r in kr])
    vdop = stats([r["vdop"] for r in kr])
    status_c = Counter(r["status"] for r in kr)
    n_fixo = sum(v for k, v in status_c.items() if k.lower().startswith("fix"))
    n_auto = sum(v for k, v in status_c.items() if not k.lower().startswith("fix"))

    topo = dict(
        arquivos=len(files), total_pontos=len(kr),
        z_min=round(min(zs), 3), z_max=round(max(zs), 3),
        z_media=round(sum(zs) / len(zs), 3),
        bbox=dict(x_min=round(min(xs), 3), x_max=round(max(xs), 3),
                  y_min=round(min(ys), 3), y_max=round(max(ys), 3)),
        por_arquivo=per_file)

    gnss = dict(
        n_pontos=len(kr), n_arquivos=len(files),
        metodo_calculo=("col5=PDOP,col6=HDOP,col7=VDOP,col8=HRMS,col9=VRMS,"
                        "col14=status; filtro outliers raio 5000m + |z-med|<=4*std"),
        metodo_RTK=("PPP-RTK (GEO PPP) - %d pts solucao 'Fixo', %d pt(s) outra solucao"
                    % (n_fixo, n_auto)),
        precisao_planimetrica_m_HRMS=hrms,
        precisao_altimetrica_m_VRMS=vrms,
        PDOP=pdop, VDOP=vdop, HDOP=hdop, HRMS=hrms, VRMS=vrms,
        detalhe_por_arquivo=[dict(arquivo=pf["file"], pontos=pf["pontos"]) for pf in per_file],
        status_contagem=dict(status_c))
    return topo, kr, gnss


# ============================================================================
#  4) Interferencias (shp agua/drenagem) -> contagem de trechos
# ============================================================================
def classificar_interferencias(int_path, tmp_root=None):
    """Resolve o caminho de interferencias (PASTA com subpastas, .shp ou .zip),
    faz busca RECURSIVA e auto-classifica os shapes pelo NOME. Retorna um dict
    com os caminhos absolutos (.shp) de cada categoria (ou None se nao achou):
        {linhas_agua, linhas_drenagem, vertices_agua, vertices_drenagem}
    Regras de nome (sem acento, case-insensitive):
      - agua  : nome contem 'agua'
      - dren  : nome contem 'dren'
      - linha : 'linha'/'line'  (ou simplesmente NAO ser vertice)
      - vert  : 'vertice'/'vert'/'vertices'
    """
    if not int_path:
        return {}
    if tmp_root is None:
        tmp_root = tempfile.mkdtemp(prefix="interf_")
    # se for um unico .shp informado direto, ainda tenta classifica-lo sozinho
    shps = _gather_shps(int_path, tmp_root, "interf")
    if not shps:
        return {}
    print("  [classif] interferencias: %d shapefile(s) varrido(s) (recursivo)" % len(shps))

    def vert_of(sp):
        return any(k in _norm(sp) for k in ("vertice", "vertices", "_vert"))

    linhas = [s for s in shps if not vert_of(s)]
    verts = [s for s in shps if vert_of(s)]

    res = {}
    # LINHAS de agua: tem 'agua', NAO e vertice. Prefere quem tb tem 'linha'.
    res["linhas_agua"] = (
        _pick_best(linhas, "LINHAS_AGUA",
                   keywords_all=["agua"], keywords_any=["linha", "line"], keywords_not=["vert"]) or
        _pick_best(linhas, "LINHAS_AGUA", keywords_all=["agua"], keywords_not=["vert"]))
    # LINHAS de drenagem
    res["linhas_drenagem"] = (
        _pick_best(linhas, "LINHAS_DRENAGEM",
                   keywords_all=["dren"], keywords_any=["linha", "line"], keywords_not=["vert"]) or
        _pick_best(linhas, "LINHAS_DRENAGEM", keywords_all=["dren"], keywords_not=["vert"]))
    # VERTICES (para corrigir geometria via Z)
    res["vertices_agua"] = _pick_best(verts, "VERTICES_AGUA", keywords_all=["agua"])
    res["vertices_drenagem"] = _pick_best(verts, "VERTICES_DRENAGEM", keywords_all=["dren"])
    return res


def extrair_interferencias(int_path, tmp_root=None):
    """Conta feicoes lineares de agua / drenagem (e vertices) — aceita PASTA
    (busca recursiva inclusive subpastas), .shp ou .zip. Auto-classifica os
    shapes pelo nome. Retorna dict do bloco de interferencias + os caminhos
    classificados (chave '_paths') p/ os mapas reusarem; None se nao achar nada.
    Robusto: se faltar uma categoria, segue com a que tem."""
    if not int_path:
        return None

    paths = classificar_interferencias(int_path, tmp_root)
    if not paths:
        if os.path.isdir(int_path) or str(int_path).lower().endswith((".shp", ".zip")):
            print("  [aviso] interferencias: nenhum shapefile encontrado em %s" % int_path)
        return None

    def count(path):
        if not path:
            return 0
        try:
            return len(shapefile.Reader(path[:-4]))
        except Exception as e:
            print("  [aviso] falha ao ler %s: %s" % (path, e))
            return 0

    n_agua = count(paths.get("linhas_agua"))
    n_dren = count(paths.get("linhas_drenagem"))
    if not paths.get("linhas_agua"):
        print("  [aviso] interferencias: LINHAS_AGUA nao encontrada (camada agua omitida)")
    if not paths.get("linhas_drenagem"):
        print("  [aviso] interferencias: LINHAS_DRENAGEM nao encontrada (camada dren omitida)")
    if n_agua == 0 and n_dren == 0:
        return None
    return dict(
        agua_trechos=n_agua, drenagem_trechos=n_dren,
        observacao=("Cruzamentos do tracado projetado com redes existentes de agua (%d "
                    "trechos) e de drenagem (%d trechos), exigindo verificacao de "
                    "cotas/travessias." % (n_agua, n_dren)),
        fonte="2S Engenharia, 2026",
        _paths=paths)


# ============================================================================
#  ESCRITA dos arquivos de saida (geo csv/geojson + JSONs)
# ============================================================================
def _escrever_geo(geo_dir, points, trechos, sol_rows, sol_shp_base, epsg):
    _ensure_dir(geo_dir)
    # estruturas PV/TL
    with open(os.path.join(geo_dir, "estruturas_pv_tl.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "tipo", "X", "Y", "Z", "prof_vala", "ose"])
        for p in sorted(points.values(), key=lambda d: d["name"]):
            w.writerow([p["name"], p["tipo"], p["x"], p["y"], p["z"], p["prof"], p["ose"]])
    # rede CSV
    with open(os.path.join(geo_dir, "rede_trechos.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ose", "dn", "comprimento_m", "montante", "jusante", "seq", "X", "Y"])
        for t in trechos:
            for i, (vx, vy) in enumerate(t["vertices"]):
                w.writerow([t["ose"], t["dn"],
                            round(t["comprimento"], 3) if t["comprimento"] else "",
                            t["montante"], t["jusante"], i, vx, vy])
    # rede GeoJSON
    feats = []
    for t in trechos:
        feats.append(dict(type="Feature",
                          properties=dict(ose=t["ose"], dn=t["dn"],
                                          comprimento_m=round(t["comprimento"], 3) if t["comprimento"] else None,
                                          montante=t["montante"], jusante=t["jusante"]),
                          geometry=dict(type="LineString",
                                        coordinates=[[vx, vy] for vx, vy in t["vertices"]])))
    gj = dict(type="FeatureCollection",
              crs=dict(type="name", properties=dict(name="urn:ogc:def:crs:EPSG::%s" % epsg)),
              features=feats)
    with open(os.path.join(geo_dir, "rede_trechos.geojson"), "w", encoding="utf-8") as f:
        json.dump(gj, f, ensure_ascii=False)
    # soleiras CSV
    with open(os.path.join(geo_dir, "soleiras.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["X", "Y", "status", "edificacao", "dif_cota", "frente", "local"])
        for s in sol_rows:
            w.writerow([s["x"], s["y"], s["status"], s["edificacao"],
                        s["dif_cota"], s["frente"], s["local"]])
    # copia o shapefile de soleiras (usado pelos mapas, se precisarem)
    if sol_shp_base:
        dst = _ensure_dir(os.path.join(geo_dir, "soleiras_shp"))
        for ext in (".shp", ".shx", ".dbf", ".prj", ".cpg", ".sbn", ".sbx"):
            src = sol_shp_base + ext
            if os.path.exists(src):
                shutil.copyfile(src, os.path.join(dst, os.path.basename(src)))


_SHP_EXTS = (".shp", ".shx", ".dbf", ".prj", ".cpg", ".sbn", ".sbx", ".qix")


def _copiar_interferencias(geo_dir, interf_paths):
    """Copia os shapes de interferencia classificados p/ geo/interferencias_shp/
    com nomes CANONICOS (LINHAS_AGUA, LINHAS_DRENAGEM, VERTICES_AGUA,
    VERTICES_DRENAGEM) p/ o mapa5 achar mesmo que a fonte estivesse em zip/tmp.
    Retorna o diretorio destino (ou None)."""
    if not interf_paths:
        return None
    canon = {"linhas_agua": "LINHAS_AGUA", "linhas_drenagem": "LINHAS_DRENAGEM",
             "vertices_agua": "VERTICES_AGUA", "vertices_drenagem": "VERTICES_DRENAGEM"}
    dst = _ensure_dir(os.path.join(geo_dir, "interferencias_shp"))
    copiou = False
    for key, nome in canon.items():
        src_shp = interf_paths.get(key)
        if not src_shp:
            continue
        base = src_shp[:-4]
        for ext in _SHP_EXTS:
            s = base + ext
            if os.path.exists(s):
                shutil.copyfile(s, os.path.join(dst, nome + ext))
                copiou = True
    return dst if copiou else None


# ============================================================================
#  5) Modelo hidraulico SewerGEMS (.sqlite/.stsw) -> parametros de projeto
# ============================================================================
def extrair_modelo_sqlite(db_path, scenario_id=None):
    """Le a rede do modelo SewerGEMS (banco SQLite, schema Bentley/Haestad) SEM
    OpenFlows/licenca. Unidades internas em PES (x0,3048 -> m). Resolve a
    alternativa do cenario ativo (raiz por padrao). Retorna um dict com os
    parametros de projeto (Manning/material/Dmin/recobrimento/declividade min)
    + extensao e contagens, ou None se nao conseguir ler. NUNCA levanta excecao
    (a geracao do memorial continua com [A PREENCHER] em caso de falha)."""
    import sqlite3
    FT = 0.3048
    try:
        con = sqlite3.connect(db_path)
        con.text_factory = lambda b: b.decode("utf-8", "replace")
        cur = con.cursor()

        def _has(t):
            return bool(cur.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (t,)).fetchone())

        def _cols(t):
            return [c[1] for c in cur.execute('PRAGMA table_info("%s")' % t)]

        if scenario_id:
            scen = scenario_id
        else:
            row = cur.execute("SELECT ScenarioID FROM HMIScenario WHERE IsDeleted=0 "
                              "AND ParentID IS NULL ORDER BY ScenarioID LIMIT 1").fetchone()
            scen = row[0] if row else None
        active_alts = set(r[0] for r in cur.execute(
            "SELECT AlternativeID FROM HMIScenarioAlternative WHERE ScenarioID=?", (scen,)))

        def load(table, fields, key="DomainElementID"):
            out = {}
            if not _has(table):
                return out
            c = _cols(table)
            sel = [f for f in fields if f in c]
            if key not in c or not sel:
                return out
            has_alt = "AlternativeID" in c
            q = 'SELECT %s,%s%s FROM "%s"' % (
                key, ",".join(sel), ",AlternativeID" if has_alt else "", table)
            for row in cur.execute(q):
                if has_alt and row[-1] not in active_alts:
                    continue
                out[row[0]] = dict(zip(sel, row[1:1 + len(sel)]))
            return out

        plink = load("PhysicalLink_Physical_Data",
                     ["Physical_UpstreamInvert", "Physical_DownstreamInvert"])
        cond = load("Conduit_Physical_Data", ["ConduitDiameter"])
        geom = load("BaseLink_HmiDataSetGeometry_Data", ["HMIGeometryScaledLength"])
        glink = load("GravityLinkBase_Physical_Data", ["Physical_ManningsN", "Physical_Material"])
        ext = load("Conduit_HMIUserDefinedExtensions_Data", ["RECOBRIMNTO"])
        invert = load("GravityNode_Physical_Data", ["Physical_InvertElevation"])

        slopes, recobr, mannings, materiais, diams = [], [], [], [], []
        ext_total = 0.0
        n_trechos = 0
        for did, t in geom.items():
            L = t.get("HMIGeometryScaledLength")
            if L:
                ext_total += L * FT
            n_trechos += 1
            up = plink.get(did, {}).get("Physical_UpstreamInvert")
            dn = plink.get(did, {}).get("Physical_DownstreamInvert")
            if up is not None and dn is not None and L:
                s = abs(up - dn) / L
                if s > 0:
                    slopes.append(round(s, 5))
            r = ext.get(did, {}).get("RECOBRIMNTO")
            if isinstance(r, (int, float)) and r > 0:
                recobr.append(round(r, 3))
            mn = glink.get(did, {}).get("Physical_ManningsN")
            if isinstance(mn, (int, float)) and mn > 0:
                mannings.append(round(mn, 4))
            mat = glink.get(did, {}).get("Physical_Material")
            if mat:
                materiais.append(str(mat).strip())
            dia = cond.get(did, {}).get("ConduitDiameter")
            if dia:
                diams.append(round(dia * FT * 1000))
        con.close()

        if n_trechos == 0:
            return None

        def _mode(lst):
            if not lst:
                return None
            return Counter(lst).most_common(1)[0][0]

        return dict(
            fonte=os.path.basename(db_path), scenario=scen,
            n_trechos=n_trechos, n_nos=len(invert),
            extensao_total_m=round(ext_total, 2) if ext_total else None,
            declividade_min_m_m=min(slopes) if slopes else None,
            declividade_max_m_m=max(slopes) if slopes else None,
            recobrimento_min_m=min(recobr) if recobr else None,
            manning=_mode(mannings),
            material=_mode(materiais),
            dn_min_mm=min(diams) if diams else None,
            dn_max_mm=max(diams) if diams else None,
            dn_lista=sorted(set(diams)))
    except Exception as e:
        sys.stderr.write("[modelo] falha ao ler SQLite (%s): %s\n" % (db_path, e))
        return None


# ============================================================================
#  ORQUESTRADOR
# ============================================================================
def extrair(cfg):
    """Roda o pipeline completo. cfg = dict (ou ja carregado de JSON).
    Retorna dict de resumo com os caminhos gerados e os numeros principais."""
    # ---- caminhos do config ----
    ose_xlsx = _get(cfg, "ose", "ose_xlsx")
    txt_dir = _get(cfg, "txt_dir", "topografia_dir")
    soleiras_path = _get(cfg, "soleiras", "soleiras_path")
    int_dir = _get(cfg, "interferencias", "interferencias_dir")
    out_dir = _get(cfg, "out_dir", "out", "saida_dir") or os.path.dirname(
        os.path.abspath(_get(cfg, "saida", default=".") or "."))
    epsg = str(_get(cfg, "epsg", "codigo_epsg", default="31982"))
    serve_buf = float(_get(cfg, "soleiras_serve_buf_m", default=90.0))

    proj = cfg.get("projeto") or {}
    municipio = _get(cfg, "municipio") or proj.get("municipio") or "—"
    uf = _get(cfg, "uf") or proj.get("uf") or ""
    subbacia = _get(cfg, "subbacia", "sub_bacia") or proj.get("subbacia") or ""
    bacia = _get(cfg, "bacia") or subbacia or ""
    codigo_ibge = _get(cfg, "codigo_ibge", "ibge") or proj.get("codigo_ibge") or ""
    datum_txt = _get(cfg, "datum", default="SIRGAS 2000 / UTM Zone 22S (EPSG:%s)" % epsg)

    # ose pode ser um caminho unico (str) ou uma lista de planilhas a mesclar.
    ose_paths = [ose_xlsx] if isinstance(ose_xlsx, str) else list(ose_xlsx or [])
    if not ose_paths:
        raise FileNotFoundError("OSE.xlsx nao informado.")
    faltando = [p for p in ose_paths if not (p and os.path.exists(p))]
    if faltando:
        raise FileNotFoundError("OSE.xlsx nao encontrado: %s" % ", ".join(faltando))

    _ensure_dir(out_dir)
    geo_dir = _ensure_dir(os.path.join(out_dir, "geo"))
    tmp_root = _ensure_dir(os.path.join(out_dir, "_tmp_extracao"))

    # ---- 1) OSE ----
    print("[extrair] OSE:", " + ".join(ose_paths))
    O = extrair_ose_multi(ose_paths)
    rede_segs = [t["vertices"] for t in O["trechos"]]
    pv_xy = [(p["x"], p["y"]) for p in O["points"].values()]
    n_pv_pts = sum(1 for p in O["points"].values() if p["tipo"] == "PV")
    n_tl_pts = sum(1 for p in O["points"].values() if p["tipo"] == "TL")

    # ---- 2) Soleiras ----
    sol_rows = []
    sol_block = None
    sol_shp_base = None
    if soleiras_path and (os.path.exists(soleiras_path) or os.path.isdir(soleiras_path)):
        print("[extrair] Soleiras:", soleiras_path)
        sol_rows, n_tot, n_pos, n_neg, idx, sol_shp_base = extrair_soleiras(
            soleiras_path, rede_segs, pv_xy, tmp_root, serve_buf=serve_buf)
        sol_block = dict(
            total_imoveis=n_tot, atendidas_positiva=n_pos, nao_atendidas_negativa=n_neg,
            indice_atendimento_pct=idx,
            escopo=("%s%s" % (bacia, (" (sub-bacia %s)" % subbacia) if subbacia and subbacia != bacia else "")) or "—",
            campo_classificacao=("STATUS (SOLEIRA-POSITIVA / SOLEIRA-NEGATIVA); "
                                 "DIF_COTA = cota soleira - frente"),
            fonte=soleiras_path, datum=datum_txt,
            observacao=("Numeros referentes a bacia do projeto (%d imoveis): %d atendidas / "
                        "%d nao atendidas; indice %s%%." %
                        (n_tot, n_pos, n_neg,
                         ("{:.1f}".format(idx).replace(".", ",") if idx is not None else "—"))))

    # ---- 3) Topografia ----
    topo_block = None
    gnss_block = None
    if txt_dir and os.path.isdir(txt_dir):
        print("[extrair] Topografia:", txt_dir)
        topo_block, _kr, gnss_block = extrair_topografia(txt_dir)
        topo_block["datum"] = "SIRGAS 2000 / UTM Zone 22S (EPSG:%s)" % epsg
        topo_block["formato"] = "PONTO,DESCRICAO,NORTHING(Y),EASTING(X),Z,+campos GNSS"
        if gnss_block is not None:
            gnss_block["fonte"] = txt_dir

    # ---- 4) Interferencias (PASTA recursiva / .shp / .zip, auto-classificado) ----
    print("[extrair] Interferencias:", int_dir)
    interf_block = extrair_interferencias(int_dir, tmp_root)
    interf_paths = None
    if interf_block:
        interf_paths = interf_block.pop("_paths", None)
        print("[extrair] Interferencias: agua=%d drenagem=%d" %
              (interf_block["agua_trechos"], interf_block["drenagem_trechos"]))

    # ---- 5) Modelo hidraulico SewerGEMS (.sqlite/.stsw) ----
    modelo_block = None
    modelo_path = _get(cfg, "modelo", "modelo_sqlite", "modelo_hidraulico", "sqlite")
    if modelo_path and os.path.exists(modelo_path):
        print("[extrair] Modelo hidraulico:", modelo_path)
        modelo_block = extrair_modelo_sqlite(
            modelo_path, _get(cfg, "modelo_cenario", "scenario_id"))
        if modelo_block:
            print("[extrair] Modelo: %d trechos, %s m, DN %s, recobr min %s, decl min %s" % (
                modelo_block["n_trechos"], modelo_block["extensao_total_m"],
                modelo_block["dn_lista"], modelo_block["recobrimento_min_m"],
                modelo_block["declividade_min_m_m"]))

    # ---- escreve geo ----
    _escrever_geo(geo_dir, O["points"], O["trechos"], sol_rows, sol_shp_base, epsg)
    # copia os shapes de interferencia classificados p/ nome canonico (estavel)
    interf_shp_dir = _copiar_interferencias(geo_dir, interf_paths)
    if interf_shp_dir:
        print("[extrair] Interferencias copiadas (nomes canonicos) ->", interf_shp_dir)

    # ========================================================================
    #  DADOS_JSON (quantitativos)
    # ========================================================================
    total_row = O["total_row"] or {}
    q = dict(
        n_oses=len(O["ose_list"]),
        extensao_total_m=round(total_row.get("ext"), 2) if total_row.get("ext") else
        round(sum([o["ext"] for o in O["ose_list"] if o["ext"]]), 2),
        extensao_total_conferencia_somatorio=round(sum([o["ext"] for o in O["ose_list"] if o["ext"]]), 2),
        n_PV=total_row.get("pv") if total_row.get("pv") is not None else n_pv_pts,
        n_TL=total_row.get("tl") if total_row.get("tl") is not None else n_tl_pts,
        n_PIT=total_row.get("pit") if total_row.get("pit") is not None else 0,
        n_TQ=len(O["tq_list"]),
        n_DEGRAU=len(O["deg_list"]),
        extensao_por_DN_m={k: round(v, 2) for k, v in O["ext_by_dn"].items()},
        profundidade_por_faixa_estruturas=O["faixas"],
        metodo_extensao=dict(
            regra="OSE/trecho com profundidade max > 3,00 m classificado como VCA",
            MND_m=round(O["ext_mnd"], 2), VCA_m=round(O["ext_vca"], 2),
            n_OSE_VCA=O["n_ose_vca"],
            VCA_planilhas_m=round(O.get("ext_vca_sheets", 0.0), 2),
            ext_total_sheets_m=round(O.get("ext_total_sheets", 0.0), 2)))

    sub_bacias = sorted(set([str(m["bacia"]).strip() for m in O["ose_meta"] if m["bacia"]]))
    data = dict(
        identificacao=dict(
            municipio=("%s%s" % (municipio, " - %s" % uf if uf else "")),
            bacia=bacia or (sub_bacias[0] if sub_bacias else "—"),
            sub_bacias=sub_bacias,
            concessionaria=_get(cfg, "concessionaria") or proj.get("concessionaria") or "ACCIONA",
            contratada=_get(cfg, "contratada") or proj.get("contratada") or "2S Engenharia e Geotecnologia",
            sistema="Rede Coletora de Esgoto (RCE)",
            metodo_executivo=dict(distribuicao=dict(O["method_set"])),
            DN=dict(distribuicao_por_ose=dict(O["DN_set"])),
            pavimentos=dict(O["pav_set"]),
            codigo_ibge=str(codigo_ibge) if codigo_ibge else None,
            datum="SIRGAS 2000 / UTM 22S",
            arquivo_fonte=O["arquivo_fonte"]),
        quantitativos=q,
        parametros_vazoes=dict(
            encontrado=False,
            observacao=("A planilha de OSEs contem apenas geometria/cotas/quantitativos da "
                        "rede. NAO ha parametros populacionais, per capita, K1/K2/K3, coef. "
                        "retorno, taxa de infiltracao, vazoes, horizonte de projeto nem indice "
                        "de atendimento populacional nesta fonte.")),
        geo_arquivos=dict(
            pvs_tls_csv="geo/estruturas_pv_tl.csv",
            rede_trechos_geojson="geo/rede_trechos.geojson",
            rede_trechos_csv="geo/rede_trechos.csv",
            soleiras_csv="geo/soleiras.csv",
            soleiras_shapefile="geo/soleiras_shp/"))
    if sol_block:
        data["soleiras"] = sol_block
    if interf_block:
        data["interferencias"] = interf_block
    if topo_block:
        data["topografia"] = topo_block

    dados_json_path = os.path.join(out_dir, "dados.json")
    with open(dados_json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # ========================================================================
    #  DADOS_EXTRA_JSON (TQ/degrau, DN, GNSS)
    # ========================================================================
    def _alt_block(lst):
        alturas = [d["altura_m"] for d in lst]
        return dict(
            quantidade=len(lst),
            soma_alturas_m=round(sum(alturas), 3) if alturas else 0.0,
            altura_min_m=round(min(alturas), 3) if alturas else None,
            altura_max_m=round(max(alturas), 3) if alturas else None,
            lista=lst)

    dn_extensao = []
    # extensao total de referencia (RESUMO!TOTAL ou somatorio das OSEs)
    ext_total_ref = (round(total_row.get("ext"), 2) if total_row.get("ext")
                     else round(sum([o["ext"] for o in O["ose_list"] if o["ext"]]), 2))
    _dn_items = list(O["ext_by_dn"].items())
    for dn_label, ext in _dn_items:
        # 'DN 150mm - PVC' -> dn='DN 150', material='PVC'
        material = "PVC"
        if " - " in dn_label:
            base, material = dn_label.split(" - ", 1)
        else:
            base = dn_label
        dn = base.replace("mm", "").strip()
        # com DN unico, usa a extensao TOTAL de referencia (RESUMO) p/ casar com
        # o quantitativo geral; com varios DN, usa a soma por DN das abas.
        ext_dn = ext_total_ref if len(_dn_items) == 1 else round(ext, 2)
        dn_extensao.append(dict(dn=dn, material=material.strip(),
                                extensao_m=round(ext_dn, 2),
                                n_oses=O["DN_set"].get(dn_label, 0)))

    extra = dict(
        municipio=municipio,
        bacia=bacia,
        data_extracao=__import__("datetime").date.today().isoformat(),
    )
    extra["1_tubos_de_queda_e_degrau"] = dict(
        fonte=O["arquivo_fonte"],
        onde_achado=("cada aba OSE-NNN, celula U12 (rotulo 'T.Q. X,XXX m' ou "
                     "'DEGRAU X,XXX m')"),
        tubos_de_queda_TQ=_alt_block(O["tq_list"]),
        degraus=_alt_block(O["deg_list"]),
        observacao=("O tubo de queda aparece abreviado 'T.Q.' no rotulo da celula U12 "
                    "de cada aba OSE; 'DEGRAU  m' (sem valor) e linha-modelo vazia e e "
                    "ignorado."))
    extra["2_dn_rede"] = dict(
        fonte="OSE: aba!B9 (DN) e RESUMO col B (Extensao)",
        dn_unico=(dn_extensao[0]["dn"] + " - " + dn_extensao[0]["material"]
                  if len(dn_extensao) == 1 else None),
        dn_extensao=dn_extensao,
        n_linhas_tabela_quantitativo_rede=len(dn_extensao),
        observacao=("Tabela de quantitativo de rede com %d linha(s) de DN." % len(dn_extensao)))
    if gnss_block is not None:
        extra["3_precisoes_gnss"] = gnss_block
    if modelo_block is not None:
        extra["4_modelo_hidraulico"] = modelo_block

    dados_extra_path = os.path.join(out_dir, "dados_extra.json")
    with open(dados_extra_path, "w", encoding="utf-8") as f:
        json.dump(extra, f, ensure_ascii=False, indent=2)

    # ---- limpeza do tmp ----
    try:
        shutil.rmtree(tmp_root, ignore_errors=True)
    except Exception:
        pass

    resumo = dict(
        ok=True,
        dados_json=dados_json_path,
        dados_extra_json=dados_extra_path,
        rede_geojson=os.path.join(geo_dir, "rede_trechos.geojson"),
        geo_dir=geo_dir,
        interferencias_shp_dir=interf_shp_dir,
        n_oses=q["n_oses"], extensao_total_m=q["extensao_total_m"],
        n_PV=q["n_PV"], n_TL=q["n_TL"], n_PIT=q["n_PIT"],
        n_TQ=q["n_TQ"], n_DEGRAU=q["n_DEGRAU"],
        MND_m=q["metodo_extensao"]["MND_m"], VCA_m=q["metodo_extensao"]["VCA_m"],
        DN=list(O["ext_by_dn"].keys()),
        soleiras=(sol_block and dict(total=sol_block["total_imoveis"],
                                     pos=sol_block["atendidas_positiva"],
                                     neg=sol_block["nao_atendidas_negativa"],
                                     idx=sol_block["indice_atendimento_pct"])),
        topo=(topo_block and dict(pontos=topo_block["total_pontos"],
                                  z_min=topo_block.get("z_min"),
                                  z_max=topo_block.get("z_max"))),
        interferencias=interf_block)
    return resumo


def _carregar_cfg(arg):
    if isinstance(arg, dict):
        return arg
    with open(arg, "r", encoding="utf-8") as f:
        return json.load(f)


def main():
    ap = argparse.ArgumentParser(description="Extracao parametrica de dados do RCE (Memorial)")
    ap.add_argument("--config", required=True, help="JSON com caminhos dos brutos + projeto")
    ap.add_argument("--out", help="Diretorio de saida (sobrescreve out_dir do config)")
    args = ap.parse_args()
    cfg = _carregar_cfg(args.config)
    if args.out:
        cfg["out_dir"] = args.out
    try:
        resumo = extrair(cfg)
    except Exception as e:
        import traceback
        sys.stderr.write(traceback.format_exc())
        print(json.dumps({"ok": False, "erro": str(e)}, ensure_ascii=False))
        return 1
    print(json.dumps(resumo, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
