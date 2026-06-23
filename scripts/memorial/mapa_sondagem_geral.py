# -*- coding: utf-8 -*-
"""Mapa GERAL das sondagens — visão de conjunto do Lote 02 Acciona.

Plota TODOS os furos de sondagem (todos os municípios) sobre imagem de
satélite, coloridos por município, com legenda lateral + contagem, no padrão
cartográfico 2S. Diferente do mapa7 (por município, com nº do furo), aqui é o
panorama do projeto inteiro.

Fonte: RELATORIO_SONDAGEM_*.xlsx (Survey123 "Sondagem - Acciona").
"""
import os, math
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D
import cartolib as C
import mapa7_sondagem as M7   # reusa COL_* e _norm

OUT_PNG = os.path.join(C.OUT, "Mapa_Sondagem_Geral.png")


def carregar_todos(xlsx):
    """Le o XLSX e devolve todos os furos com coordenada: {num,munic,lon,lat,agua}."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sondagem"] if "Sondagem" in wb.sheetnames else wb[wb.sheetnames[0]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        lon, lat = r[M7.COL_LON], r[M7.COL_LAT]
        if lon in (None, "") or lat in (None, ""):
            continue
        try:
            lon = float(str(lon).replace(",", ".")); lat = float(str(lat).replace(",", "."))
        except ValueError:
            continue
        # sanidade: descarta coordenada fora do Paraná (zerada/trocada/erro)
        if not (-55.0 < lon < -48.0 and -27.0 < lat < -22.0):
            continue
        mun = str(r[M7.COL_MUN] or "").strip()
        out.append(dict(num=str(r[M7.COL_NUM] or "").strip(), munic=mun, lon=lon, lat=lat))
    return out


def main():
    furos = carregar_todos(M7.XLSX)
    if not furos:
        print("[geral] nenhum furo."); return None
    from pyproj import Transformer
    tr = Transformer.from_crs("EPSG:4326", "EPSG:31982", always_xy=True)
    for f in furos:
        f["x"], f["y"] = tr.transform(f["lon"], f["lat"])

    # agrupa por municipio (ordena por contagem desc)
    from collections import defaultdict
    grp = defaultdict(list)
    for f in furos:
        grp[f["munic"] or "(s/ município)"].append(f)
    munics = sorted(grp, key=lambda m: -len(grp[m]))
    n_tot = len(furos)
    print(f"[geral] {n_tot} furos | {len(munics)} municípios")

    # paleta categorica (até 20)
    cmap = plt.get_cmap("tab20")
    cores = {m: cmap(i % 20) for i, m in enumerate(munics)}

    xs = np.array([f["x"] for f in furos]); ys = np.array([f["y"] for f in furos])
    xmin, xmax = xs.min(), xs.max(); ymin, ymax = ys.min(), ys.max()
    span = max(xmax - xmin, ymax - ymin)
    mx = span * 0.06
    cx, cy = (xmin + xmax) / 2, (ymin + ymax) / 2
    half = span / 2 + mx
    xmin, xmax = cx - half, cx + half
    ymin, ymax = cy - half, cy + half

    fig = plt.figure(figsize=(11.69, 8.27), dpi=300)
    # mapa à esquerda (deixa faixa direita p/ legenda dos municípios)
    MAP_TOP = C.MAP_TOP
    ax = fig.add_axes([0.045, 0.075, 0.66, MAP_TOP - 0.075])
    ax.set_xlim(xmin, xmax); ax.set_ylim(ymin, ymax)
    ax.set_aspect("equal", adjustable="box")

    src = C.add_basemap(ax, provider="imagery")

    full = xmax - xmin
    step = C._nice_step(full)
    while full / step > 6: step *= 2
    while full / step < 3: step /= 2
    C.utm_grid(ax, xmin, xmax, ymin, ymax, step=step, grid_alpha=0.45)
    C.frame(ax)

    # furos coloridos por municipio
    for m in munics:
        gx = [f["x"] for f in grp[m]]; gy = [f["y"] for f in grp[m]]
        ax.plot(gx, gy, ls="none", marker="o", ms=3.0, mfc=cores[m],
                mec=C.INK, mew=0.4, zorder=18, clip_on=True)

    C.north_arrow(ax)
    C.scale_bar(ax, xmin, xmax, ymin, ymax, frac=0.30, loc="lower right", y_frac=0.035)
    C.title_block(fig, "MAPA GERAL DAS SONDAGENS (SPT)",
                  f"Investigação Geotécnica · Lote 02 Acciona · {n_tot} furos · {len(munics)} municípios",
                  project="REDE COLETORA DE ESGOTO · LOTE 02 — PR")

    # ---- legenda lateral (municipio + contagem) ----
    lrect = [0.725, 0.075, 0.235, MAP_TOP - 0.075]
    handles = [Line2D([0], [0], marker="o", ls="none", mfc=cores[m], mec=C.INK,
                      mew=0.4, ms=7,
                      label=f"{(m[:22]+'…') if len(m) > 23 else m}  ({len(grp[m])})")
               for m in munics]
    C.legend_panel(fig, lrect, handles, title="MUNICÍPIOS",
                   leg_fs=C.FS_LEG - 0.8, labelspacing=0.9)
    C.datum_box(ax, corner="lower left")
    C.credits(fig, fonte=(src or "Esri World Imagery"))

    os.makedirs(C.OUT, exist_ok=True)
    fig.savefig(OUT_PNG, dpi=300)
    plt.close(fig)
    print("[geral] salvo:", OUT_PNG)
    return OUT_PNG


if __name__ == "__main__":
    main()
