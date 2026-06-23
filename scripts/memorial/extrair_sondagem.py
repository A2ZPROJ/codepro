# -*- coding: utf-8 -*-
"""Extração e síntese da investigação geotécnica (sondagens) por município.

Lê o RELATORIO_SONDAGEM_*.xlsx (Survey123 "Sondagem - Acciona") e devolve um
dicionário com os parâmetros para a seção de Sondagem do Memorial Descritivo:
contagem, método, profundidades, perfil de solo, nível d'água e (quando houver)
classificação SPT. Usado por gerar_memorial_descritivo (tokens {{SOND_*}}).
"""
import os, unicodedata, statistics as st

_DEF_XLSX = (r"C:\Users\lcabd\OneDrive - 2S ENGENHARIA DE AGRIMENSURA E GEOTECNOLOGIA"
             r"\001. SERVIDOR PARANÁ\002. ACCIONA\004. CT-027.2025 - PROJETOS"
             r"\000. CRONOGRAMAS\DASHBOARD ONLINE\RELATORIO_SONDAGEM_2026-05-29.xlsx")

# colunas (0-based) do XLSX Survey123
C_NUM, C_MUN, C_INTERV, C_PROF, C_AGUA, C_PROFAGUA, C_SPT = 6, 3, 10, 12, 50, 51, 52
# camadas (col_tipo, col_cor, rótulo de profundidade)
LAYERS = [(13, 15, "0,00–0,50 m"), (17, 19, "0,50–1,00 m"), (21, 23, "1,00–1,50 m"),
          (25, 27, "1,50–2,00 m"), (29, 31, "2,00–2,50 m"), (33, 35, "2,50–3,00 m"),
          (37, 39, "3,00–3,50 m"), (41, 43, "3,50–4,00 m")]
GOLPES_MIN, GOLPES_MAX = 53, 635   # faixa de colunas golpes_spt_*


def _norm(s):
    return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().strip().upper()


def _br(v, casas=2):
    return ("{:.%df}" % casas).format(v).replace(".", ",")


def _limpa(v):
    """'1 - Arenoso' -> 'Arenoso'; 'other'/'' -> ''."""
    s = str(v or "").strip()
    if not s or s.lower() == "other":
        return ""
    if " - " in s:
        s = s.split(" - ", 1)[1].strip()
    return s


def _fnum(v):
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return None


def resumo_sondagem(municipio, xlsx=None):
    """Devolve dict com os parâmetros da sondagem do município (ou None se não houver)."""
    import openpyxl
    xlsx = xlsx or os.environ.get("MEMORIAL_SONDAGEM_XLSX", _DEF_XLSX)
    if not os.path.exists(xlsx):
        return None
    alvo = _norm(municipio)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb["Sondagem"] if "Sondagem" in wb.sheetnames else wb[wb.sheetnames[0]]

    furos, profs, solos, cores = [], [], [], []
    golpes_all, n_spt_furos, n_agua, profs_agua = [], 0, 0, []
    interv = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        mun = _norm(r[C_MUN])
        if not (mun == alvo or mun.startswith(alvo) or (alvo and alvo in mun)):
            continue
        prof = _fnum(r[C_PROF])
        # camadas (tipo/cor) preenchidas neste furo
        camadas = []
        for ct, cc, faixa in LAYERS:
            ti = _limpa(r[ct]) if ct < len(r) else ""
            co = _limpa(r[cc]) if cc < len(r) else ""
            if ti:
                camadas.append((faixa, ti, co))
                solos.append(ti)
                if co:
                    cores.append(co)
        ag = _norm(r[C_AGUA]).startswith("SIM")
        if ag:
            n_agua += 1
            pa = _fnum(r[C_PROFAGUA])
            if pa is not None:
                profs_agua.append(pa)
        # golpes SPT deste furo
        gf = [_fnum(r[i]) for i in range(GOLPES_MIN, min(GOLPES_MAX, len(r)))]
        gf = [g for g in gf if g is not None and g >= 0]
        tem_spt_furo = bool(gf) or _norm(r[C_SPT]).startswith("SIM")
        if gf:
            golpes_all += gf
            n_spt_furos += 1
        it = _norm(r[C_INTERV])
        interv[it] = interv.get(it, 0) + 1
        if prof is not None:
            profs.append(prof)
        furos.append(dict(num=str(r[C_NUM] or "").strip(), prof=prof,
                          solo=(camadas[0][1] if camadas else ""),
                          cor=(camadas[0][2] if camadas else ""),
                          agua=ag, spt=tem_spt_furo))
    if not furos:
        return None

    from collections import Counter
    solo_pred = Counter(solos).most_common(1)[0][0] if solos else "—"
    cor_pred = Counter(cores).most_common(1)[0][0] if cores else "—"
    tem_spt = n_spt_furos > 0
    tem_na = n_agua > 0

    # método de investigação
    if tem_spt:
        metodo = "sondagem a trado, poços de inspeção e sondagem à percussão com SPT"
    else:
        metodo = "sondagem a trado / poços de inspeção"

    # classificação SPT (aprox.: média dos golpes registrados)
    nspt_med = round(st.mean(golpes_all)) if golpes_all else None
    compacidade = "—"
    if nspt_med is not None:
        aren = solo_pred.lower().startswith("aren")
        if aren:
            compacidade = ("fofa" if nspt_med <= 4 else "pouco compacta" if nspt_med <= 8
                           else "medianamente compacta" if nspt_med <= 18
                           else "compacta" if nspt_med <= 40 else "muito compacta")
        else:
            compacidade = ("muito mole" if nspt_med <= 2 else "mole" if nspt_med <= 5
                           else "média" if nspt_med <= 10 else "rija" if nspt_med <= 19 else "dura")

    # nível d'água
    if tem_na:
        na_sit = "Detectado em %d furo(s)" % n_agua
        na_prof_med = _br(st.mean(profs_agua), 2) if profs_agua else "—"
    else:
        na_sit = "Não detectado até a profundidade investigada"
        na_prof_med = "—"

    pmin, pmax = min(profs), max(profs)
    pmed = st.mean(profs)

    return dict(
        MUNICIPIO=municipio, N_FUROS=len(furos), METODO_INVEST=metodo,
        PROF_MIN=_br(pmin), PROF_MED=_br(pmed), PROF_MAX=_br(pmax),
        SOLO_PRED=solo_pred, COR_PRED=cor_pred,
        SPT_SIM_NAO="Sim" if tem_spt else "Não",
        NSPT_MED=(str(nspt_med) if nspt_med is not None else "N/A"),
        COMPACIDADE_PRED=compacidade,
        NA_SITUACAO=na_sit, N_FUROS_NA=n_agua,
        NA_PROF_MED=na_prof_med, DATUM="SIRGAS 2000",
        tem_spt=tem_spt, tem_na=tem_na,
        n_spt_furos=n_spt_furos, furos=furos,
        prof_uniforme=(abs(pmax - pmin) < 0.05),
    )


if __name__ == "__main__":
    import sys, json
    m = sys.argv[1] if len(sys.argv) > 1 else "Amaporã"
    d = resumo_sondagem(m)
    if not d:
        print("Sem sondagem para", m)
    else:
        d2 = {k: v for k, v in d.items() if k != "furos"}
        print(json.dumps(d2, ensure_ascii=False, indent=2))
        print("furos:", len(d["furos"]), "| ex:", d["furos"][0])
